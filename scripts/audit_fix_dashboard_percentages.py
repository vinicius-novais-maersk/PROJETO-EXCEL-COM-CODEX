from __future__ import annotations

import json
import os
import shutil
import subprocess
import sys
import tempfile
import time
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import pythoncom
import win32com.client as win32
import win32process
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


ALLIANCE = "Alian" + chr(231) + "a"
PORT_ROWS = [2, 3, 6, 7, 8, 11, 12, 13, 16, 17, 18, 19, 20]
REGION_ROWS = [4, 9, 14, 21]
ALL_ROWS = PORT_ROWS + REGION_ROWS + [23]

XL_CALC_MANUAL = -4135
XL_CALC_AUTOMATIC = -4105
XL_OPENXML_WORKBOOK_MACRO_ENABLED = 52


def contiguous_blocks(rows: list[int]) -> list[tuple[int, int]]:
    if not rows:
        return []
    rows = sorted(rows)
    blocks: list[tuple[int, int]] = []
    start = end = rows[0]
    for row in rows[1:]:
        if row == end + 1:
            end = row
            continue
        blocks.append((start, end))
        start = end = row
    blocks.append((start, end))
    return blocks


def kill_excel_process(pid: int | None) -> None:
    if not pid:
        return
    subprocess.run(
        ["taskkill", "/PID", str(pid), "/F"],
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        check=False,
    )


def open_workbook(excel, path: Path):
    try:
        return excel.Workbooks.Open(str(path), UpdateLinks=0, ReadOnly=False, IgnoreReadOnlyRecommended=True)
    except TypeError:
        return excel.Workbooks.Open(str(path), 0, False, 5, "", "", True)
    except Exception:
        return excel.Workbooks.Open(str(path), 0, False, 5, "", "", True, None, None, False, False, None, False, True, 1)


def set_formula(ws, cell_ref: str, formula: str) -> None:
    ws.Range(cell_ref).Formula = formula


def set_value(ws, cell_ref: str, value) -> None:
    ws.Range(cell_ref).Value = value


def build_countifs(*parts: str) -> str:
    return "COUNTIFS(" + ",".join(parts) + ")"


def criterion(range_ref: str, crit: str) -> str:
    return f"{range_ref},{crit}"


def sheet_range(sheet_name: str, address: str) -> str:
    return f"'{sheet_name}'!{address}"


def q(value: str) -> str:
    return f'"{value}"'


def build_week_denominator(scope_column: str | None = None, scope_ref: str | None = None, alliance_filter: str | None = None) -> str:
    parts = [criterion(sheet_range("ROE_wk", "$AV:$AV"), q("Ok"))]
    if scope_column and scope_ref:
        parts.append(criterion(sheet_range("ROE_wk", f"${scope_column}:${scope_column}"), scope_ref))
    parts.append(criterion(sheet_range("ROE_wk", "$AR:$AR"), "$AG$1"))
    if alliance_filter is not None:
        parts.append(criterion(sheet_range("ROE_wk", "$AI:$AI"), q(alliance_filter)))
    parts.extend(
        [
            criterion(sheet_range("ROE_wk", "$BB:$BB"), q("<>Sem Preenchimento")),
            criterion(sheet_range("ROE_wk", "$BL:$BL"), q("N")),
            criterion(sheet_range("ROE_wk", "$BO:$BO"), q("<>Especial")),
        ]
    )
    return build_countifs(*parts)


def build_week_late(scope_column: str | None = None, scope_ref: str | None = None, alliance_filter: str | None = None) -> str:
    parts = [criterion(sheet_range("ROE_wk", "$AV:$AV"), q("Ok"))]
    if scope_column and scope_ref:
        parts.append(criterion(sheet_range("ROE_wk", f"${scope_column}:${scope_column}"), scope_ref))
    parts.append(criterion(sheet_range("ROE_wk", "$AR:$AR"), "$AG$1"))
    if alliance_filter is not None:
        parts.append(criterion(sheet_range("ROE_wk", "$AI:$AI"), q(alliance_filter)))
    parts.extend(
        [
            criterion(sheet_range("ROE_wk", "$BB:$BB"), q("Atrasado")),
            criterion(sheet_range("ROE_wk", "$BL:$BL"), q("N")),
            criterion(sheet_range("ROE_wk", "$BO:$BO"), q("<>Especial")),
        ]
    )
    return build_countifs(*parts)


def build_week_ratio_formula(scope_column: str | None = None, scope_ref: str | None = None, alliance_filter: str | None = None) -> str:
    denominator = build_week_denominator(scope_column, scope_ref, alliance_filter)
    late = build_week_late(scope_column, scope_ref, alliance_filter)
    return f'=IF({denominator}=0,"",1-IFERROR({late}/{denominator},0))'


def build_port_denominator(alliance_filter: str | None, day_ref: str | None = None, include_special: bool = False) -> str:
    parts = [
        criterion(sheet_range("ROE_wk", "$AV:$AV"), q("Ok")),
        criterion(sheet_range("ROE_wk", "$AY:$AY"), "$M$3"),
        criterion(sheet_range("ROE_wk", "$AR:$AR"), "$M$2"),
    ]
    if day_ref is not None:
        parts.append(criterion(sheet_range("ROE_wk", "$AP:$AP"), day_ref))
    if alliance_filter is not None:
        parts.append(criterion(sheet_range("ROE_wk", "$AI:$AI"), q(alliance_filter)))
    parts.extend(
        [
            criterion(sheet_range("ROE_wk", "$BB:$BB"), q("<>Sem Preenchimento")),
            criterion(sheet_range("ROE_wk", "$BL:$BL"), q("N")),
        ]
    )
    if not include_special:
        parts.append(criterion(sheet_range("ROE_wk", "$BO:$BO"), q("<>Especial")))
    return build_countifs(*parts)


def build_port_late(alliance_filter: str | None, day_ref: str | None = None, include_special: bool = False) -> str:
    parts = [
        criterion(sheet_range("ROE_wk", "$AV:$AV"), q("Ok")),
        criterion(sheet_range("ROE_wk", "$AY:$AY"), "$M$3"),
        criterion(sheet_range("ROE_wk", "$AR:$AR"), "$M$2"),
    ]
    if day_ref is not None:
        parts.append(criterion(sheet_range("ROE_wk", "$AP:$AP"), day_ref))
    if alliance_filter is not None:
        parts.append(criterion(sheet_range("ROE_wk", "$AI:$AI"), q(alliance_filter)))
    parts.extend(
        [
            criterion(sheet_range("ROE_wk", "$BB:$BB"), q("Atrasado")),
            criterion(sheet_range("ROE_wk", "$BL:$BL"), q("N")),
        ]
    )
    if not include_special:
        parts.append(criterion(sheet_range("ROE_wk", "$BO:$BO"), q("<>Especial")))
    return build_countifs(*parts)


def build_port_ratio_formula(alliance_filter: str | None, day_ref: str | None = None, include_special: bool = False) -> str:
    denominator = build_port_denominator(alliance_filter, day_ref, include_special=include_special)
    late = build_port_late(alliance_filter, day_ref, include_special=include_special)
    return f'=IF({denominator}=0,"",1-IFERROR({late}/{denominator},0))'


def build_48h_formula(denominator_ref: str, alliance_filter: str | None, day_ref: str | None = None, include_special: bool = True) -> str:
    parts = [
        criterion(sheet_range("ROE_wk", "$AV:$AV"), q("Ok")),
        criterion(sheet_range("ROE_wk", "$AY:$AY"), "$M$3"),
        criterion(sheet_range("ROE_wk", "$AR:$AR"), "$M$2"),
    ]
    if day_ref is not None:
        parts.append(criterion(sheet_range("ROE_wk", "$AP:$AP"), day_ref))
    if alliance_filter is not None:
        parts.append(criterion(sheet_range("ROE_wk", "$AI:$AI"), q(alliance_filter)))
    if not include_special:
        parts.append(criterion(sheet_range("ROE_wk", "$BO:$BO"), q("<>Especial")))
    parts.append(criterion(sheet_range("ROE_wk", "$AT:$AT"), q("N")))
    numerator = build_countifs(*parts)
    return f'=IF({denominator_ref}=0,"",1-IFERROR({numerator}/{denominator_ref},0))'


def build_port_lookup_formula(result_column: str) -> str:
    return f'=IFERROR(XLOOKUP($M$3,Week_Overview!AG:AG,Week_Overview!{result_column}:{result_column},""),"")'


def build_hist_lookup_formula(week_cell: str, result_column: str) -> str:
    return (
        f'=IFERROR(XLOOKUP($M$3,INDIRECT("Week_"&${week_cell}&"!AG:AG"),'
        f'INDIRECT("Week_"&${week_cell}&"!{result_column}:{result_column}"),""),"")'
    )


def apply_week_overview(ws) -> None:
    for row in PORT_ROWS:
        set_formula(ws, f"N{row}", build_week_ratio_formula("AY", f"$AG{row}", ALLIANCE))
        set_formula(ws, f"P{row}", build_week_ratio_formula("AY", f"$AG{row}", f"<>{ALLIANCE}"))
        set_formula(ws, f"Q{row}", build_week_ratio_formula("AY", f"$AG{row}", None))
    for row in REGION_ROWS:
        set_formula(ws, f"N{row}", build_week_ratio_formula("AZ", f"$A{row}", ALLIANCE))
        set_formula(ws, f"P{row}", build_week_ratio_formula("AZ", f"$A{row}", f"<>{ALLIANCE}"))
        set_formula(ws, f"Q{row}", build_week_ratio_formula("AZ", f"$A{row}", None))
    set_formula(ws, "N23", build_week_ratio_formula(None, None, ALLIANCE))
    set_formula(ws, "P23", build_week_ratio_formula(None, None, f"<>{ALLIANCE}"))
    set_formula(ws, "Q23", build_week_ratio_formula())

    # 48h Scheduled CAB aligned to CAB volume only.
    for row in PORT_ROWS:
        set_formula(
            ws,
            f"R{row}",
            (
                f'=IF(D{row}=0,"",1-IFERROR(COUNTIFS('
                "'ROE_wk'!$AV:$AV,\"Ok\","
                f"'ROE_wk'!$AY:$AY,$AG{row},"
                "'ROE_wk'!$AR:$AR,$AG$1,"
                f"'ROE_wk'!$AI:$AI,\"{ALLIANCE}\","
                "'ROE_wk'!$AT:$AT,\"N\")/D"
                f"{row},0))"
            ),
        )
    for row in REGION_ROWS:
        set_formula(
            ws,
            f"R{row}",
            (
                f'=IF(D{row}=0,"",1-IFERROR(COUNTIFS('
                "'ROE_wk'!$AV:$AV,\"Ok\","
                f"'ROE_wk'!$AZ:$AZ,$A{row},"
                "'ROE_wk'!$AR:$AR,$AG$1,"
                f"'ROE_wk'!$AI:$AI,\"{ALLIANCE}\","
                "'ROE_wk'!$AT:$AT,\"N\")/D"
                f"{row},0))"
            ),
        )
    set_formula(
        ws,
        "R23",
        (
            '=IF(D23=0,"",1-IFERROR(COUNTIFS('
            "'ROE_wk'!$AV:$AV,\"Ok\","
            "'ROE_wk'!$AR:$AR,$AG$1,"
            f"'ROE_wk'!$AI:$AI,\"{ALLIANCE}\","
            "'ROE_wk'!$AT:$AT,\"N\")/D23,0))"
        ),
    )

    # Rescheduled OP and active-week drivers.
    for row in PORT_ROWS:
        set_formula(
            ws,
            f"S{row}",
            (
                f'=SUMIFS(\'Reagendas\'!Q:Q,\'Reagendas\'!$F:$F,Week_Overview!AG{row},'
                f'\'Reagendas\'!$R:$R,$AI$1,\'Reagendas\'!$K:$K,"Reagenda")'
            ),
        )
        set_formula(
            ws,
            f"W{row}",
            (
                '=COUNTIFS(Validacao_Final!O:O,"<>NÃ£o",'
                'Validacao_Final!A:A,"Landside",'
                'Validacao_Final!J:J,"Reagendamento",'
                f'Validacao_Final!I:I,$AG{row},'
                'Validacao_Final!M:M,$AG$1)'
            ),
        )
        set_formula(
            ws,
            f"X{row}",
            (
                '=COUNTIFS(Validacao_Final!O:O,"<>NÃ£o",'
                'Validacao_Final!A:A,"Landside",'
                'Validacao_Final!J:J,"No-Show",'
                f'Validacao_Final!I:I,$AG{row},'
                'Validacao_Final!M:M,$AG$1)'
                '+COUNTIFS(Reagendas_2!BD:BD,"Negativo",'
                'Reagendas_2!AX:AX,"Landside",'
                f'Reagendas_2!AV:AV,$AG{row},'
                'Reagendas_2!AY:AY,$AG$1)'
            ),
        )
        set_formula(
            ws,
            f"Y{row}",
            (
                '=COUNTIFS(Reagendas_2!AX:AX,"CX",'
                f'Reagendas_2!AV:AV,$AG{row},'
                'Reagendas_2!AY:AY,$AG$1,'
                'Reagendas_2!AP:AP,"Agendamento - Falha de CX - AlianÃ§a")'
                '+COUNTIFS(Reagendas_2!AX:AX,"CX",'
                f'Reagendas_2!AV:AV,$AG{row},'
                'Reagendas_2!AY:AY,$AG$1,'
                'Reagendas_2!AP:AP,"Agendamento - Erro de Doc - AlianÃ§a")'
            ),
        )
        set_formula(
            ws,
            f"Z{row}",
            (
                '=COUNTIFS(Reagendas_2!AX:AX,"Outros",'
                f'Reagendas_2!AV:AV,$AG{row},'
                'Reagendas_2!AY:AY,$AG$1,'
                'Reagendas_2!AP:AP,"Agendamento - Falta de janela - AlianÃ§a")'
                '+COUNTIFS(Reagendas_2!AX:AX,"Outros",'
                f'Reagendas_2!AV:AV,$AG{row},'
                'Reagendas_2!AY:AY,$AG$1,'
                'Reagendas_2!AP:AP,"Erro Sistema â€“ AlianÃ§a")'
                '+COUNTIFS(Reagendas_2!AX:AX,"Outros",'
                f'Reagendas_2!AV:AV,$AG{row},'
                'Reagendas_2!AY:AY,$AG$1,'
                'Reagendas_2!AP:AP,"Agendamento - Indisp de equip - AlianÃ§a")'
                '+COUNTIFS(Reagendas_2!AX:AX,"Outros",'
                f'Reagendas_2!AV:AV,$AG{row},'
                'Reagendas_2!AY:AY,$AG$1,'
                'Reagendas_2!AP:AP,"Agendamento - Atraso do navio - AlianÃ§a")'
                '+COUNTIFS(Reagendas_2!AX:AX,"Outros",'
                f'Reagendas_2!AV:AV,$AG{row},'
                'Reagendas_2!AY:AY,$AG$1,'
                'Reagendas_2!AP:AP,"Agendamento - Falha PortuÃ¡ria - AlianÃ§a")'
            ),
        )
        set_formula(
            ws,
            f"AA{row}",
            (
                '=COUNTIFS(Reagendas_2!AX:AX,"Outros",'
                f'Reagendas_2!AV:AV,$AG{row},'
                'Reagendas_2!AY:AY,$AG$1,'
                'Reagendas_2!AP:AP,"Agendamento - Indisp. VeÃ­culo - AlianÃ§a")'
            ),
        )
        set_formula(
            ws,
            f"AB{row}",
            (
                '=COUNTIFS(Reagendas_2!AX:AX,"Cliente",'
                f'Reagendas_2!AV:AV,$AG{row},'
                'Reagendas_2!AY:AY,$AG$1,'
                'Reagendas_2!AP:AP,"Agendamento - Erro de Doc - Cliente")'
                '+COUNTIFS(Reagendas_2!AX:AX,"Cliente",'
                f'Reagendas_2!AV:AV,$AG{row},'
                'Reagendas_2!AY:AY,$AG$1,'
                'Reagendas_2!AP:AP,"Agendamento - Reagendamento - Cliente")'
            ),
        )
        set_formula(
            ws,
            f"AC{row}",
            (
                '=COUNTIFS(Reagendas_2!AX:AX,"Landside",'
                f'Reagendas_2!AV:AV,$AG{row},'
                'Reagendas_2!AY:AY,$AG$1,'
                'Reagendas_2!BC:BC,"Negativo")'
            ),
        )


def apply_volume_ds(ws) -> None:
    set_value(ws, "E18", "Rescheduled")
    for col in "FGHIJKL":
        set_formula(ws, f"{col}16", build_port_ratio_formula(f"<>{ALLIANCE}", day_ref=f"{col}$8"))
        set_formula(ws, f"{col}17", build_48h_formula(f"{col}12", f"<>{ALLIANCE}", day_ref=f"{col}$8"))
        set_formula(
            ws,
            f"{col}18",
            (
                f'=SUMIFS(\'Reagendas\'!$Q:$Q,\'Reagendas\'!$F:$F,$M$3,'
                f'\'Reagendas\'!$R:$R,$M$2,\'Reagendas\'!$O:$O,{col}$8,'
                '\'Reagendas\'!$K:$K,"Reagenda")'
            ),
        )
    set_formula(ws, "M16", build_port_ratio_formula(f"<>{ALLIANCE}"))
    set_formula(ws, "M17", build_48h_formula("M12", f"<>{ALLIANCE}"))
    set_formula(ws, "M18", "=SUM(F18:L18)")
    for row in [25, 26]:
        set_formula(ws, f"J{row}", build_hist_lookup_formula(f"E{row}", "N"))
        set_formula(ws, f"N{row}", build_hist_lookup_formula(f"E{row}", "P"))
        set_formula(ws, f"R{row}", build_hist_lookup_formula(f"E{row}", "Q"))
    set_formula(ws, "J27", build_port_lookup_formula("N"))
    set_formula(ws, "N27", build_port_lookup_formula("P"))
    set_formula(ws, "R27", build_port_lookup_formula("Q"))


def apply_volume_graph(ws) -> None:
    set_value(ws, "E19", "Rescheduled")
    for col in "FGHIJKL":
        set_formula(ws, f"{col}17", build_port_ratio_formula(ALLIANCE, day_ref=f"{col}$8"))
        set_formula(ws, f"{col}18", build_48h_formula(f"{col}12", ALLIANCE, day_ref=f"{col}$8"))
        set_formula(
            ws,
            f"{col}19",
            (
                '=COUNTIFS('
                'Validacao_Final!$A:$A,"Landside",'
                'Validacao_Final!$I:$I,$M$3,'
                'Validacao_Final!$M:$M,$M$2,'
                f'Validacao_Final!$Q:$Q,{col}$8,'
                'Validacao_Final!$J:$J,"Reagendamento",'
                'Validacao_Final!$O:$O,"<>NÃ£o")'
            ),
        )
    set_formula(ws, "M17", build_port_ratio_formula(ALLIANCE))
    set_formula(ws, "M18", build_48h_formula("M12", ALLIANCE))
    set_formula(ws, "M19", "=SUM(F19:L19)")
    for row in [25, 26]:
        set_formula(ws, f"J{row}", build_hist_lookup_formula(f"E{row}", "N"))
        set_formula(ws, f"N{row}", build_hist_lookup_formula(f"E{row}", "P"))
        set_formula(ws, f"R{row}", build_hist_lookup_formula(f"E{row}", "Q"))
    set_formula(ws, "J27", build_port_lookup_formula("N"))
    set_formula(ws, "N27", build_port_lookup_formula("P"))
    set_formula(ws, "R27", build_port_lookup_formula("Q"))


def apply_volume_mao(ws) -> None:
    set_value(ws, "E21", "Rescheduled")
    for col in "FGHIJKL":
        set_formula(ws, f"{col}19", build_port_ratio_formula(ALLIANCE, day_ref=f"{col}$8"))
        set_formula(ws, f"{col}20", build_48h_formula(f"{col}12+{col}13", ALLIANCE, day_ref=f"{col}$8"))
        set_formula(
            ws,
            f"{col}21",
            (
                '=COUNTIFS('
                'Validacao_Final!$A:$A,"Landside",'
                'Validacao_Final!$I:$I,$M$3,'
                'Validacao_Final!$M:$M,$M$2,'
                f'Validacao_Final!$Q:$Q,{col}$8,'
                'Validacao_Final!$J:$J,"Reagendamento",'
                'Validacao_Final!$O:$O,"<>NÃ£o")'
            ),
        )
    set_formula(ws, "M19", build_port_ratio_formula(ALLIANCE))
    set_formula(ws, "M20", build_48h_formula("M12+M13", ALLIANCE))
    set_formula(ws, "M21", "=SUM(F21:L21)")
    for row in [27, 28]:
        set_formula(ws, f"J{row}", build_hist_lookup_formula(f"E{row}", "N"))
        set_formula(ws, f"N{row}", build_hist_lookup_formula(f"E{row}", "P"))
        set_formula(ws, f"R{row}", build_hist_lookup_formula(f"E{row}", "Q"))
    set_formula(ws, "J29", build_port_lookup_formula("N"))
    set_formula(ws, "N29", build_port_lookup_formula("P"))
    set_formula(ws, "R29", build_port_lookup_formula("Q"))


@dataclass
class RoeRow:
    week: int | None
    day_week: str | None
    alliance: str | None
    sla_flag: str | None
    volume_flag: str | None
    port: str | None
    region: str | None
    otd_status: str | None
    oto_out: str | None
    special: str | None


def read_roe_rows(workbook_path: Path) -> list[RoeRow]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True, keep_vba=True)
    ws = wb["ROE_wk"]
    idx = {letter: column_index_from_string(letter) - 1 for letter in ["AI", "AP", "AR", "AT", "AV", "AY", "AZ", "BB", "BL", "BO"]}
    rows: list[RoeRow] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        rows.append(
            RoeRow(
                week=values[idx["AR"]],
                day_week=values[idx["AP"]],
                alliance=values[idx["AI"]],
                sla_flag=values[idx["AT"]],
                volume_flag=values[idx["AV"]],
                port=values[idx["AY"]],
                region=values[idx["AZ"]],
                otd_status=values[idx["BB"]],
                oto_out=values[idx["BL"]],
                special=values[idx["BO"]],
            )
        )
    return rows


def count_rows(
    rows: list[RoeRow],
    *,
    week,
    port: str | None = None,
    region: str | None = None,
    alliance: str | None = None,
    exclude_special: bool = False,
    exclude_sem_preench: bool = False,
    require_oto_out_n: bool = False,
    late_only: bool = False,
    day_week: str | None = None,
    sla_n_only: bool = False,
) -> int:
    total = 0
    for row in rows:
        if row.volume_flag != "Ok":
            continue
        if row.week != week:
            continue
        if port is not None and row.port != port:
            continue
        if region is not None and row.region != region:
            continue
        if alliance == ALLIANCE and row.alliance != ALLIANCE:
            continue
        if alliance == f"<>{ALLIANCE}" and row.alliance == ALLIANCE:
            continue
        if exclude_special and row.special == "Especial":
            continue
        if exclude_sem_preench and row.otd_status == "Sem Preenchimento":
            continue
        if require_oto_out_n and row.oto_out != "N":
            continue
        if late_only and row.otd_status != "Atrasado":
            continue
        if day_week is not None and row.day_week != day_week:
            continue
        if sla_n_only and row.sla_flag != "N":
            continue
        total += 1
    return total


def ratio(blank_denominator: int, numerator: int) -> float | None:
    if blank_denominator == 0:
        return None
    return 1 - (numerator / blank_denominator)


def audit_after_fix(workbook_path: Path) -> dict[str, object]:
    rows = read_roe_rows(workbook_path)
    wb = load_workbook(workbook_path, read_only=True, data_only=True, keep_vba=True)
    week_ws = wb["Week_Overview"]
    week = week_ws["AG1"].value
    volume_ds = wb["Volume_DS"]
    volume_graph = wb["Volume_Graph"]
    volume_mao = wb["Volume_MAO"]

    week_checks: dict[str, dict[str, float | None]] = {}

    row_definitions = {
        2: ("port", "Manaus"),
        3: ("port", "Vila do Conde"),
        4: ("region", "North"),
        6: ("port", "Pecem"),
        7: ("port", "Suape"),
        8: ("port", "Salvador"),
        9: ("region", "Northeast"),
        11: ("port", "Vitoria"),
        12: ("port", "Rio"),
        13: ("port", "Santos"),
        14: ("region", "Southeast"),
        16: ("port", "Itapoa"),
        17: ("port", "Itajai"),
        18: ("port", "Paranagua"),
        19: ("port", "Imbituba"),
        20: ("port", "Rio Grande"),
        21: ("region", "South"),
        23: ("total", None),
    }

    for row_num, (scope_type, scope_value) in row_definitions.items():
        kwargs = {"week": week}
        if scope_type == "port":
            kwargs["port"] = scope_value
        elif scope_type == "region":
            kwargs["region"] = scope_value

        cab_den = count_rows(rows, alliance=ALLIANCE, exclude_special=True, exclude_sem_preench=True, require_oto_out_n=True, **kwargs)
        cab_late = count_rows(rows, alliance=ALLIANCE, exclude_special=True, require_oto_out_n=True, late_only=True, **kwargs)
        ds_den = count_rows(rows, alliance=f"<>{ALLIANCE}", exclude_special=True, exclude_sem_preench=True, require_oto_out_n=True, **kwargs)
        ds_late = count_rows(rows, alliance=f"<>{ALLIANCE}", exclude_special=True, require_oto_out_n=True, late_only=True, **kwargs)
        ttl_den = count_rows(rows, exclude_special=True, exclude_sem_preench=True, require_oto_out_n=True, **kwargs)
        ttl_late = count_rows(rows, exclude_special=True, require_oto_out_n=True, late_only=True, **kwargs)
        cab_48_den = count_rows(rows, alliance=ALLIANCE, **kwargs)
        cab_48_n = count_rows(rows, alliance=ALLIANCE, sla_n_only=True, **kwargs)

        week_checks[f"row_{row_num}"] = {
            "expected_N": ratio(cab_den, cab_late),
            "sheet_N": week_ws[f"N{row_num}"].value,
            "expected_P": ratio(ds_den, ds_late),
            "sheet_P": week_ws[f"P{row_num}"].value,
            "expected_Q": ratio(ttl_den, ttl_late),
            "sheet_Q": week_ws[f"Q{row_num}"].value,
            "expected_R": ratio(cab_48_den, cab_48_n),
            "sheet_R": week_ws[f"R{row_num}"].value,
            "ttl_den": ttl_den,
            "ttl_late": ttl_late,
        }

    port = volume_ds["M3"].value
    graph_checks: dict[str, dict[str, float | None]] = {}

    for col in "FGHIJKL":
        day = volume_ds[f"{col}8"].value
        ds_den = count_rows(rows, week=week, port=port, alliance=f"<>{ALLIANCE}", exclude_special=True, exclude_sem_preench=True, require_oto_out_n=True, day_week=day)
        ds_late = count_rows(rows, week=week, port=port, alliance=f"<>{ALLIANCE}", exclude_special=True, require_oto_out_n=True, late_only=True, day_week=day)
        ds_48_den = count_rows(rows, week=week, port=port, alliance=f"<>{ALLIANCE}", day_week=day)
        ds_48_n = count_rows(rows, week=week, port=port, alliance=f"<>{ALLIANCE}", day_week=day, sla_n_only=True)
        cab_den = count_rows(rows, week=week, port=port, alliance=ALLIANCE, exclude_special=True, exclude_sem_preench=True, require_oto_out_n=True, day_week=day)
        cab_late = count_rows(rows, week=week, port=port, alliance=ALLIANCE, exclude_special=True, require_oto_out_n=True, late_only=True, day_week=day)
        cab_48_den = count_rows(rows, week=week, port=port, alliance=ALLIANCE, day_week=day)
        cab_48_n = count_rows(rows, week=week, port=port, alliance=ALLIANCE, day_week=day, sla_n_only=True)

        graph_checks[f"day_{day}"] = {
            "expected_ds_oto": ratio(ds_den, ds_late),
            "sheet_ds_oto": volume_ds[f"{col}16"].value,
            "expected_ds_48h": ratio(ds_48_den, ds_48_n),
            "sheet_ds_48h": volume_ds[f"{col}17"].value,
            "expected_cab_oto": ratio(cab_den, cab_late),
            "sheet_cab_oto": volume_graph[f"{col}17"].value,
            "sheet_mao_oto": volume_mao[f"{col}19"].value,
            "expected_cab_48h": ratio(cab_48_den, cab_48_n),
            "sheet_cab_48h": volume_graph[f"{col}18"].value,
            "sheet_mao_48h": volume_mao[f"{col}20"].value,
        }

    ds_den_total = count_rows(rows, week=week, port=port, alliance=f"<>{ALLIANCE}", exclude_special=True, exclude_sem_preench=True, require_oto_out_n=True)
    ds_late_total = count_rows(rows, week=week, port=port, alliance=f"<>{ALLIANCE}", exclude_special=True, require_oto_out_n=True, late_only=True)
    ds_48_den_total = count_rows(rows, week=week, port=port, alliance=f"<>{ALLIANCE}")
    ds_48_n_total = count_rows(rows, week=week, port=port, alliance=f"<>{ALLIANCE}", sla_n_only=True)
    cab_den_total = count_rows(rows, week=week, port=port, alliance=ALLIANCE, exclude_special=True, exclude_sem_preench=True, require_oto_out_n=True)
    cab_late_total = count_rows(rows, week=week, port=port, alliance=ALLIANCE, exclude_special=True, require_oto_out_n=True, late_only=True)
    cab_48_den_total = count_rows(rows, week=week, port=port, alliance=ALLIANCE)
    cab_48_n_total = count_rows(rows, week=week, port=port, alliance=ALLIANCE, sla_n_only=True)

    totals = {
        "port": port,
        "week": week,
        "ds_oto_expected": ratio(ds_den_total, ds_late_total),
        "ds_oto_sheet": volume_ds["M16"].value,
        "ds_48_expected": ratio(ds_48_den_total, ds_48_n_total),
        "ds_48_sheet": volume_ds["M17"].value,
        "cab_oto_expected": ratio(cab_den_total, cab_late_total),
        "cab_oto_graph_sheet": volume_graph["M17"].value,
        "cab_oto_mao_sheet": volume_mao["M19"].value,
        "cab_48_expected": ratio(cab_48_den_total, cab_48_n_total),
        "cab_48_graph_sheet": volume_graph["M18"].value,
        "cab_48_mao_sheet": volume_mao["M20"].value,
        "selected_port_weekoverview_Q": week_ws["Q18"].value if port == "Paranagua" else None,
    }

    # 100% legitimacy report.
    true_100 = []
    for row_num, values in week_checks.items():
        for key in ["expected_N", "expected_P", "expected_Q", "expected_R"]:
            val = values[key]
            if val == 1:
                true_100.append({"scope": row_num, "metric": key, "denominator": values["ttl_den"], "late": values["ttl_late"]})

    return {
        "week_overview": week_checks,
        "graphs": graph_checks,
        "totals": totals,
        "true_100_examples": true_100[:20],
    }


def main() -> int:
    if len(sys.argv) != 2:
        print("Usage: python scripts/audit_fix_dashboard_percentages.py <workbook_path>", file=sys.stderr)
        return 2

    workbook_path = Path(sys.argv[1]).resolve()
    if not workbook_path.exists():
        print(f"Workbook not found: {workbook_path}", file=sys.stderr)
        return 2

    backup_dir = Path(__file__).resolve().parents[1] / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = backup_dir / f"{workbook_path.stem}_backup_{timestamp}{workbook_path.suffix}"
    shutil.copy2(workbook_path, backup_path)

    temp_dir = Path(tempfile.mkdtemp(prefix="codex_dashboard_fix_"))
    temp_path = temp_dir / workbook_path.name

    pythoncom.CoInitialize()
    excel = None
    workbook = None
    excel_pid: int | None = None
    try:
        excel = win32.DispatchEx("Excel.Application")
        try:
            _, excel_pid = win32process.GetWindowThreadProcessId(excel.Hwnd)
        except Exception:
            excel_pid = None
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.ScreenUpdating = False
        excel.EnableEvents = False
        excel.AskToUpdateLinks = False

        workbook = open_workbook(excel, workbook_path)
        try:
            excel.Calculation = XL_CALC_MANUAL
        except Exception:
            pass

        apply_week_overview(workbook.Worksheets("Week_Overview"))
        apply_volume_ds(workbook.Worksheets("Volume_DS"))
        apply_volume_graph(workbook.Worksheets("Volume_Graph"))
        apply_volume_mao(workbook.Worksheets("Volume_MAO"))

        try:
            excel.Calculation = XL_CALC_AUTOMATIC
        except Exception:
            pass
        excel.CalculateFullRebuild()
        workbook.SaveCopyAs(str(temp_path))
    finally:
        if workbook is not None:
            try:
                workbook.Close(SaveChanges=False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()

    if excel_pid:
        for _ in range(5):
            try:
                os.kill(excel_pid, 0)
                time.sleep(1)
            except OSError:
                break
        else:
            kill_excel_process(excel_pid)

    for _ in range(20):
        try:
            os.replace(temp_path, workbook_path)
            break
        except PermissionError:
            time.sleep(2)
    else:
        raise PermissionError(f"Could not replace workbook: {workbook_path}")

    audit = audit_after_fix(workbook_path)
    audit["backup"] = str(backup_path)
    print(json.dumps(audit, ensure_ascii=True))
    shutil.rmtree(temp_dir, ignore_errors=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
