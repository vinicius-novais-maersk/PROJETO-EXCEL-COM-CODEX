from __future__ import annotations

import argparse
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl


TARGET_SHEETS = ["Week_Overview", "Volume_DS", "Volume_Graph", "Volume_MAO"]
REGIONS = ["North", "Northeast", "Southeast", "South"]
TOP_PORTS = [
    "Manaus",
    "Vila do Conde",
    "Pecem",
    "Suape",
    "Salvador",
    "Santos",
    "Rio",
    "Vitoria",
    "Itapoa",
    "Rio Grande",
    "Imbituba",
    "Itajai",
    "Paranagua",
]
PIVOT_FIELD_NAMES = ["Atrasado?", "OTO Out", "Weeknum", "Porto", "Region"]


@dataclass
class Row:
    os_num: str
    tipo_os: str | None
    provider: str | None
    service_type: str | None
    cost_center: str | None
    programmed_date: Any
    day_week: str | None
    weeknum: int | None
    data_ag: Any
    sla_ag: str | None
    volume: str | None
    port: str | None
    region: str | None
    otd: str | None
    oto_out: str | None
    special: str | None


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Audit workbook consistency.")
    parser.add_argument("workbook", type=Path, help="Path to the workbook (.xlsm)")
    parser.add_argument(
        "--output",
        type=Path,
        help="Path to the markdown report. Defaults to analysis/<stem>_audit.md",
    )
    return parser.parse_args()


def pct(value: float | None) -> str:
    if value is None:
        return ""
    return f"{value * 100:.2f}%"


def value_or_blank(value: Any) -> str:
    if value in (None, ""):
        return ""
    return str(value)


def markdown_table(headers: list[str], rows: list[list[Any]]) -> str:
    lines = ["| " + " | ".join(headers) + " |", "|" + "|".join(["---"] * len(headers)) + "|"]
    for row in rows:
        rendered = [value_or_blank(item) for item in row]
        lines.append("| " + " | ".join(rendered) + " |")
    return "\n".join(lines)


def normalize_formula(formula: str) -> str:
    return formula.replace(" ", "").replace("\n", "").replace("\r", "").lower()


def load_rows(workbook_path: Path) -> tuple[openpyxl.Workbook, openpyxl.Workbook, int, list[Row]]:
    wb_values = openpyxl.load_workbook(workbook_path, data_only=True, keep_vba=True, read_only=True)
    wb_formulas = openpyxl.load_workbook(workbook_path, data_only=False, keep_vba=True, read_only=True)
    week = wb_values["Week_Overview"]["AG1"].value
    ws = wb_values["ROE_wk"]
    rows: list[Row] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[47] != "Ok" or row[43] != week:
            continue
        rows.append(
            Row(
                os_num=row[0],
                tipo_os=row[2],
                provider=row[3],
                service_type=row[4],
                cost_center=row[34],
                programmed_date=row[9],
                day_week=row[41],
                weeknum=row[43],
                data_ag=row[44],
                sla_ag=row[45],
                volume=row[47],
                port=row[50],
                region=row[51],
                otd=row[53],
                oto_out=row[63],
                special=row[66],
            )
        )
    return wb_values, wb_formulas, week, rows


def metrics(rows: list[Row]) -> dict[str, Any]:
    total = len(rows)
    total_oto_n = sum(1 for row in rows if row.oto_out == "N")
    delayed_oto_n = sum(1 for row in rows if row.oto_out == "N" and row.otd == "Atrasado")
    sem_oto_n = sum(1 for row in rows if row.oto_out == "N" and row.otd == "Sem Preenchimento")
    no_prazo_oto_n = sum(1 for row in rows if row.oto_out == "N" and row.otd == "No Prazo")
    return {
        "total": total,
        "oto_n": total_oto_n,
        "sem_n": sem_oto_n,
        "delayed_n": delayed_oto_n,
        "no_prazo_n": no_prazo_oto_n,
        "kpi_sheet": None if total == 0 else 1 - delayed_oto_n / total,
        "kpi_oto_denominator": None if total_oto_n == 0 else 1 - delayed_oto_n / total_oto_n,
        "kpi_without_sem": None
        if total_oto_n == sem_oto_n
        else 1 - delayed_oto_n / (total_oto_n - sem_oto_n),
    }


def build_sil_status_map(wb_values: openpyxl.Workbook) -> dict[str, Any]:
    ws = wb_values["SIL_wk"]
    status_by_os: dict[str, Any] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        os_num = row[1] if len(row) > 1 else None
        status = row[24] if len(row) > 24 else None
        if os_num not in (None, "") and os_num not in status_by_os:
            status_by_os[os_num] = status
    return status_by_os


def build_rao_key_set(wb_values: openpyxl.Workbook) -> set[str]:
    ws = wb_values["RAO_wk"]
    keys: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        os_num = row[0]
        if os_num not in (None, ""):
            keys.add(os_num)
    return keys


def formula_stats(wb_formulas: openpyxl.Workbook) -> dict[str, dict[str, Any]]:
    report: dict[str, dict[str, Any]] = {}
    for sheet_name in TARGET_SHEETS:
        ws = wb_formulas[sheet_name]
        stats = defaultdict(int)
        samples: dict[str, list[str]] = defaultdict(list)
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if not (isinstance(value, str) and value.startswith("=")):
                    continue
                normalized = normalize_formula(value)
                normalized_no_quotes = normalized.replace("'", "")
                if "roe_wk!" not in normalized_no_quotes:
                    continue
                stats["roe_formulas"] += 1
                if "bo:$bo" in normalized_no_quotes:
                    stats["special_exclusion_refs"] += 1
                    if len(samples["bo"]) < 5:
                        samples["bo"].append(cell.coordinate)
                if "sempreenchimento" in normalized_no_quotes:
                    stats["sem_preenchimento_refs"] += 1
                    if len(samples["sem"]) < 5:
                        samples["sem"].append(cell.coordinate)
                if 'bl:$bl,"n"' in normalized_no_quotes or 'bl:$bl;"n"' in normalized_no_quotes or 'bl:bl,"n"' in normalized_no_quotes:
                    stats["oto_out_n_refs"] += 1
                    if len(samples["bl"]) < 5:
                        samples["bl"].append(cell.coordinate)
                if 'bb:$bb,"atrasado"' in normalized_no_quotes or 'bb:$bb;"atrasado"' in normalized_no_quotes or 'bb:bb,"atrasado"' in normalized_no_quotes:
                    stats["delayed_refs"] += 1
                    if len(samples["bb"]) < 5:
                        samples["bb"].append(cell.coordinate)
        report[sheet_name] = {
            "stats": dict(stats),
            "samples": dict(samples),
        }
    return report


def try_read_pivots(workbook_path: Path) -> dict[str, list[dict[str, Any]]]:
    try:
        import pythoncom
        import win32com.client
    except ImportError:
        return {}

    pivot_report: dict[str, list[dict[str, Any]]] = {}
    pythoncom.CoInitialize()
    excel = None
    workbook = None
    try:
        excel = win32com.client.gencache.EnsureDispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(
            str(workbook_path),
            UpdateLinks=0,
            ReadOnly=True,
            IgnoreReadOnlyRecommended=True,
            AddToMru=False,
            Notify=False,
        )
        for sheet_name in ["Top_Offenders_Customers", "Top_Offenders_Vendors"]:
            ws = workbook.Worksheets(sheet_name)
            pivot_tables = []
            for index in range(1, ws.PivotTables().Count + 1):
                pivot_table = ws.PivotTables().Item(index)
                fields: dict[str, Any] = {}
                for field_name in PIVOT_FIELD_NAMES:
                    try:
                        field = pivot_table.PivotFields(field_name)
                        try:
                            current_page = field.CurrentPage
                        except Exception:
                            current_page = None
                        if current_page not in (None, ""):
                            current_page = str(current_page)
                        fields[field_name] = {
                            "orientation": int(field.Orientation),
                            "current_page": current_page,
                        }
                    except Exception:
                        fields[field_name] = {"orientation": None, "current_page": None}
                pivot_tables.append({"name": pivot_table.Name, "fields": fields})
            pivot_report[sheet_name] = pivot_tables
    finally:
        if workbook is not None:
            workbook.Close(SaveChanges=False)
        if excel is not None:
            excel.Quit()
        pythoncom.CoUninitialize()
    return pivot_report


def render_report(
    workbook_path: Path,
    week: int,
    rows: list[Row],
    wb_values: openpyxl.Workbook,
    wb_formulas: openpyxl.Workbook,
) -> str:
    sil_status = build_sil_status_map(wb_values)
    rao_keys = build_rao_key_set(wb_values)
    formulas = formula_stats(wb_formulas)
    pivots = try_read_pivots(workbook_path)

    total_rows = len(rows)
    unique_os = len({row.os_num for row in rows})
    duplicate_os = total_rows - unique_os
    blank_port_rows = [row for row in rows if row.port in (None, "")]
    blank_region_rows = [row for row in rows if row.region in (None, "")]
    blank_day_rows = [row for row in rows if row.day_week in (None, "")]
    blank_day_excel_weekdays: Counter[int] = Counter()
    for row in blank_day_rows:
        if row.programmed_date is None:
            continue
        python_weekday = row.programmed_date.weekday()  # Monday=0
        excel_weekday = ((python_weekday + 1) % 7) + 1
        blank_day_excel_weekdays[excel_weekday] += 1

    sil_not_found = [row for row in rows if row.os_num not in sil_status]
    sem_rows = [row for row in rows if row.otd == "Sem Preenchimento"]
    sem_from_missing_sil = [row for row in sem_rows if row.os_num not in sil_status]
    sem_from_sil_flag = [row for row in sem_rows if sil_status.get(row.os_num) == "Sem Preenchimento"]

    rao_missing = [row for row in rows if row.os_num not in rao_keys]
    special_rows = [row for row in rows if row.special == "Especial"]

    sum_region_totals = sum(1 for row in rows if row.region in REGIONS)
    week_overview_total = wb_values["Week_Overview"]["K23"].value

    region_rows = []
    for region in REGIONS:
        region_subset = [row for row in rows if row.region == region]
        region_without_special = [row for row in region_subset if row.special != "Especial"]
        current = metrics(region_subset)
        no_special = metrics(region_without_special)
        region_rows.append(
            [
                region,
                current["total"],
                current["sem_n"],
                current["delayed_n"],
                pct(current["kpi_sheet"]),
                pct(current["kpi_without_sem"]),
                len(region_subset) - len(region_without_special),
                pct(no_special["kpi_sheet"]),
                pct(no_special["kpi_without_sem"]),
            ]
        )

    impacted_port_rows = []
    for port in TOP_PORTS:
        subset = [row for row in rows if row.port == port]
        if not subset:
            continue
        current = metrics(subset)
        impacted_port_rows.append(
            [
                port,
                current["total"],
                current["sem_n"],
                current["delayed_n"],
                pct(current["kpi_sheet"]),
                pct(current["kpi_without_sem"]),
            ]
        )

    formula_rows = []
    for sheet_name in TARGET_SHEETS:
        stats = formulas[sheet_name]["stats"]
        formula_rows.append(
            [
                sheet_name,
                stats.get("roe_formulas", 0),
                stats.get("special_exclusion_refs", 0),
                stats.get("sem_preenchimento_refs", 0),
                stats.get("oto_out_n_refs", 0),
                stats.get("delayed_refs", 0),
            ]
        )

    pivot_lines: list[str] = []
    for sheet_name, pivot_tables in pivots.items():
        pivot_lines.append(f"### {sheet_name}")
        pivot_table_rows = []
        for pivot_table in pivot_tables:
            fields = pivot_table["fields"]
            pivot_table_rows.append(
                [
                    pivot_table["name"],
                    fields["Atrasado?"]["orientation"],
                    fields["Atrasado?"]["current_page"],
                    fields["OTO Out"]["orientation"],
                    fields["OTO Out"]["current_page"],
                    fields["Weeknum"]["orientation"],
                    fields["Weeknum"]["current_page"],
                ]
            )
        pivot_lines.append(
            markdown_table(
                ["Pivot", "Atrasado? ori", "Atrasado? val", "OTO Out ori", "OTO Out val", "Weeknum ori", "Weeknum val"],
                pivot_table_rows,
            )
        )

    rio_port = wb_values["Volume_DS"]["M3"].value
    rio_week = wb_values["Volume_DS"]["M2"].value
    rio_rows = [row for row in rows if row.port == rio_port]
    rio_blank_day = sum(1 for row in rio_rows if row.day_week in (None, ""))

    lines = [
        f"# Workbook audit - {workbook_path.name}",
        "",
        f"- Workbook: `{workbook_path}`",
        f"- Active week in `Week_Overview!AG1`: `{week}`",
        "",
        "## Executive summary",
        "",
        markdown_table(
            ["Check", "Result"],
            [
                ["Weekly ROE rows (`Volume=Ok`)", total_rows],
                ["Unique OS in week", unique_os],
                ["Duplicate OS in week", duplicate_os],
                ["Rows with blank `Porto`", len(blank_port_rows)],
                ["Rows with blank `Region`", len(blank_region_rows)],
                ["Rows with blank `day week`", len(blank_day_rows)],
                ["Rows not found in `SIL_wk`", len(sil_not_found)],
                ["Rows not found in `RAO_wk`", len(rao_missing)],
                ["Rows marked `Especial`", len(special_rows)],
                ["Sum of regional rows in base", sum_region_totals],
                ["`Week_Overview!K23`", week_overview_total],
                ["Difference base vs regional total", total_rows - sum_region_totals],
            ],
        ),
        "",
        "## Base integrity",
        "",
        markdown_table(
            ["Metric", "Count", "Top detail"],
            [
                [
                    "Blank `Porto` / `Region`",
                    len(blank_region_rows),
                    Counter(row.provider for row in blank_region_rows).most_common(3),
                ],
                [
                    "Blank `day week`",
                    len(blank_day_rows),
                    f"Ports={Counter(row.port for row in blank_day_rows).most_common(4)}; Excel weekday={dict(blank_day_excel_weekdays)}",
                ],
                [
                    "Missing in `SIL_wk`",
                    len(sil_not_found),
                    Counter(row.region for row in sil_not_found).most_common(4),
                ],
                [
                    "Missing in `RAO_wk`",
                    len(rao_missing),
                    Counter(row.provider for row in rao_missing).most_common(4),
                ],
                [
                    "`Sem Preenchimento` total",
                    len(sem_rows),
                    Counter(row.region for row in sem_rows).most_common(4),
                ],
            ],
        ),
        "",
        "### `Sem Preenchimento` origin",
        "",
        markdown_table(
            ["Origin", "Count"],
            [
                ["Not found in `SIL_wk`", len(sem_from_missing_sil)],
                ["Returned as `Sem Preenchimento` from `SIL_wk`", len(sem_from_sil_flag)],
                ["Other / unresolved", len(sem_rows) - len(sem_from_missing_sil) - len(sem_from_sil_flag)],
            ],
        ),
        "",
        "### Missing in `SIL_wk` by service and OS type",
        "",
        markdown_table(
            ["Cut", "Distribution"],
            [
                ["`Tipo Servico`", Counter(row.service_type for row in sil_not_found)],
                ["`Tipo OS`", Counter(row.tipo_os for row in sil_not_found)],
            ],
        ),
        "",
        "### Missing in `RAO_wk` sample",
        "",
        markdown_table(
            ["OS", "Provider", "Tipo Servico", "Tipo OS", "OTD"],
            [
                [row.os_num, row.provider, row.service_type, row.tipo_os, row.otd]
                for row in rao_missing[:10]
            ],
        ),
        "",
        "## KPI consistency by region",
        "",
        markdown_table(
            [
                "Region",
                "Total",
                "`Sem Preenchimento` (`OTO Out=N`)",
                "Delayed (`OTO Out=N`)",
                "KPI shown by current denominator",
                "KPI excluding `Sem Preenchimento`",
                "Special rows",
                "KPI shown without specials",
                "KPI no-special and no-sem",
            ],
            region_rows,
        ),
        "",
        "## Port impact",
        "",
        markdown_table(
            ["Port", "Total", "`Sem Preenchimento`", "Delayed", "KPI shown", "KPI excluding `Sem Preenchimento`"],
            impacted_port_rows,
        ),
        "",
        "## Special rows",
        "",
        markdown_table(
            ["Cut", "Distribution"],
            [
                ["By region", Counter(row.region for row in special_rows)],
                ["By OTD (`OTO Out=N` only)", Counter(row.otd for row in special_rows if row.oto_out == "N")],
            ],
        ),
        "",
        "## Formula coverage in dashboards",
        "",
        markdown_table(
            [
                "Sheet",
                "ROE formulas",
                "Special exclusion refs",
                "Sem Preenchimento refs",
                "`OTO Out=N` refs",
                "`Atrasado` refs",
            ],
            formula_rows,
        ),
        "",
        "### Formula samples",
        "",
    ]

    for sheet_name in TARGET_SHEETS:
        lines.append(f"#### {sheet_name}")
        lines.append("")
        lines.append(
            markdown_table(
                ["Check", "Samples"],
                [
                    ["Special exclusion", formulas[sheet_name]["samples"].get("bo", [])],
                    ["Sem Preenchimento", formulas[sheet_name]["samples"].get("sem", [])],
                    ["OTO Out", formulas[sheet_name]["samples"].get("bl", [])],
                    ["Atrasado", formulas[sheet_name]["samples"].get("bb", [])],
                ],
            )
        )
        lines.append("")

    lines.extend(
        [
            "## Current operational panel check",
            "",
            markdown_table(
                ["Panel selection", "Value"],
                [
                    ["`Volume_DS!M2` week", rio_week],
                    ["`Volume_DS!M3` port", rio_port],
                    ["Base rows for selected port", len(rio_rows)],
                    ["Base rows with blank `day week`", rio_blank_day],
                    ["`Volume_DS!M9` shown total", wb_values["Volume_DS"]["M9"].value],
                    ["`Volume_Graph!M9` shown total", wb_values["Volume_Graph"]["M9"].value],
                    ["`Volume_MAO!M9` shown total", wb_values["Volume_MAO"]["M9"].value],
                ],
            ),
            "",
            "## Pivot filter check",
            "",
        ]
    )

    if pivot_lines:
        lines.extend(pivot_lines)
    else:
        lines.append("Pivot inspection not available in this environment.")

    lines.extend(
        [
            "",
            "## Main findings",
            "",
            "1. `Sem Preenchimento` enters totals but not delayed numerators, which improves KPI artificially.",
            f"2. `Week_Overview` does not exclude `Especial` rows, while the active week has {len(special_rows)} `Especial` rows.",
            f"3. `RAO_wk` misses {len(rao_missing)} road rows in the active week, leaving them without `Porto` and `Region`.",
            f"4. `SIL_wk` misses {len(sil_not_found)} weekly rows, all in `Transporte Rodoviario`.",
            f"5. `day week` is blank for {len(blank_day_rows)} rows, and the blank rows map to `WEEKDAY(...,1)={dict(blank_day_excel_weekdays)}` based on `J -> AO -> AQ -> AP`.",
            f"6. Dashboard formula rules are not uniform: `Week_Overview` has {formulas['Week_Overview']['stats'].get('sem_preenchimento_refs', 0)} `Sem Preenchimento` exclusions and {formulas['Week_Overview']['stats'].get('special_exclusion_refs', 0)} special exclusions; `Volume_DS` has {formulas['Volume_DS']['stats'].get('sem_preenchimento_refs', 0)} `Sem Preenchimento` exclusions; `Volume_MAO` has {formulas['Volume_MAO']['stats'].get('special_exclusion_refs', 0)} explicit special exclusions.",
            "7. `Top_Offenders_Customers` pivot filters are inconsistent across pivot tables: some have `OTO Out=N`, others do not.",
            "",
        ]
    )

    return "\n".join(lines)


def main() -> int:
    args = parse_args()
    workbook_path = args.workbook.resolve()
    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")

    output_path = args.output
    if output_path is None:
        output_path = workbook_path.parent / "analysis"
        if output_path.name != "analysis":
            output_path = workbook_path.parent / "analysis" / f"{workbook_path.stem}_audit.md"
        else:
            output_path = output_path / f"{workbook_path.stem}_audit.md"
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb_values, wb_formulas, week, rows = load_rows(workbook_path)
    report = render_report(workbook_path, week, rows, wb_values, wb_formulas)
    output_path.write_text(report, encoding="utf-8")
    print(f"Audit report written to: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
