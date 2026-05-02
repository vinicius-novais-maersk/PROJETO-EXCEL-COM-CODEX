from __future__ import annotations

import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl
import pythoncom
import win32com.client
from openpyxl.utils import get_column_letter

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from scripts.update_special_clients import WORKBOOK_PATH, backup_workbook, call_with_retry


ALLIANCE = "Alian" + chr(231) + "a"
RODO = "Rodo"
OK = "Ok"
ATRASADO = "Atrasado"
SEM_PREENCHIMENTO = "Sem Preenchimento"
ESPECIAL = "Especial"
ON_TIME = "On time"
REAGENDA = "Reagenda"
SUN = "Sun"

MONTH_PORT_ROWS = [2, 3, 6, 7, 8, 11, 12, 13, 16, 17, 18, 19]
MONTH_REGION_GROUPS = {
    4: [2, 3],
    9: [6, 7, 8],
    14: [11, 12, 13],
    20: [16, 17, 18, 19],
}
WEEKLY_BLOCK_STARTS = [2, 11, 20, 29, 38]
WEEKLY_LABEL_COLUMNS = [2, 12, 22]
WEEKLY_PORT_GROUP_WIDTH = 9

PORT_TO_REGION = {
    "Manaus": "North",
    "Vila do Conde": "North",
    "Pecem": "Northeast",
    "Suape": "Northeast",
    "Salvador": "Northeast",
    "Santos": "Southeast",
    "Rio": "Southeast",
    "Vitoria": "Southeast",
    "Itapoa": "South",
    "Imbituba": "South",
    "Rio Grande": "South",
    "Itajai": "South",
    "Paranagua": "South",
    "Navegantes": "South",
}


def log(message: str) -> None:
    print(f"[full-sweep] {message}", flush=True)


def q(value: str) -> str:
    return f'"{value}"'


def sheet_range(sheet_name: str, address: str) -> str:
    return f"'{sheet_name}'!{address}"


def criterion(range_ref: str, value: str) -> str:
    return f"{range_ref};{value}"


def countifs(*parts: str) -> str:
    return "CONT.SES(" + ";".join(parts) + ")"


def sumifs(sum_range: str, *parts: str) -> str:
    return "SOMASES(" + ";".join([sum_range, *parts]) + ")"


def sum_exprs(expressions: list[str]) -> str:
    if not expressions:
        return "0"
    if len(expressions) == 1:
        return expressions[0]
    return "(" + ")+(".join(expressions) + ")"


def ratio_formula(port_ref: str, denominator: str, late: str) -> str:
    return f'=SE({port_ref}="";"";SE({denominator}=0;"";1-SEERRO({late}/{denominator};0)))'


def count_formula(port_ref: str, denominator: str, late: str) -> str:
    return f'=SE({port_ref}="";"";SE({denominator}=0;"";{late}))'


def volume_formula(port_ref: str, volume_count: str) -> str:
    return f'=SE({port_ref}="";"";SE({volume_count}=0;"";{volume_count}))'


def set_formula(worksheet, cell_ref: str, formula: str) -> None:
    call_with_retry(setattr, worksheet.Range(cell_ref), "FormulaLocal", formula)


def set_value(worksheet, cell_ref: str, value) -> None:
    call_with_retry(setattr, worksheet.Range(cell_ref), "Value", value)


def clear_cell(worksheet, cell_ref: str) -> None:
    call_with_retry(setattr, worksheet.Range(cell_ref), "Value", "")


def abs_cell(row: int, col: int) -> str:
    return f"${get_column_letter(col)}${row}"


def rel_cell(row: int, col: int) -> str:
    return f"{get_column_letter(col)}{row}"


def build_weekly_volume(port_ref: str, *, day_ref: str | None = None, product: str | None = None) -> str:
    parts = [
        criterion(sheet_range("ROE_wk", "$AN:$AN"), q(RODO)),
        criterion(sheet_range("ROE_wk", "$AV:$AV"), q(OK)),
        criterion(sheet_range("ROE_wk", "$AY:$AY"), port_ref),
    ]
    if day_ref is not None:
        parts.append(criterion(sheet_range("ROE_wk", "$AP:$AP"), day_ref))
    if product is not None:
        parts.append(criterion(sheet_range("ROE_wk", "$BR:$BR"), q(product)))
    return countifs(*parts)


def build_weekly_denominator(port_ref: str, *, day_ref: str | None = None, product: str | None = None) -> str:
    parts = [
        criterion(sheet_range("ROE_wk", "$AN:$AN"), q(RODO)),
        criterion(sheet_range("ROE_wk", "$AV:$AV"), q(OK)),
        criterion(sheet_range("ROE_wk", "$AY:$AY"), port_ref),
    ]
    if day_ref is not None:
        parts.append(criterion(sheet_range("ROE_wk", "$AP:$AP"), day_ref))
    if product is not None:
        parts.append(criterion(sheet_range("ROE_wk", "$BR:$BR"), q(product)))
    parts.extend(
        [
            criterion(sheet_range("ROE_wk", "$BL:$BL"), q("N")),
            criterion(sheet_range("ROE_wk", "$BO:$BO"), q(f"<>{ESPECIAL}")),
            criterion(sheet_range("ROE_wk", "$BB:$BB"), q(f"<>{SEM_PREENCHIMENTO}")),
        ]
    )
    return countifs(*parts)


def build_weekly_late(port_ref: str, *, day_ref: str | None = None, product: str | None = None) -> str:
    parts = [
        criterion(sheet_range("ROE_wk", "$AN:$AN"), q(RODO)),
        criterion(sheet_range("ROE_wk", "$AV:$AV"), q(OK)),
        criterion(sheet_range("ROE_wk", "$AY:$AY"), port_ref),
    ]
    if day_ref is not None:
        parts.append(criterion(sheet_range("ROE_wk", "$AP:$AP"), day_ref))
    if product is not None:
        parts.append(criterion(sheet_range("ROE_wk", "$BR:$BR"), q(product)))
    parts.extend(
        [
            criterion(sheet_range("ROE_wk", "$BL:$BL"), q("N")),
            criterion(sheet_range("ROE_wk", "$BO:$BO"), q(f"<>{ESPECIAL}")),
            criterion(sheet_range("ROE_wk", "$BB:$BB"), q(ATRASADO)),
        ]
    )
    return countifs(*parts)


def build_monthly_denominator(port_ref: str, *, alliance_filter: str | None = None) -> str:
    parts = [
        criterion(sheet_range("ROE_wk_monthly", "$AV:$AV"), q(OK)),
        criterion(sheet_range("ROE_wk_monthly", "$AY:$AY"), port_ref),
        criterion(sheet_range("ROE_wk_monthly", "$BG:$BG"), "$Q$1"),
        criterion(sheet_range("ROE_wk_monthly", "$AZ:$AZ"), q(f"<>{SEM_PREENCHIMENTO}")),
        criterion(sheet_range("ROE_wk_monthly", "$BI:$BI"), q(f"<>{ESPECIAL}")),
    ]
    if alliance_filter is not None:
        parts.append(criterion(sheet_range("ROE_wk_monthly", "$AI:$AI"), q(alliance_filter)))
    return countifs(*parts)


def build_monthly_late(port_ref: str, *, alliance_filter: str | None = None) -> str:
    parts = [
        criterion(sheet_range("ROE_wk_monthly", "$AV:$AV"), q(OK)),
        criterion(sheet_range("ROE_wk_monthly", "$AY:$AY"), port_ref),
        criterion(sheet_range("ROE_wk_monthly", "$BG:$BG"), "$Q$1"),
        criterion(sheet_range("ROE_wk_monthly", "$BI:$BI"), q(f"<>{ESPECIAL}")),
        criterion(sheet_range("ROE_wk_monthly", "$AZ:$AZ"), q(ATRASADO)),
    ]
    if alliance_filter is not None:
        parts.append(criterion(sheet_range("ROE_wk_monthly", "$AI:$AI"), q(alliance_filter)))
    return countifs(*parts)


def build_monthly_schedule_denominator(port_ref: str) -> str:
    return countifs(
        criterion(sheet_range("ROE_wk_monthly", "$AV:$AV"), q(OK)),
        criterion(sheet_range("ROE_wk_monthly", "$AY:$AY"), port_ref),
        criterion(sheet_range("ROE_wk_monthly", "$BG:$BG"), "$Q$1"),
        criterion(sheet_range("ROE_wk_monthly", "$AI:$AI"), q(ALLIANCE)),
    )


def build_monthly_schedule_late(port_ref: str) -> str:
    return countifs(
        criterion(sheet_range("ROE_wk_monthly", "$AV:$AV"), q(OK)),
        criterion(sheet_range("ROE_wk_monthly", "$AY:$AY"), port_ref),
        criterion(sheet_range("ROE_wk_monthly", "$BG:$BG"), "$Q$1"),
        criterion(sheet_range("ROE_wk_monthly", "$AI:$AI"), q(ALLIANCE)),
        criterion(sheet_range("ROE_wk_monthly", "$AT:$AT"), q("N")),
    )


def build_monthly_reagenda_formula(port_ref: str) -> str:
    start_date = "DATA(ANO($R$1);$Q$1;1)"
    end_date = "FIMMÊS(DATA(ANO($R$1);$Q$1;1);0)+1"
    return (
        "="
        + sumifs(
            sheet_range("Reagendas", "$Q:$Q"),
            criterion(sheet_range("Reagendas", "$F:$F"), port_ref),
            criterion(sheet_range("Reagendas", "$K:$K"), q(REAGENDA)),
            criterion(sheet_range("Reagendas", "$J:$J"), q(">=") + "&" + start_date),
            criterion(sheet_range("Reagendas", "$J:$J"), q("<") + "&" + end_date),
        )
    )


def build_monthly_hypercare(port_ref: str) -> str:
    return (
        "="
        + countifs(
            criterion(sheet_range("ROE_wk_monthly", "$AV:$AV"), q(OK)),
            criterion(sheet_range("ROE_wk_monthly", "$AY:$AY"), port_ref),
            criterion(sheet_range("ROE_wk_monthly", "$BG:$BG"), "$Q$1"),
            criterion(sheet_range("ROE_wk_monthly", "$AI:$AI"), q(ALLIANCE)),
            criterion(sheet_range("ROE_wk_monthly", "$N:$N"), q("*Reagenda*")),
        )
    )


def build_month_ratio_from_ports(child_rows: list[int], column: str, *, alliance_filter: str | None = None) -> str:
    port_refs = [f"$Q${row}" for row in child_rows]
    if column == "N":
        denominator_exprs = [build_monthly_schedule_denominator(port_ref) for port_ref in port_refs]
        late_exprs = [build_monthly_schedule_late(port_ref) for port_ref in port_refs]
    else:
        if column == "K":
            alliance_filter = ALLIANCE
        elif column == "L":
            alliance_filter = f"<>{ALLIANCE}"
        elif column == "M":
            alliance_filter = None
        denominator_exprs = [build_monthly_denominator(port_ref, alliance_filter=alliance_filter) for port_ref in port_refs]
        late_exprs = [build_monthly_late(port_ref, alliance_filter=alliance_filter) for port_ref in port_refs]
    denominator = sum_exprs(denominator_exprs)
    late = sum_exprs(late_exprs)
    return f'=SE(({denominator})=0;"";1-SEERRO(({late})/({denominator});0))'


def update_month_sheet(worksheet) -> int:
    changed = 0
    for row in MONTH_PORT_ROWS:
        port_ref = f"$Q${row}"
        formulas = {
            f"K{row}": ratio_formula(
                port_ref,
                build_monthly_denominator(port_ref, alliance_filter=ALLIANCE),
                build_monthly_late(port_ref, alliance_filter=ALLIANCE),
            ),
            f"L{row}": ratio_formula(
                port_ref,
                build_monthly_denominator(port_ref, alliance_filter=f"<>{ALLIANCE}"),
                build_monthly_late(port_ref, alliance_filter=f"<>{ALLIANCE}"),
            ),
            f"M{row}": ratio_formula(
                port_ref,
                build_monthly_denominator(port_ref),
                build_monthly_late(port_ref),
            ),
            f"N{row}": ratio_formula(
                port_ref,
                build_monthly_schedule_denominator(port_ref),
                build_monthly_schedule_late(port_ref),
            ),
            f"O{row}": f'=SE({port_ref}="";"";{build_monthly_reagenda_formula(port_ref)[1:]})',
            f"P{row}": f'=SE({port_ref}="";"";{build_monthly_hypercare(port_ref)[1:]})',
        }
        for cell_ref, formula in formulas.items():
            set_formula(worksheet, cell_ref, formula)
            changed += 1

    for row, child_rows in MONTH_REGION_GROUPS.items():
        set_formula(worksheet, f"K{row}", build_month_ratio_from_ports(child_rows, "K"))
        set_formula(worksheet, f"L{row}", build_month_ratio_from_ports(child_rows, "L"))
        set_formula(worksheet, f"M{row}", build_month_ratio_from_ports(child_rows, "M"))
        set_formula(worksheet, f"N{row}", build_month_ratio_from_ports(child_rows, "N"))
        set_formula(worksheet, f"O{row}", f'=SOMA(O{child_rows[0]}:O{child_rows[-1]})')
        set_formula(worksheet, f"P{row}", f'=SOMA(P{child_rows[0]}:P{child_rows[-1]})')
        changed += 6

    # Clear legacy helper rows that carried stale formulas from older layouts.
    for cell_ref in ["K22", "L22", "M22", "N22", "O22", "P22"]:
        clear_cell(worksheet, cell_ref)
        changed += 1

    return changed


def update_weekly_sheet(worksheet) -> int:
    changed = 0
    for block_start in WEEKLY_BLOCK_STARTS:
        day_row = block_start + 1
        oto_row = block_start + 2
        volume_row = block_start + 3
        oto_cab_row = block_start + 4
        qtd_cab_row = block_start + 5
        oto_ds_row = block_start + 6
        qtd_ds_row = block_start + 7

        for label_col in WEEKLY_LABEL_COLUMNS:
            port_col = label_col
            day_start_col = label_col + 1
            total_col = label_col + WEEKLY_PORT_GROUP_WIDTH - 1
            port_cell = worksheet.Cells(block_start, port_col)
            port_value = call_with_retry(lambda: port_cell.Value)
            target_columns = list(range(day_start_col, day_start_col + 7)) + [total_col]
            if port_value in (None, ""):
                for row in [oto_row, volume_row, oto_cab_row, qtd_cab_row, oto_ds_row, qtd_ds_row]:
                    for col in target_columns:
                        clear_cell(worksheet, rel_cell(row, col))
                continue

            port_ref = abs_cell(block_start, port_col)
            for col in range(day_start_col, day_start_col + 7):
                day_ref = abs_cell(day_row, col)
                total_expr = build_weekly_volume(port_ref, day_ref=day_ref)
                denominator_expr = build_weekly_denominator(port_ref, day_ref=day_ref)
                late_expr = build_weekly_late(port_ref, day_ref=day_ref)
                cab_den = build_weekly_denominator(port_ref, day_ref=day_ref, product="Cab")
                cab_late = build_weekly_late(port_ref, day_ref=day_ref, product="Cab")
                ds_den = build_weekly_denominator(port_ref, day_ref=day_ref, product="DS")
                ds_late = build_weekly_late(port_ref, day_ref=day_ref, product="DS")

                formulas = {
                    rel_cell(oto_row, col): ratio_formula(port_ref, denominator_expr, late_expr),
                    rel_cell(volume_row, col): volume_formula(port_ref, total_expr),
                    rel_cell(oto_cab_row, col): ratio_formula(port_ref, cab_den, cab_late),
                    rel_cell(qtd_cab_row, col): count_formula(port_ref, cab_den, cab_late),
                    rel_cell(oto_ds_row, col): ratio_formula(port_ref, ds_den, ds_late),
                    rel_cell(qtd_ds_row, col): count_formula(port_ref, ds_den, ds_late),
                }
                for cell_ref, formula in formulas.items():
                    set_formula(worksheet, cell_ref, formula)
                    changed += 1

            total_volume = build_weekly_volume(port_ref)
            total_den = build_weekly_denominator(port_ref)
            total_late = build_weekly_late(port_ref)
            total_cab_den = build_weekly_denominator(port_ref, product="Cab")
            total_cab_late = build_weekly_late(port_ref, product="Cab")
            total_ds_den = build_weekly_denominator(port_ref, product="DS")
            total_ds_late = build_weekly_late(port_ref, product="DS")

            total_formulas = {
                rel_cell(oto_row, total_col): ratio_formula(port_ref, total_den, total_late),
                rel_cell(volume_row, total_col): volume_formula(port_ref, total_volume),
                rel_cell(oto_cab_row, total_col): ratio_formula(port_ref, total_cab_den, total_cab_late),
                rel_cell(qtd_cab_row, total_col): count_formula(port_ref, total_cab_den, total_cab_late),
                rel_cell(oto_ds_row, total_col): ratio_formula(port_ref, total_ds_den, total_ds_late),
                rel_cell(qtd_ds_row, total_col): count_formula(port_ref, total_ds_den, total_ds_late),
            }
            for cell_ref, formula in total_formulas.items():
                set_formula(worksheet, cell_ref, formula)
                changed += 1

    return changed


def infer_missing_port_region(workbook_path: Path) -> list[dict[str, str | int]]:
    wb = openpyxl.load_workbook(workbook_path, keep_vba=True, data_only=True, read_only=True)
    week = wb["Week_Overview"]["AG1"].value
    ws = wb["ROE_wk"]

    populated_rows = []
    missing_rows = []
    for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[47] != OK or row[43] != week:
            continue
        entry = {
            "row": row_index,
            "booking": row[7],
            "client": row[12],
            "provider": row[3],
            "port": row[50],
            "region": row[51],
        }
        if row[50] in (None, "") or row[51] in (None, ""):
            missing_rows.append(entry)
        else:
            populated_rows.append(entry)

    by_booking: dict[object, list[str]] = defaultdict(list)
    by_client_provider: dict[tuple[object, object], list[str]] = defaultdict(list)
    by_provider: dict[object, list[str]] = defaultdict(list)

    for entry in populated_rows:
        by_booking[entry["booking"]].append(str(entry["port"]))
        by_client_provider[(entry["client"], entry["provider"])].append(str(entry["port"]))
        by_provider[entry["provider"]].append(str(entry["port"]))

    assignments = []
    for entry in missing_rows:
        candidates = Counter(by_booking.get(entry["booking"], []))
        reason = "booking"
        if len(candidates) != 1:
            candidates = Counter(by_client_provider.get((entry["client"], entry["provider"]), []))
            reason = "client_provider"
        if len(candidates) != 1:
            candidates = Counter(by_provider.get(entry["provider"], []))
            reason = "provider"
        if len(candidates) != 1:
            continue
        port = next(iter(candidates))
        region = PORT_TO_REGION.get(port)
        if region is None:
            continue
        assignments.append(
            {
                "row": entry["row"],
                "port": port,
                "region": region,
                "reason": reason,
            }
        )

    return assignments


def apply_inferred_port_region(worksheet, assignments: list[dict[str, str | int]]) -> None:
    for assignment in assignments:
        row = int(assignment["row"])
        set_value(worksheet, f"AY{row}", assignment["port"])
        set_value(worksheet, f"AZ{row}", assignment["region"])


def update_aux_and_day_week(workbook) -> int:
    changed = 0
    aux_ws = workbook.Worksheets("Aux")
    set_value(aux_ws, "A7", 1)
    set_value(aux_ws, "B7", SUN)
    changed += 2

    roe_ws = workbook.Worksheets("ROE_wk")
    day_body = call_with_retry(lambda: roe_ws.ListObjects("ROE_wk").ListColumns("day week").DataBodyRange)
    call_with_retry(setattr, day_body, "FormulaLocal", '=SEERRO(PROCX(AQ2;Aux!$A$1:$A$7;Aux!$B$1:$B$7;"");"")')
    changed += 1

    monthly_ws = workbook.Worksheets("ROE_wk_monthly")
    last_monthly_row = call_with_retry(lambda: monthly_ws.Cells(monthly_ws.Rows.Count, 1).End(-4162).Row)
    monthly_range = monthly_ws.Range(f"AP2:AP{last_monthly_row}")
    call_with_retry(setattr, monthly_range, "FormulaLocal", '=SEERRO(PROCX(AQ2;Aux!$A$1:$A$7;Aux!$B$1:$B$7;"");"")')
    changed += 1

    return changed


def refresh_all_pivots(workbook) -> int:
    refreshed = 0
    target_sheets = {
        "Top_Offenders_Customers",
        "Top_Offenders_Vendors",
        "Weekly_DS",
        "Din",
    }
    for worksheet in workbook.Worksheets:
        if worksheet.Name not in target_sheets:
            continue
        try:
            pivot_tables = worksheet.PivotTables()
            pivot_count = pivot_tables.Count
        except Exception:
            continue
        for index in range(1, pivot_count + 1):
            pivot_table = pivot_tables.Item(index)
            try:
                call_with_retry(pivot_table.RefreshTable)
                refreshed += 1
            except Exception:
                continue
    return refreshed


def calculate_targets(workbook, excel) -> None:
    target_sheets = [
        "ROE_wk",
        "ROE_wk_monthly",
        "Week_Overview",
        "Volume_DS",
        "Volume_Graph",
        "Volume_MAO",
        "Weekly",
        "Month2date",
        "LastMonth_Overview",
    ]
    for sheet_name in target_sheets:
        log(f"Calculating {sheet_name}...")
        call_with_retry(workbook.Worksheets(sheet_name).Calculate)
    log("Calculating workbook...")
    call_with_retry(excel.CalculateFull)


def main() -> int:
    if not WORKBOOK_PATH.exists():
        log(f"Workbook not found: {WORKBOOK_PATH}")
        return 1

    assignments = infer_missing_port_region(WORKBOOK_PATH)
    reason_counts = Counter(str(item["reason"]) for item in assignments)
    log(f"Inferred {len(assignments)} blank current-week port/region rows: {dict(reason_counts)}")

    backup_path = backup_workbook(WORKBOOK_PATH)
    log(f"Backup created at: {backup_path}")

    pythoncom.CoInitialize()
    excel = None
    workbook = None
    try:
        excel = win32com.client.gencache.EnsureDispatch("Excel.Application")
        call_with_retry(setattr, excel, "Visible", False)
        call_with_retry(setattr, excel, "DisplayAlerts", False)
        call_with_retry(setattr, excel, "EnableEvents", False)
        call_with_retry(setattr, excel, "ScreenUpdating", False)

        workbook = call_with_retry(
            excel.Workbooks.Open,
            str(WORKBOOK_PATH),
            UpdateLinks=0,
            ReadOnly=False,
            IgnoreReadOnlyRecommended=True,
            AddToMru=False,
            Notify=False,
        )
        call_with_retry(setattr, excel, "Calculation", -4135)
        call_with_retry(setattr, excel, "CalculateBeforeSave", False)

        changed = {}
        log("Updating Aux/day week...")
        changed["Aux/day week"] = update_aux_and_day_week(workbook)
        log("Updating Month2date...")
        changed["Month2date"] = update_month_sheet(workbook.Worksheets("Month2date"))
        log("Updating LastMonth_Overview...")
        changed["LastMonth_Overview"] = update_month_sheet(workbook.Worksheets("LastMonth_Overview"))
        log("Updating Weekly...")
        changed["Weekly"] = update_weekly_sheet(workbook.Worksheets("Weekly"))

        log("Applying inferred port/region fallbacks...")
        apply_inferred_port_region(workbook.Worksheets("ROE_wk"), assignments)
        changed["ROE_wk inferred port/region"] = len(assignments)

        calculate_targets(workbook, excel)

        log("Refreshing pivots...")
        pivot_count = refresh_all_pivots(workbook)
        changed["Pivot refresh"] = pivot_count

        call_with_retry(setattr, excel, "Calculation", -4105)
        call_with_retry(excel.CalculateFull)
        call_with_retry(workbook.Save)

        checks = {
            "Week_Overview!K23": call_with_retry(lambda: workbook.Worksheets("Week_Overview").Range("K23").Value),
            "Week_Overview!Q23": call_with_retry(lambda: workbook.Worksheets("Week_Overview").Range("Q23").Value),
            "Month2date!K2": call_with_retry(lambda: workbook.Worksheets("Month2date").Range("K2").Value),
            "Month2date!O2": call_with_retry(lambda: workbook.Worksheets("Month2date").Range("O2").Value),
            "LastMonth_Overview!K2": call_with_retry(lambda: workbook.Worksheets("LastMonth_Overview").Range("K2").Value),
            "Weekly!C4": call_with_retry(lambda: workbook.Worksheets("Weekly").Range("C4").Value),
            "Weekly!C4_formula": call_with_retry(lambda: workbook.Worksheets("Weekly").Range("C4").FormulaLocal),
            "ROE_wk!AY29": call_with_retry(lambda: workbook.Worksheets("ROE_wk").Range("AY29").Value),
            "ROE_wk!AZ29": call_with_retry(lambda: workbook.Worksheets("ROE_wk").Range("AZ29").Value),
            "Aux!A7": call_with_retry(lambda: workbook.Worksheets("Aux").Range("A7").Value),
            "Aux!B7": call_with_retry(lambda: workbook.Worksheets("Aux").Range("B7").Value),
        }
        log(f"Changes applied: {changed}")
        log(f"Validation: {checks}")
        return 0
    finally:
        if workbook is not None:
            try:
                call_with_retry(workbook.Close, SaveChanges=True)
            except Exception:
                pass
        if excel is not None:
            try:
                call_with_retry(excel.Quit)
            except Exception:
                pass
        pythoncom.CoUninitialize()


if __name__ == "__main__":
    sys.exit(main())
