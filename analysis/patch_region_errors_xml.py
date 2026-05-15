from __future__ import annotations

from pathlib import Path
from zipfile import ZipFile, ZIP_DEFLATED
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from collections import Counter
import html, json, re, shutil
from datetime import datetime

ROOT=Path(r'C:\Users\VNO024\Downloads\Github\PROJETO EXCEL COM CODEX')
WORKBOOK=Path(r'C:\Users\VNO024\Downloads\Base_DSU2026 - TbM - WK19.xlsm')
BACKUP_DIR=ROOT/'backups'
ANALYSIS_DIR=ROOT/'analysis'
REPORT=ANALYSIS_DIR/'Base_DSU2026_WK19_region_errors_xml_patch_report.json'
OLD_SHEET='Sem_Preenchimento_Regiao'
NEW_SHEET='Region errors'
DAYS=['Sun','Mon','Tue','Wed','Thu','Fri','Sat']
REGIONS=['North','Northeast','Southeast','South','Sem Porto/Region']
NS_MAIN='http://schemas.openxmlformats.org/spreadsheetml/2006/main'
NS_REL='http://schemas.openxmlformats.org/officeDocument/2006/relationships'
NS_PKG_REL='http://schemas.openxmlformats.org/package/2006/relationships'
ET.register_namespace('', NS_MAIN); ET.register_namespace('r', NS_REL)

def col_letter(n):
    s=''
    while n:
        n,rem=divmod(n-1,26); s=chr(65+rem)+s
    return s

def inline_cell(ref, value):
    return f'<c r="{ref}" t="inlineStr"><is><t>{html.escape(str(value))}</t></is></c>'

def num_formula_cell(ref, formula, value):
    val='' if value is None else str(int(value) if isinstance(value,(int,float)) and value==int(value) else value)
    return f'<c r="{ref}"><f>{html.escape(formula)}</f><v>{html.escape(val)}</v></c>'

def row_xml(r, cells):
    return f'<row r="{r}">' + ''.join(cells) + '</row>'

def norm_target(t):
    t=t.replace('\\','/')
    if t.startswith('/'): t=t.lstrip('/')
    elif not t.startswith('xl/'): t='xl/'+t
    return t

def workbook_sheet_map(z):
    wb=ET.fromstring(z.read('xl/workbook.xml'))
    rels=ET.fromstring(z.read('xl/_rels/workbook.xml.rels'))
    rm={rel.attrib['Id']: norm_target(rel.attrib['Target']) for rel in rels.findall(f'{{{NS_PKG_REL}}}Relationship')}
    result={}
    for s in wb.findall(f'.//{{{NS_MAIN}}}sheet'):
        result[s.attrib['name']]= {'rid': s.attrib[f'{{{NS_REL}}}id'], 'path': rm[s.attrib[f'{{{NS_REL}}}id']]}
    return result

def compute_counts():
    wb=load_workbook(WORKBOOK, read_only=True, keep_vba=False, data_only=True)
    week=wb['Week_Overview']['AG1'].value
    ws=wb['ROE_wk']
    headers=[c.value for c in next(ws.iter_rows(min_row=1,max_row=1))]
    idx={h:i for i,h in enumerate(headers) if h}
    counts=Counter(); totals_region=Counter(); totals_day=Counter(); total=0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx['Weeknum']] == week and row[idx['Volume']] == 'Ok' and row[idx['OTD ajustado']] == 'Sem Preenchimento':
            reg=row[idx['Region']] or 'Sem Porto/Region'
            day=row[idx['day week']]
            counts[(reg,day)]+=1; totals_region[reg]+=1; totals_day[day]+=1; total+=1
    wb.close()
    return week, counts, totals_region, totals_day, total

def build_sheet_data(week, counts, totals_region, totals_day, total):
    rows=[]
    rows.append(row_xml(1,[inline_cell('A1','Sem Preenchimento - Region errors')]))
    rows.append(row_xml(2,[inline_cell('A2','Semana ativa'), num_formula_cell('B2','Week_Overview!$AG$1',week)]))
    rows.append(row_xml(3,[inline_cell('A3','Regra'), inline_cell('B3','Volume=Ok + OTD ajustado=Sem Preenchimento')]))
    rows.append(row_xml(4,[inline_cell('A4','Observacao'), inline_cell('B4','Resumo por regiao e dia da semana da semana ativa')]))
    header=[inline_cell('A6','Region')]
    for c,day in enumerate(DAYS,start=2): header.append(inline_cell(f'{col_letter(c)}6',day))
    header.append(inline_cell('I6','Total'))
    rows.append(row_xml(6,header))
    start=7
    for r,reg in enumerate(REGIONS,start=start):
        cells=[inline_cell(f'A{r}',reg)]
        for c,day in enumerate(DAYS,start=2):
            ref=f'{col_letter(c)}{r}'
            if reg=='Sem Porto/Region': regcrit='""'
            else: regcrit=f'$A{r}'
            formula=f'COUNTIFS(ROE_wk[Volume],"Ok",ROE_wk[OTD ajustado],"Sem Preenchimento",ROE_wk[Weeknum],$B$2,ROE_wk[Region],{regcrit},ROE_wk[day week],{col_letter(c)}$6)'
            cells.append(num_formula_cell(ref,formula,counts.get((reg,day),0)))
        cells.append(num_formula_cell(f'I{r}',f'SUM(B{r}:H{r})',totals_region.get(reg,0)))
        rows.append(row_xml(r,cells))
    total_row=start+len(REGIONS)
    cells=[inline_cell(f'A{total_row}','Total Geral')]
    for c,day in enumerate(DAYS,start=2):
        letter=col_letter(c)
        cells.append(num_formula_cell(f'{letter}{total_row}',f'SUM({letter}{start}:{letter}{total_row-1})',totals_day.get(day,0)))
    cells.append(num_formula_cell(f'I{total_row}',f'SUM(I{start}:I{total_row-1})',total))
    rows.append(row_xml(total_row,cells))
    return ''.join(rows), total_row

def replace_between(text, pattern, replacement, insert_before=None):
    if re.search(pattern,text,flags=re.S):
        return re.sub(pattern,replacement,text,count=1,flags=re.S)
    if insert_before and insert_before in text:
        return text.replace(insert_before,replacement+insert_before,1)
    return text

def set_calc_attrs(text):
    def repl(m):
        tag=m.group(0)
        for k,v in [('calcMode','auto'),('forceFullCalc','0'),('fullCalcOnLoad','0'),('calcOnSave','1')]:
            if re.search(rf'\b{k}="[^"]*"',tag): tag=re.sub(rf'\b{k}="[^"]*"',f'{k}="{v}"',tag)
            else: tag=tag[:-2]+f' {k}="{v}"/>' if tag.endswith('/>') else tag[:-1]+f' {k}="{v}">'
        return tag
    if re.search(r'<calcPr\b[^>]*>',text): return re.sub(r'<calcPr\b[^>]*>',repl,text,count=1)
    return text.replace('</workbook>','<calcPr calcMode="auto" forceFullCalc="0" fullCalcOnLoad="0" calcOnSave="1"/></workbook>',1)

def main():
    BACKUP_DIR.mkdir(exist_ok=True); ANALYSIS_DIR.mkdir(exist_ok=True)
    backup=BACKUP_DIR/f"Base_DSU2026 - TbM - WK19_before_region_errors_xml_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsm"
    shutil.copy2(WORKBOOK,backup)
    week, counts, totals_region, totals_day, total=compute_counts()
    sheet_data,total_row=build_sheet_data(week,counts,totals_region,totals_day,total)
    with ZipFile(WORKBOOK,'r') as zin:
        smap=workbook_sheet_map(zin)
        sheet_info=smap.get(NEW_SHEET) or smap.get(OLD_SHEET)
        if not sheet_info: raise RuntimeError('Region errors source sheet not found')
        sheet_path=sheet_info['path']
        tmp=WORKBOOK.with_suffix('.xmlpatch.xlsm')
        with ZipFile(tmp,'w',compression=ZIP_DEFLATED,allowZip64=True) as zout:
            for item in zin.infolist():
                name=item.filename
                data=zin.read(name)
                if name=='xl/workbook.xml':
                    text=data.decode('utf-8',errors='replace')
                    text=text.replace(f'name="{html.escape(OLD_SHEET)}"',f'name="{NEW_SHEET}"')
                    text=set_calc_attrs(text)
                    data=text.encode('utf-8')
                elif name==sheet_path:
                    text=data.decode('utf-8',errors='replace')
                    text=re.sub(r'<dimension\b[^>]*/>',f'<dimension ref="A1:I{total_row}"/>',text,count=1)
                    # Replace columns with reasonable widths.
                    cols='<cols><col min="1" max="1" width="18" customWidth="1"/><col min="2" max="8" width="8" customWidth="1"/><col min="9" max="9" width="10" customWidth="1"/><col min="11" max="11" width="22" customWidth="1"/></cols>'
                    text=replace_between(text,r'<cols\b.*?</cols>',cols,insert_before='<sheetData')
                    text=replace_between(text,r'<sheetData\b.*?</sheetData>',f'<sheetData>{sheet_data}</sheetData>')
                    # Remove old merge cells to avoid stale merged title conflicts.
                    text=re.sub(r'<mergeCells\b.*?</mergeCells>','',text,flags=re.S)
                    data=text.encode('utf-8')
                # Drop calcChain to avoid stale dependency chain.
                if name=='xl/calcChain.xml':
                    continue
                zout.writestr(item,data)
    tmp.replace(WORKBOOK)
    report={
        'workbook':str(WORKBOOK),'backup':str(backup),'renamed_to':NEW_SHEET,'week':week,
        'summary_total':total,'regions':dict(totals_region),'days':dict(totals_day),'sheet_path':sheet_path,
        'workbook_size':WORKBOOK.stat().st_size,'workbook_mtime':datetime.fromtimestamp(WORKBOOK.stat().st_mtime).isoformat(timespec='seconds')
    }
    REPORT.write_text(json.dumps(report,ensure_ascii=False,indent=2),encoding='utf-8')
    print(json.dumps(report,ensure_ascii=False,indent=2))
    print('REPORT='+str(REPORT))
if __name__=='__main__': main()
