from __future__ import annotations

import json
import shutil
import time
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
import pythoncom
import pywintypes
import win32com.client

ROOT = Path(r"C:\Users\VNO024\Downloads\Github\PROJETO EXCEL COM CODEX")
WORKBOOK = Path(r"C:\Users\VNO024\Downloads\Base_DSU2026 - TbM - WK19.xlsm")
BACKUP_DIR = ROOT / "backups"
ANALYSIS_DIR = ROOT / "analysis"
REPORT = ANALYSIS_DIR / "Base_DSU2026_WK19_downloads_region_errors_update_report.json"

OLD_SHEET = "Sem_Preenchimento_Regiao"
NEW_SHEET = "Region errors"
MODULE_NAME = "modCodexAtualizacaoDashboards"
MACRO_NAME = "AtualizarDashboardsSemPreenchimento"
BUTTON_NAME = "btnAtualizarDashboardsSemPreenchimento"
BUTTON_TEXT = "Atualizar Dashboards"
BUTTON_ANCHOR = "L2"
REGIONS = ["North", "Northeast", "Southeast", "South", "Sem Porto/Region"]
TARGET_SHEETS = [
    NEW_SHEET,
    "Volume_DS",
    "Volume_MAO",
    "Volume_Graph",
    "Week_Overview",
    "Top_Offenders_Customers",
    "Top_Offenders_Vendors",
]

VBA_CODE = r'''
Option Explicit

Private Function GetActiveWeekNumber() As Variant
    On Error GoTo Fallback
    GetActiveWeekNumber = ThisWorkbook.Worksheets("Week_Overview").Range("AG1").Value
    Exit Function
Fallback:
    GetActiveWeekNumber = Empty
End Function

Private Function IsWeekField(ByVal fieldName As String) As Boolean
    Dim n As String
    n = LCase$(fieldName)
    IsWeekField = (InStr(1, n, "week", vbTextCompare) > 0 Or _
                   InStr(1, n, "semana", vbTextCompare) > 0 Or _
                   InStr(1, n, "weeknum", vbTextCompare) > 0)
End Function

Private Sub ApplyPageFilterIfExists(ByVal pf As PivotField, ByVal wantedValue As String)
    Dim pi As PivotItem
    On Error GoTo Sair

    If pf.Orientation <> xlPageField Then Exit Sub
    If Len(wantedValue) = 0 Then Exit Sub

    For Each pi In pf.PivotItems
        If CStr(pi.Name) = wantedValue Or CStr(pi.Value) = wantedValue Then
            pf.CurrentPage = pi.Name
            Exit Sub
        End If
    Next pi

Sair:
    Err.Clear
End Sub

Private Sub ApplyFirstAvailablePageFilter(ByVal pf As PivotField, ByVal options As Variant)
    Dim optionValue As Variant
    Dim pi As PivotItem
    On Error GoTo Sair

    If pf.Orientation <> xlPageField Then Exit Sub

    For Each optionValue In options
        For Each pi In pf.PivotItems
            If CStr(pi.Name) = CStr(optionValue) Or CStr(pi.Value) = CStr(optionValue) Then
                pf.CurrentPage = pi.Name
                Exit Sub
            End If
        Next pi
    Next optionValue

Sair:
    Err.Clear
End Sub

Private Sub ResetPivotFiltersSafely(ByVal pt As PivotTable, ByVal sheetName As String, ByVal activeWeek As Variant)
    Dim pf As PivotField
    Dim fieldName As String

    On Error Resume Next
    pt.ManualUpdate = True

    ' Limpa filtros temporarios para evitar pivot presa em selecao antiga.
    For Each pf In pt.PivotFields
        If pf.Orientation <> xlDataField Then
            pf.ClearAllFilters
        End If
    Next pf

    ' Reaplica filtros oficiais quando eles existem como filtro de pagina.
    For Each pf In pt.PivotFields
        fieldName = pf.Name

        If IsWeekField(fieldName) Then
            ApplyPageFilterIfExists pf, CStr(activeWeek)
        End If

        If InStr(1, sheetName, "Top_Offenders", vbTextCompare) > 0 Then
            If InStr(1, fieldName, "OTO Out", vbTextCompare) > 0 Then
                ApplyPageFilterIfExists pf, "N"
            End If
            If InStr(1, fieldName, "Atrasado", vbTextCompare) > 0 Then
                ApplyFirstAvailablePageFilter pf, Array("1", "Atrasado", "Sim", "S")
            End If
        End If
    Next pf

    pt.ManualUpdate = False
    On Error GoTo 0
End Sub

Public Sub AtualizarDashboardsSemPreenchimento()
    Dim targetSheets As Variant
    Dim sheetName As Variant
    Dim ws As Worksheet
    Dim pt As PivotTable
    Dim co As ChartObject
    Dim refreshedCaches As Object
    Dim cacheKey As String
    Dim activeWeek As Variant
    Dim updatedPivots As Long
    Dim updatedCharts As Long
    Dim updatedSheets As Long
    Dim updatedCaches As Long
    Dim oldScreenUpdating As Boolean
    Dim oldEnableEvents As Boolean
    Dim oldDisplayAlerts As Boolean
    Dim oldStatusBar As Variant
    Dim oldCalculation As XlCalculation
    Dim errors As String

    On Error GoTo FalhaGeral

    targetSheets = Array( _
        "Region errors", _
        "Volume_DS", _
        "Volume_MAO", _
        "Volume_Graph", _
        "Week_Overview", _
        "Top_Offenders_Customers", _
        "Top_Offenders_Vendors" _
    )

    Set refreshedCaches = CreateObject("Scripting.Dictionary")
    activeWeek = GetActiveWeekNumber()

    oldScreenUpdating = Application.ScreenUpdating
    oldEnableEvents = Application.EnableEvents
    oldDisplayAlerts = Application.DisplayAlerts
    oldStatusBar = Application.StatusBar
    oldCalculation = Application.Calculation

    Application.ScreenUpdating = False
    Application.EnableEvents = False
    Application.DisplayAlerts = False
    Application.Calculation = xlCalculationManual
    Application.StatusBar = "Atualizando dashboards DSU..."

    ' Atualiza cada cache uma unica vez. Isso evita atualizar somente o filtro visivel e reduz lentidao.
    For Each sheetName In targetSheets
        Set ws = ThisWorkbook.Worksheets(CStr(sheetName))
        For Each pt In ws.PivotTables
            On Error Resume Next
            cacheKey = CStr(pt.PivotCache.Index)
            If Not refreshedCaches.Exists(cacheKey) Then
                pt.PivotCache.Refresh
                refreshedCaches.Add cacheKey, True
                If Err.Number <> 0 Then
                    errors = errors & vbCrLf & CStr(sheetName) & ": erro ao atualizar cache da pivot '" & pt.Name & "' - " & Err.Description
                    Err.Clear
                Else
                    updatedCaches = updatedCaches + 1
                End If
            End If
            On Error GoTo FalhaGeral
        Next pt
    Next sheetName

    For Each sheetName In targetSheets
        Set ws = ThisWorkbook.Worksheets(CStr(sheetName))
        updatedSheets = updatedSheets + 1

        For Each pt In ws.PivotTables
            On Error Resume Next
            pt.PreserveFormatting = True
            ResetPivotFiltersSafely pt, CStr(sheetName), activeWeek
            pt.RefreshTable
            If Err.Number <> 0 Then
                errors = errors & vbCrLf & CStr(sheetName) & ": erro ao atualizar pivot '" & pt.Name & "' - " & Err.Description
                Err.Clear
            Else
                updatedPivots = updatedPivots + 1
            End If
            On Error GoTo FalhaGeral
        Next pt

        On Error Resume Next
        ws.Calculate
        If Err.Number <> 0 Then
            errors = errors & vbCrLf & CStr(sheetName) & ": erro ao recalcular aba - " & Err.Description
            Err.Clear
        End If
        On Error GoTo FalhaGeral

        For Each co In ws.ChartObjects
            On Error Resume Next
            co.Chart.Refresh
            If Err.Number <> 0 Then
                errors = errors & vbCrLf & CStr(sheetName) & ": erro ao atualizar grafico '" & co.Name & "' - " & Err.Description
                Err.Clear
            Else
                updatedCharts = updatedCharts + 1
            End If
            On Error GoTo FalhaGeral
        Next co
    Next sheetName

    Application.CalculateFull
    ThisWorkbook.Save

Finalizar:
    Application.StatusBar = oldStatusBar
    Application.DisplayAlerts = oldDisplayAlerts
    Application.EnableEvents = oldEnableEvents
    Application.ScreenUpdating = oldScreenUpdating
    Application.Calculation = oldCalculation

    If Len(errors) > 0 Then
        MsgBox "Atualizacao concluida com avisos." & vbCrLf & _
               "Abas atualizadas: " & updatedSheets & vbCrLf & _
               "Caches atualizados: " & updatedCaches & vbCrLf & _
               "Pivots atualizadas: " & updatedPivots & vbCrLf & _
               "Graficos atualizados: " & updatedCharts & vbCrLf & vbCrLf & _
               "Avisos:" & errors, vbExclamation, "Atualizar dashboards"
    Else
        MsgBox "Dashboards atualizados com sucesso." & vbCrLf & _
               "Abas atualizadas: " & updatedSheets & vbCrLf & _
               "Caches atualizados: " & updatedCaches & vbCrLf & _
               "Pivots atualizadas: " & updatedPivots & vbCrLf & _
               "Graficos atualizados: " & updatedCharts, vbInformation, "Atualizar dashboards"
    End If
    Exit Sub

FalhaGeral:
    errors = errors & vbCrLf & "Erro geral: " & Err.Description
    Resume Finalizar
End Sub
'''.strip()


def log(message: str) -> None:
    print(f"[region-errors] {message}", flush=True)


def call_with_retry(func, *args, retries: int = 30, delay: float = 0.35, **kwargs):
    last_exc = None
    for _ in range(retries):
        try:
            return func(*args, **kwargs)
        except pywintypes.com_error as exc:
            last_exc = exc
            time.sleep(delay)
            pythoncom.PumpWaitingMessages()
    raise last_exc


def backup_workbook() -> Path:
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = BACKUP_DIR / f"Base_DSU2026 - TbM - WK19_before_region_errors_{timestamp}.xlsm"
    shutil.copy2(WORKBOOK, backup)
    return backup


def compute_expected_counts() -> dict:
    wb = load_workbook(WORKBOOK, read_only=True, data_only=True, keep_vba=False)
    try:
        week = wb["Week_Overview"]["AG1"].value
        ws = wb["ROE_wk"]
        headers = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}
        required = ["Volume", "Weeknum", "OTD ajustado", "OTO Out", "Region", "Dia"]
        missing = [name for name in required if name not in headers]
        if missing:
            raise RuntimeError(f"Missing ROE_wk headers: {missing}")
        counts = Counter()
        min_day = None
        total = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            def val(name):
                return row[headers[name] - 1]
            if val("Volume") == "Ok" and val("Weeknum") == week and val("OTD ajustado") == "Sem Preenchimento" and val("OTO Out") == "N":
                region = val("Region") or "Sem Porto/Region"
                day = val("Dia")
                if hasattr(day, "date"):
                    day_key = day.date().isoformat()
                else:
                    day_key = str(day)
                counts[(day_key, region)] += 1
                counts[(day_key, "Total")] += 1
                total += 1
                if day is not None and (min_day is None or day < min_day):
                    min_day = day
        return {
            "week": week,
            "min_day": min_day.isoformat() if hasattr(min_day, "isoformat") else str(min_day),
            "total": total,
            "counts": {f"{k[0]}|{k[1]}": v for k, v in counts.items()},
        }
    finally:
        wb.close()


def get_or_rename_sheet(workbook, report: dict):
    sheet_names = [workbook.Worksheets.Item(i).Name for i in range(1, workbook.Worksheets.Count + 1)]
    if NEW_SHEET in sheet_names:
        report["sheet_renamed"] = False
        return workbook.Worksheets(NEW_SHEET)
    if OLD_SHEET not in sheet_names:
        raise RuntimeError(f"Neither '{OLD_SHEET}' nor '{NEW_SHEET}' exists")
    ws = workbook.Worksheets(OLD_SHEET)
    ws.Name = NEW_SHEET
    report["sheet_renamed"] = True
    return ws


def ensure_vba_module(workbook, report: dict) -> None:
    vbproject = workbook.VBProject
    component = None
    for idx in range(1, vbproject.VBComponents.Count + 1):
        candidate = vbproject.VBComponents.Item(idx)
        if candidate.Name == MODULE_NAME:
            component = candidate
            break
    if component is None:
        component = vbproject.VBComponents.Add(1)
        component.Name = MODULE_NAME
        report["module_created"] = True
    else:
        report["module_created"] = False
    code_module = component.CodeModule
    if code_module.CountOfLines:
        code_module.DeleteLines(1, code_module.CountOfLines)
    code_module.AddFromString(VBA_CODE)
    report["module_name"] = MODULE_NAME
    report["macro_name"] = MACRO_NAME
    report["module_lines"] = code_module.CountOfLines


def ensure_button(ws, workbook, report: dict) -> None:
    removed = 0
    for idx in range(ws.Shapes.Count, 0, -1):
        shape = ws.Shapes.Item(idx)
        if shape.Name == BUTTON_NAME:
            shape.Delete()
            removed += 1
    anchor = ws.Range(BUTTON_ANCHOR)
    shape = ws.Shapes.AddShape(5, anchor.Left, anchor.Top, 170, 34)
    shape.Name = BUTTON_NAME
    shape.OnAction = f"'{workbook.Name}'!{MACRO_NAME}"
    try:
        shape.TextFrame2.TextRange.Text = BUTTON_TEXT
        shape.TextFrame2.TextRange.Font.Size = 11
        shape.TextFrame2.TextRange.Font.Bold = True
        shape.TextFrame2.TextRange.Font.Fill.ForeColor.RGB = 16777215
        shape.TextFrame2.VerticalAnchor = 3
        shape.TextFrame2.TextRange.ParagraphFormat.Alignment = 2
        shape.Fill.ForeColor.RGB = 10498160
        shape.Line.ForeColor.RGB = 10498160
    except Exception:
        shape.TextFrame.Characters().Text = BUTTON_TEXT
    report["buttons_removed"] = removed
    report["button_name"] = BUTTON_NAME
    report["button_anchor"] = BUTTON_ANCHOR
    report["button_on_action"] = shape.OnAction


def set_region_errors_layout(ws, report: dict) -> None:
    # Keep the existing region summary in A:B and add the day-by-region panel from C:K.
    ws.Range("C1:K20").Clear()
    ws.Range("C1").Value = "Sem Preenchimento por dia e região"
    ws.Range("C2").Value = "Week"
    ws.Range("D2").Formula = "=Week_Overview!$AG$1"
    ws.Range("C4").Value = "Day"
    ws.Range("D4").Value = "Date"
    for offset, region in enumerate(REGIONS, start=5):
        ws.Cells(4, offset).Value = region
    ws.Range("J4").Value = "Total"

    # Base date for current week.
    ws.Range("D3").Formula = '=IFERROR(MINIFS(ROE_wk[Dia],ROE_wk[Weeknum],$D$2),"")'
    ws.Range("C3").Value = "First date"

    for i in range(7):
        row = 5 + i
        ws.Cells(row, 3).Value = f"Day {i + 1}"
        ws.Cells(row, 4).Formula = f'=IF($D$3="","",$D$3+{i})'
        # Region formulas E:I.
        for col, region in zip(range(5, 10), REGIONS):
            if region == "Sem Porto/Region":
                criterion = '""'
            else:
                criterion = f'{ws.Cells(4, col).Address}'
            ws.Cells(row, col).Formula = (
                f'=IF($D{row}="","",COUNTIFS('
                f'ROE_wk[Weeknum],$D$2,'
                f'ROE_wk[Volume],"Ok",'
                f'ROE_wk[OTO Out],"N",'
                f'ROE_wk[OTD ajustado],"Sem Preenchimento",'
                f'ROE_wk[Dia],$D{row},'
                f'ROE_wk[Region],{criterion}))'
            )
        ws.Cells(row, 10).Formula = f'=IF($D{row}="","",SUM(E{row}:I{row}))'

    ws.Range("C12").Value = "Total"
    for col in range(5, 11):
        col_letter = ws.Cells(4, col).Address.split("$")[1]
        ws.Cells(12, col).Formula = f"=SUM({col_letter}5:{col_letter}11)"

    # Formatting.
    ws.Range("C1:J1").Merge()
    ws.Range("C1:J1").Font.Bold = True
    ws.Range("C1:J1").Font.Size = 13
    ws.Range("C4:J4").Font.Bold = True
    ws.Range("C12:J12").Font.Bold = True
    ws.Range("D5:D11").NumberFormat = "ddd dd/mm"
    ws.Range("C4:J12").Borders.LineStyle = 1
    ws.Range("C:K").Columns.AutoFit()
    report["region_errors_panel_range"] = "C1:J12"


def collect_panel_values(ws) -> list[dict]:
    rows = []
    for row in range(5, 12):
        rows.append({
            "row": row,
            "day": ws.Cells(row, 3).Value,
            "date": str(ws.Cells(row, 4).Value),
            "North": ws.Cells(row, 5).Value,
            "Northeast": ws.Cells(row, 6).Value,
            "Southeast": ws.Cells(row, 7).Value,
            "South": ws.Cells(row, 8).Value,
            "Sem Porto/Region": ws.Cells(row, 9).Value,
            "Total": ws.Cells(row, 10).Value,
        })
    return rows


def main() -> int:
    if not WORKBOOK.exists():
        raise FileNotFoundError(WORKBOOK)
    ANALYSIS_DIR.mkdir(parents=True, exist_ok=True)

    report: dict = {
        "workbook": str(WORKBOOK),
        "started_at": datetime.now().isoformat(timespec="seconds"),
        "old_sheet": OLD_SHEET,
        "new_sheet": NEW_SHEET,
        "target_sheets": TARGET_SHEETS,
    }

    expected = compute_expected_counts()
    report["expected"] = {"week": expected["week"], "min_day": expected["min_day"], "total": expected["total"]}

    backup = backup_workbook()
    report["backup"] = str(backup)
    log(f"Backup criado: {backup}")

    pythoncom.CoInitialize()
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.EnableEvents = False
    excel.AskToUpdateLinks = False

    workbook = None
    try:
        workbook = call_with_retry(
            excel.Workbooks.Open,
            str(WORKBOOK),
            UpdateLinks=False,
            ReadOnly=False,
            IgnoreReadOnlyRecommended=True,
            AddToMru=False,
            Notify=False,
        )
        report["opened"] = True
        try:
            workbook.ForceFullCalculation = False
            report["force_full_calculation_set_false"] = True
        except Exception as exc:
            report["force_full_calculation_set_false"] = False
            report["force_full_calculation_error"] = str(exc)

        ws = get_or_rename_sheet(workbook, report)
        set_region_errors_layout(ws, report)
        ensure_vba_module(workbook, report)
        ensure_button(ws, workbook, report)

        excel.Calculation = -4105  # xlCalculationAutomatic
        excel.CalculateFull()
        report["calculate_full"] = True
        call_with_retry(workbook.Save)
        report["saved"] = True

        # Validate after calculation/save.
        sheet_names = [workbook.Worksheets.Item(i).Name for i in range(1, workbook.Worksheets.Count + 1)]
        report["old_sheet_exists_after"] = OLD_SHEET in sheet_names
        report["new_sheet_exists_after"] = NEW_SHEET in sheet_names
        report["missing_target_sheets_after"] = [s for s in TARGET_SHEETS if s not in sheet_names]
        report["panel_values"] = collect_panel_values(ws)
        report["panel_total"] = ws.Range("J12").Value
        report["panel_total_matches_expected"] = (ws.Range("J12").Value == expected["total"])

        # Validate button and macro.
        button_found = False
        button_on_action = None
        for idx in range(1, ws.Shapes.Count + 1):
            shape = ws.Shapes.Item(idx)
            if shape.Name == BUTTON_NAME:
                button_found = True
                button_on_action = shape.OnAction
                break
        report["button_found_after"] = button_found
        report["button_on_action_after"] = button_on_action
        module_found = False
        macro_found = False
        vbproject = workbook.VBProject
        for idx in range(1, vbproject.VBComponents.Count + 1):
            comp = vbproject.VBComponents.Item(idx)
            if comp.Name == MODULE_NAME:
                module_found = True
                text = comp.CodeModule.Lines(1, comp.CodeModule.CountOfLines)
                macro_found = f"Sub {MACRO_NAME}" in text or f"Sub {MACRO_NAME}(" in text
                break
        report["module_found_after"] = module_found
        report["macro_found_after"] = macro_found
    finally:
        if workbook is not None:
            try:
                workbook.Close(SaveChanges=False)
            except Exception:
                pass
        try:
            excel.Quit()
        except Exception:
            pass
        pythoncom.CoUninitialize()

    report["finished_at"] = datetime.now().isoformat(timespec="seconds")
    report["workbook_size"] = WORKBOOK.stat().st_size
    report["workbook_mtime"] = datetime.fromtimestamp(WORKBOOK.stat().st_mtime).isoformat(timespec="seconds")
    REPORT.write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2, default=str))
    print(f"REPORT={REPORT}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
