
import json, math, re, zipfile, xml.etree.ElementTree as ET
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook

WORKBOOK = Path(r"C:\Users\VNO024\Downloads\Base_DSU2026 - TbM - WK19.xlsm")
OUTDIR = Path(r"C:\Users\VNO024\Downloads\Github\PROJETO EXCEL COM CODEX\analysis")
OUTDIR.mkdir(parents=True, exist_ok=True)
TS = datetime.now().strftime('%Y%m%d_%H%M%S')
REPORT_JSON = OUTDIR / f"Base_DSU2026_WK19_final_audit_{TS}.json"
REPORT_MD = OUTDIR / f"Base_DSU2026_WK19_final_audit_{TS}.md"

NS = {
    'main':'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
    'rel':'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
    'pkgrel':'http://schemas.openxmlformats.org/package/2006/relationships',
    'c':'http://schemas.openxmlformats.org/drawingml/2006/chart',
}

DAYS = ['Sun','Mon','Tue','Wed','Thu','Fri','Sat']
TOL = 1e-9
ALI = 'Alian' + chr(0xE7) + 'a'
BAD_ALI = 'Alian?a'

def norm(v):
    if v is None:
        return ''
    if isinstance(v, str):
        return v.strip()
    return v

def normstr(v):
    v = norm(v)
    return '' if v == '' else str(v).strip()

def weeknum(v):
    try:
        if v is None or v == '': return None
        return int(float(v))
    except Exception:
        return None

def is_ok(a,b,tol=TOL):
    if a in (None,'') and b in (None,''):
        return True
    if isinstance(a,(int,float)) and isinstance(b,(int,float)):
        return abs(float(a)-float(b)) <= tol
    return a == b

def as_num(v):
    if v in (None,''):
        return None
    try: return float(v)
    except Exception: return v

def safe_round(v, nd=2):
    if v in (None,''): return ''
    return round(v, nd)

def scan_workbook_xml():
    result = {
        'worksheet_error_cells': {},
        'parts_with_formula_error_text': [],
        'formulas_with_mojibake_alianca': [],
        'defined_name_info_only': [],
        'pivot_cache_error_items': [],
    }
    with zipfile.ZipFile(WORKBOOK) as z:
        wb_xml = ET.fromstring(z.read('xl/workbook.xml'))
        rels = ET.fromstring(z.read('xl/_rels/workbook.xml.rels'))
        rid_target = {r.attrib['Id']: r.attrib['Target'] for r in rels}
        sheet_map = {}
        for s in wb_xml.findall('main:sheets/main:sheet', NS):
            name = s.attrib['name']
            rid = s.attrib['{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id']
            target = rid_target[rid]
            if not target.startswith('/'):
                target = 'xl/' + target.lstrip('/')
            sheet_map[target] = name
        for name in z.namelist():
            if not name.endswith('.xml'):
                continue
            text = z.read(name).decode('utf-8', errors='replace')
            if any(e in text for e in ['#REF!','#DIV/0!','#VALUE!']):
                result['parts_with_formula_error_text'].append(name)
            if name.startswith('xl/pivotCache/') and any(e in text for e in ['#N/A','#REF!','#DIV/0!','#VALUE!']):
                for token in ['#N/A','#REF!','#DIV/0!','#VALUE!']:
                    if token in text:
                        result['pivot_cache_error_items'].append({'part': name, 'token': token, 'count': text.count(token)})
            if name == 'xl/workbook.xml':
                # _xleta.* and slicer #N/A are metadata and not active worksheet formula errors.
                for token in ['#NAME?','#N/A']:
                    if token in text:
                        result['defined_name_info_only'].append({'token': token, 'count': text.count(token)})
            if name.startswith('xl/worksheets/sheet'):
                sh = sheet_map.get(name, name)
                root = ET.fromstring(z.read(name))
                errors = []
                for c in root.iter('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c'):
                    addr = c.attrib.get('r')
                    t = c.attrib.get('t')
                    f = c.find('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}f')
                    v = c.find('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v')
                    ft = f.text if f is not None and f.text else ''
                    vt = v.text if v is not None and v.text else ''
                    if BAD_ALI in ft:
                        result['formulas_with_mojibake_alianca'].append({'sheet': sh, 'cell': addr, 'formula': ft})
                    if t == 'e' or vt in ['#DIV/0!','#REF!','#VALUE!','#NAME?','#N/A'] or any(e in ft for e in ['#REF!','#DIV/0!','#VALUE!']):
                        errors.append({'cell': addr, 'type': t, 'formula': ft[:250], 'value': vt})
                if errors:
                    result['worksheet_error_cells'][sh] = errors[:50]
    return result

def load_base_and_sheets():
    wb = load_workbook(WORKBOOK, read_only=True, data_only=True, keep_vba=False)
    wbf = load_workbook(WORKBOOK, read_only=True, data_only=False, keep_vba=False)
    active_week = wb['Week_Overview']['AG1'].value
    active_week = weeknum(active_week)
    headers = [c.value for c in next(wb['ROE_wk'].iter_rows(min_row=1, max_row=1))]
    idx = {h:i for i,h in enumerate(headers) if h is not None}
    needed = ['Provedor','Centro de Custo','day week','Weeknum','SLA Ag','Volume','Porto','Region','OTD ajustado','Atrasado?','OTO Out','Especiais','AtrasoRev']
    missing = [h for h in needed if h not in idx]
    rows=[]
    counters=Counter()
    for row in wb['ROE_wk'].iter_rows(min_row=2, values_only=True):
        if missing:
            break
        if normstr(row[idx['Volume']]) == 'Ok' and weeknum(row[idx['Weeknum']]) == active_week:
            r = {h: row[idx[h]] for h in needed}
            rows.append(r)
            counters[(normstr(r['OTO Out']), normstr(r['Atrasado?']), normstr(r['OTD ajustado']))] += 1
    return wb, wbf, active_week, idx, rows, counters

def filter_rows(rows, port=None, region=None, centro=None, day=None, oto_scope=False, eligible=False, delayed_penalty=False, sem_preench=False, sla_n=False, special=None):
    res=[]
    for r in rows:
        if port is not None and normstr(r['Porto']) != port: continue
        if region is not None:
            if region == '':
                if normstr(r['Region']) != '': continue
            elif normstr(r['Region']) != region: continue
        if day is not None and normstr(r['day week']) != day: continue
        if centro is not None:
            cc = normstr(r['Centro de Custo'])
            if centro == ALI:
                if cc != ALI: continue
            elif centro == '<>'+ALI:
                if cc == ALI: continue
            else:
                if cc != centro: continue
        if oto_scope and normstr(r['OTO Out']) != 'N': continue
        if eligible and normstr(r['OTD ajustado']) == 'Sem Preenchimento': continue
        if delayed_penalty:
            if normstr(r['OTD ajustado']) != 'Atrasado': continue
            if normstr(r['Especiais']) == 'Especial': continue
        if sem_preench and normstr(r['OTD ajustado']) != 'Sem Preenchimento': continue
        if sla_n and normstr(r['SLA Ag']) != 'N': continue
        if special is not None:
            if special and normstr(r['Especiais']) != 'Especial': continue
            if (not special) and normstr(r['Especiais']) == 'Especial': continue
        res.append(r)
    return res

def oto_calc(rows, port=None, region=None, centro=None, day=None, round2=False):
    eligible_rows = filter_rows(rows, port=port, region=region, centro=centro, day=day, oto_scope=True, eligible=True)
    if not eligible_rows:
        return ''
    delayed = filter_rows(rows, port=port, region=region, centro=centro, day=day, oto_scope=True, delayed_penalty=True)
    val = 1 - len(delayed) / len(eligible_rows)
    return round(val,2) if round2 else val

def sla48(rows, port=None, region=None, centro=ALI, day=None):
    denom = filter_rows(rows, port=port, region=region, centro=centro, day=day)
    if not denom: return ''
    bad = filter_rows(rows, port=port, region=region, centro=centro, day=day, sla_n=True)
    return 1 - len(bad)/len(denom)

def compare_week_overview(wb, rows, active_week):
    ws = wb['Week_Overview']
    checks=[]
    # rows and their source filter
    rowdefs = {2:('MAO','port'),3:('VDC','port'),4:('North','region'),6:('PEC','port'),7:('SUA','port'),8:('SSA','port'),9:('Northeast','region'),11:('VIX','port'),12:('RIO','port'),13:('SSZ','port'),14:('Southeast','region'),16:('IOA','port'),17:('ITJ','port'),18:('PNG','port'),19:('IBB','port'),20:('RIG','port'),21:('South','region'),23:('Total BR','total')}
    for r,(label,typ) in rowdefs.items():
        port = region = None
        if typ == 'port': port = normstr(ws.cell(r,33).value)
        elif typ == 'region': region = normstr(ws.cell(r,1).value)
        # compute counts
        calc_D = len(filter_rows(rows, port=port, region=region, centro=ALI))
        calc_H = len(filter_rows(rows, port=port, region=region, centro='<>'+ALI))
        calc_K = len(filter_rows(rows, port=port, region=region))
        calc_N = oto_calc(rows, port=port, region=region, centro=ALI, round2=(r==23))
        calc_P = oto_calc(rows, port=port, region=region, centro='<>'+ALI, round2=(r==23))
        calc_Q = oto_calc(rows, port=port, region=region)
        calc_R = sla48(rows, port=port, region=region, centro=ALI)
        for col, name, expected in [(4,'Moves CAB',calc_D),(8,'Moves DS',calc_H),(11,'Moves Total',calc_K),(14,'OTO CAB',calc_N),(16,'OTO DS',calc_P),(17,'OTO Total',calc_Q),(18,'48h Schedule CAB',calc_R)]:
            actual = ws.cell(r,col).value
            checks.append({'row': r, 'label': label, 'filter_type': typ, 'filter': port or region or 'ALL', 'metric': name, 'cell': ws.cell(r,col).coordinate, 'actual': actual, 'expected_from_base': expected, 'ok': is_ok(as_num(actual), as_num(expected), tol=1e-8)})
    return checks

def compare_region_errors(wb, wbf, rows):
    ws = wb['Region errors']; wsf = wbf['Region errors']
    checks=[]; formula_issues=[]
    # Verify formulas exist in expected cells
    for r in range(7,13):
        for c in range(2,10):
            f = wsf.cell(r,c).value
            if not isinstance(f,str) or not f.startswith('='):
                formula_issues.append({'cell': wsf.cell(r,c).coordinate, 'value': f})
    for r in range(7,12):
        label=normstr(ws.cell(r,1).value)
        region = '' if label == 'Sem Porto/Region' else label
        for c,day in zip(range(2,9),DAYS):
            expected=len(filter_rows(rows, region=region, day=day, sem_preench=True))
            actual=ws.cell(r,c).value
            checks.append({'cell': ws.cell(r,c).coordinate, 'region':label, 'day':day, 'actual':actual, 'expected_from_base':expected, 'ok':actual==expected})
        actual_total=ws.cell(r,9).value
        expected_total=sum(x['expected_from_base'] for x in checks if x.get('region')==label and 'day' in x)
        checks.append({'cell': ws.cell(r,9).coordinate, 'region':label, 'day':'Total', 'actual':actual_total, 'expected_from_base':expected_total, 'ok':actual_total==expected_total})
    # total row
    for c,day in zip(range(2,9),DAYS):
        expected=sum(ch['expected_from_base'] for ch in checks if ch.get('day')==day and ch.get('region')!='Total Geral')
        actual=ws.cell(12,c).value
        checks.append({'cell':ws.cell(12,c).coordinate, 'region':'Total Geral','day':day,'actual':actual,'expected_from_base':expected,'ok':actual==expected})
    actual=ws.cell(12,9).value
    expected=sum(ch['expected_from_base'] for ch in checks if ch.get('day')=='Total' and ch.get('region')!='Total Geral')
    checks.append({'cell':'I12','region':'Total Geral','day':'Total','actual':actual,'expected_from_base':expected,'ok':actual==expected})
    return {'checks': checks, 'formula_issues': formula_issues}

def compare_volume_sheets(wb, rows):
    results=[]
    for s in ['Volume_DS','Volume_MAO','Volume_Graph']:
        ws=wb[s]
        week=weeknum(ws['M2'].value)
        port=normstr(ws['M3'].value)
        if week is None:
            continue
        # Audit only if selected week equals active base rows already loaded.
        for c,day in zip(range(6,13),DAYS):
            expected=len(filter_rows(rows, port=port, day=day))
            actual=ws.cell(9,c).value
            results.append({'sheet':s,'cell':ws.cell(9,c).coordinate,'metric':'Volume Total','port':port,'day':day,'actual':actual,'expected_from_base':expected,'ok':actual==expected})
        results.append({'sheet':s,'cell':'M9','metric':'Volume Total','port':port,'day':'Total','actual':ws['M9'].value,'expected_from_base':len(filter_rows(rows, port=port)),'ok':ws['M9'].value==len(filter_rows(rows, port=port))})
        # relevant CAB/DS rows per sheet
        if s == 'Volume_DS':
            row_cab,row_ds,row_oto,row_sla = 11,12,16,17
            metrics=[(row_cab,'Volume CAB',ALI),(row_ds,'Volume DS','<>'+ALI)]
            oto_center='<>'+ALI; oto_row=row_oto; sla_center='<>'+ALI; sla_row=row_sla
        elif s == 'Volume_MAO':
            metrics=[(12,'Volume CAB Diversos',ALI+'_non_special'),(13,'Volume CAB Special',ALI+'_special'),(14,'Volume DS','<>'+ALI)]
            oto_center=ALI; oto_row=19; sla_center=ALI; sla_row=20
        else:
            metrics=[(12,'Volume CAB',ALI),(13,'Volume DS','<>'+ALI)]
            oto_center=ALI; oto_row=17; sla_center=ALI; sla_row=18
        for rownum,metric,center in metrics:
            for c,day in zip(range(6,13),DAYS):
                if center==ALI+'_non_special': expected=len(filter_rows(rows, port=port, day=day, centro=ALI, special=False))
                elif center==ALI+'_special': expected=len(filter_rows(rows, port=port, day=day, centro=ALI, special=True))
                else: expected=len(filter_rows(rows, port=port, day=day, centro=center))
                actual=ws.cell(rownum,c).value
                results.append({'sheet':s,'cell':ws.cell(rownum,c).coordinate,'metric':metric,'port':port,'day':day,'actual':actual,'expected_from_base':expected,'ok':actual==expected})
            # Total cell in M: exact formula may be a COUNTIFS or SUM; compare to row sum/base.
            if center==ALI+'_non_special': expected=len(filter_rows(rows, port=port, centro=ALI, special=False))
            elif center==ALI+'_special': expected=len(filter_rows(rows, port=port, centro=ALI, special=True))
            else: expected=len(filter_rows(rows, port=port, centro=center))
            actual=ws.cell(rownum,13).value
            results.append({'sheet':s,'cell':ws.cell(rownum,13).coordinate,'metric':metric,'port':port,'day':'Total','actual':actual,'expected_from_base':expected,'ok':actual==expected})
        # OTO/SLA totals where visible
        actual=ws.cell(oto_row,13).value
        expected=round(oto_calc(rows, port=port, centro=oto_center),2) if oto_calc(rows, port=port, centro=oto_center) != '' else None
        ok=(actual in (None,'') and expected is None) or is_ok(as_num(actual),as_num(expected),1e-8)
        results.append({'sheet':s,'cell':ws.cell(oto_row,13).coordinate,'metric':'OTO total','port':port,'day':'Total','actual':actual,'expected_from_base':expected,'ok':ok})
        actual=ws.cell(sla_row,13).value
        expected=sla48(rows, port=port, centro=sla_center)
        ok=(actual in (None,'') and expected == '') or is_ok(as_num(actual),as_num(expected),1e-8)
        results.append({'sheet':s,'cell':ws.cell(sla_row,13).coordinate,'metric':'48h total','port':port,'day':'Total','actual':actual,'expected_from_base':expected,'ok':ok})
    return results

def inspect_top_offenders(wb, rows):
    result={}
    # Customer totals and helper cell checks
    ws=wb['Top_Offenders_Customers']
    result['Top_Offenders_Customers']={
        'visao_geral_total_row': {'row':510, 'A:G':[ws.cell(510,c).value for c in range(1,8)]},
        'hypercare_norte_mirror_total': {'row':441,'T:W':[ws.cell(441,c).value for c in range(20,24)]},
        'g27': ws['G27'].value,
    }
    # Vendors should not be blank if configured as in older good backup: base has providers and delays.
    wv=wb['Top_Offenders_Vendors']
    vendor_rows=[r for r in range(8, min(40,wv.max_row)+1) if normstr(wv.cell(r,1).value) not in ('','Total Geral','Grand Total')]
    delayed_vendor_count=len(filter_rows(rows, oto_scope=True, delayed_penalty=True))
    total_provider_count=len({normstr(r['Provedor']) for r in rows if normstr(r['Provedor'])})
    result['Top_Offenders_Vendors']={
        'filters_visible': {
            'B2':wv['B2'].value,'B3':wv['B3'].value,'B4':wv['B4'].value,'B5':wv['B5'].value,
            'M2':wv['M2'].value,'M3':wv['M3'].value,'M4':wv['M4'].value,'M5':wv['M5'].value,
        },
        'visible_vendor_data_rows_A8_A40': len(vendor_rows),
        'a8_e8':[wv.cell(8,c).value for c in range(1,6)],
        'base_distinct_providers_week': total_provider_count,
        'base_penalized_delayed_rows_oto_scope': delayed_vendor_count,
        'status': 'OK' if vendor_rows else 'BLANK_PIVOT_WHILE_BASE_HAS_DATA'
    }
    return result

def parse_charts():
    # Extract chart source formulas for target sheets and flag ranges with error tokens or blank formulas.
    result=[]
    with zipfile.ZipFile(WORKBOOK) as z:
        wb_xml=ET.fromstring(z.read('xl/workbook.xml'))
        rels=ET.fromstring(z.read('xl/_rels/workbook.xml.rels'))
        rid_target={r.attrib['Id']:r.attrib['Target'] for r in rels}
        sheet_targets={}
        for s in wb_xml.findall('main:sheets/main:sheet',NS):
            rid=s.attrib['{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id']
            target='xl/'+rid_target[rid].lstrip('/')
            sheet_targets[target]=s.attrib['name']
        for sheet_path, sheet_name in sheet_targets.items():
            if sheet_name not in ['Volume_DS','Volume_MAO','Volume_Graph','Week_Overview','Top_Offenders_Customers','Top_Offenders_Vendors','Menu','Region errors']:
                continue
            rel_path=sheet_path.replace('xl/worksheets/','xl/worksheets/_rels/')+'.rels'
            if rel_path not in z.namelist():
                continue
            relroot=ET.fromstring(z.read(rel_path))
            drawing_targets=[]
            for r in relroot:
                if 'drawing' in r.attrib.get('Type',''):
                    target=r.attrib['Target']
                    if target.startswith('../'):
                        target='xl/'+target[3:]
                    elif not target.startswith('xl/'):
                        target='xl/worksheets/'+target
                    drawing_targets.append(target)
            for drawing in drawing_targets:
                draw_rel=drawing.replace('xl/drawings/','xl/drawings/_rels/')+'.rels'
                if draw_rel not in z.namelist():
                    continue
                dr=ET.fromstring(z.read(draw_rel))
                for r in dr:
                    if 'chart' in r.attrib.get('Type',''):
                        target=r.attrib['Target']
                        if target.startswith('../'):
                            chart_path='xl/'+target[3:]
                        else:
                            chart_path='xl/drawings/'+target
                        if chart_path in z.namelist():
                            chart_xml=ET.fromstring(z.read(chart_path))
                            formulas=[]
                            for f in chart_xml.iter('{http://schemas.openxmlformats.org/drawingml/2006/chart}f'):
                                if f.text:
                                    formulas.append(f.text)
                            result.append({'sheet':sheet_name,'chart_part':chart_path,'formula_count':len(formulas),'sample_formulas':formulas[:12],'has_error_ref':any('#REF!' in x or '#DIV/0!' in x for x in formulas)})
    return result

xml_scan=scan_workbook_xml()
wb,wbf,active_week,idx,rows,counters=load_base_and_sheets()
base_summary={
    'active_week':active_week,
    'roe_wk_rows_volume_ok_active_week':len(rows),
    'header_missing': [h for h in ['Provedor','Centro de Custo','day week','Weeknum','SLA Ag','Volume','Porto','Region','OTD ajustado','Atrasado?','OTO Out','Especiais','AtrasoRev'] if h not in idx],
    'status_patterns':[{ 'oto_out':k[0],'atrasado':k[1],'otd_ajustado':k[2],'count':v} for k,v in counters.most_common()],
    'sem_preenchimento_count':len(filter_rows(rows, sem_preench=True)),
    'kpi_denominator_excludes_sem_preenchimento':len(filter_rows(rows, oto_scope=True, eligible=True)),
    'kpi_penalized_delays_excludes_special':len(filter_rows(rows, oto_scope=True, delayed_penalty=True)),
    'oto_total_expected':oto_calc(rows),
    'special_rows_active_week':len(filter_rows(rows, special=True)),
    'special_delayed_active_week':len([r for r in filter_rows(rows, special=True) if normstr(r['OTD ajustado'])=='Atrasado']),
}
week_checks=compare_week_overview(wb, rows, active_week)
region_checks=compare_region_errors(wb, wbf, rows)
volume_checks=compare_volume_sheets(wb, rows)
top_info=inspect_top_offenders(wb, rows)
charts=parse_charts()

failures={
    'week_overview':[x for x in week_checks if not x['ok']],
    'region_errors':[x for x in region_checks['checks'] if not x['ok']],
    'region_formula_issues':region_checks['formula_issues'],
    'volume_sheets':[x for x in volume_checks if not x['ok']],
    'formula_mojibake_alianca':xml_scan['formulas_with_mojibake_alianca'],
    'worksheet_error_cells':xml_scan['worksheet_error_cells'],
    'charts_with_ref_errors':[x for x in charts if x['has_error_ref']],
    'top_offenders_vendor_blank': top_info['Top_Offenders_Vendors'] if top_info['Top_Offenders_Vendors']['status']!='OK' else None,
}
summary={
    'workbook':str(WORKBOOK),
    'generated_at':datetime.now().isoformat(timespec='seconds'),
    'base_summary':base_summary,
    'counts':{
        'week_overview_checks':len(week_checks),'week_overview_failures':len(failures['week_overview']),
        'region_errors_checks':len(region_checks['checks']),'region_errors_failures':len(failures['region_errors']),
        'volume_sheet_checks':len(volume_checks),'volume_sheet_failures':len(failures['volume_sheets']),
        'formulas_with_mojibake_alianca':len(failures['formula_mojibake_alianca']),
        'worksheet_error_sheets':len(xml_scan['worksheet_error_cells']),
        'charts_total':len(charts),'charts_with_ref_errors':len(failures['charts_with_ref_errors']),
    },
    'xml_scan':xml_scan,
    'failures':failures,
    'top_offenders':top_info,
    'charts':charts,
    'week_overview_checks':week_checks,
    'region_errors':region_checks,
    'volume_sheets':volume_checks,
}
REPORT_JSON.write_text(json.dumps(summary, ensure_ascii=False, indent=2, default=str), encoding='utf-8')

md=[]
md.append(f"# Auditoria final WK19 - {TS}\n")
md.append(f"Workbook: `{WORKBOOK}`\n")
md.append(f"Semana ativa: **{active_week}**\n")
md.append("## Resumo da base ROE_wk\n")
for k,v in base_summary.items():
    if k!='status_patterns': md.append(f"- {k}: {v}\n")
md.append("\n## Status dos testes\n")
for k,v in summary['counts'].items(): md.append(f"- {k}: {v}\n")
md.append("\n## Falhas principais\n")
for k,v in failures.items():
    if v:
        if isinstance(v,list): md.append(f"- {k}: {len(v)} ocorr?ncia(s). Exemplo: `{v[0]}`\n")
        else: md.append(f"- {k}: `{v}`\n")
md.append("\n## Observa??o sobre erros XML\n")
md.append(f"- Worksheet error cells: {xml_scan['worksheet_error_cells']}\n")
md.append(f"- Defined names/pivot cache info-only: {xml_scan['defined_name_info_only']} / {xml_scan['pivot_cache_error_items']}\n")
REPORT_MD.write_text(''.join(md), encoding='utf-8')
wb.close(); wbf.close()
print(json.dumps({'report_json':str(REPORT_JSON),'report_md':str(REPORT_MD),'counts':summary['counts'],'base_summary':base_summary,'main_failures':{k:(len(v) if isinstance(v,list) else bool(v)) for k,v in failures.items()}}, ensure_ascii=False, indent=2, default=str))
