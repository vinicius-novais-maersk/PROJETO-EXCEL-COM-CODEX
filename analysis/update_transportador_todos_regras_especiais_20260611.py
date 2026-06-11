from __future__ import annotations

import json
import os
import shutil
import sys
import zipfile
from datetime import datetime
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

try:
    from scripts.workbook_paths import resolve_workbook_path
except ModuleNotFoundError:
    sys.path.append(str(Path(__file__).resolve().parents[1]))
    from scripts.workbook_paths import resolve_workbook_path

ROOT_DIR = Path(__file__).resolve().parents[1]
BACKUP_DIR = ROOT_DIR / 'backups'
ANALYSIS_DIR = ROOT_DIR / 'analysis'
SHEET_NAME = 'Regras_Especiais'
MAIN_NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
REL_NS = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
PKG_REL_NS = 'http://schemas.openxmlformats.org/package/2006/relationships'

ET.register_namespace('', MAIN_NS)
ET.register_namespace('r', REL_NS)
ET.register_namespace('', PKG_REL_NS)


def parse_xml(data: bytes) -> ET.Element:
    return ET.fromstring(data)


def find_sheet_part(zf: zipfile.ZipFile, sheet_name: str) -> str:
    workbook = parse_xml(zf.read('xl/workbook.xml'))
    rels = parse_xml(zf.read('xl/_rels/workbook.xml.rels'))
    rel_map = {
        rel.attrib['Id']: rel.attrib['Target']
        for rel in rels.findall(f'{{{PKG_REL_NS}}}Relationship')
    }
    sheets = workbook.find(f'{{{MAIN_NS}}}sheets')
    if sheets is None:
        raise RuntimeError('workbook.xml sem lista de abas')
    for sheet in sheets.findall(f'{{{MAIN_NS}}}sheet'):
        if sheet.attrib.get('name') == sheet_name:
            rid = sheet.attrib.get(f'{{{REL_NS}}}id')
            target = rel_map.get(rid)
            if not target:
                raise RuntimeError(f'Relacionamento não encontrado para {sheet_name}')
            if target.startswith('/'):
                return target.lstrip('/')
            return 'xl/' + target.lstrip('/')
    raise RuntimeError(f'Aba {sheet_name} não encontrada')


def patch_sheet_xml(data: bytes) -> tuple[bytes, dict]:
    root = parse_xml(data)
    changed_cells = 0
    changed_validation = 0

    # Inline string cells.
    for t in root.findall(f'.//{{{MAIN_NS}}}t'):
        if t.text == 'Diversas':
            t.text = 'Todos'
            changed_cells += 1
        elif t.text and 'Diversas,Maersk/Frota Própria,Específico' in t.text:
            t.text = t.text.replace('Diversas,Maersk/Frota Própria,Específico', 'Todos,Maersk/Frota Própria,Específico')
            changed_validation += 1

    # Data validation formula nodes.
    for formula in root.findall(f'.//{{{MAIN_NS}}}formula1'):
        if formula.text and 'Diversas,Maersk/Frota Própria,Específico' in formula.text:
            formula.text = formula.text.replace('Diversas,Maersk/Frota Própria,Específico', 'Todos,Maersk/Frota Própria,Específico')
            changed_validation += 1

    return ET.tostring(root, encoding='utf-8', xml_declaration=True), {
        'changed_cells': changed_cells,
        'changed_validation': changed_validation,
    }


def patch_workbook(input_path: Path, output_path: Path) -> dict:
    tmp_path = output_path.with_suffix(output_path.suffix + '.tmp')
    with zipfile.ZipFile(input_path, 'r') as zin:
        sheet_part = find_sheet_part(zin, SHEET_NAME)
        patch_info = None
        with zipfile.ZipFile(tmp_path, 'w', compression=zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == sheet_part:
                    data, patch_info = patch_sheet_xml(data)
                zout.writestr(item, data)
    if patch_info is None:
        raise RuntimeError('Aba localizada, mas XML não foi patchado')
    os.replace(tmp_path, output_path)
    patch_info['sheet_part'] = sheet_part
    return patch_info


def validate(path: Path) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True, keep_vba=True)
    try:
        ws = wb[SHEET_NAME]
        headers = [ws.cell(4, c).value for c in range(1, 14)]
        idx = {name: i + 1 for i, name in enumerate(headers)}
        tipo_col = idx['Transportador_Tipo']
        values = []
        divers_count = 0
        todos_count = 0
        rows = 0
        for r in range(5, ws.max_row + 1):
            if ws.cell(r, 1).value or ws.cell(r, 5).value:
                rows += 1
                value = ws.cell(r, tipo_col).value
                values.append(value)
                if value == 'Diversas':
                    divers_count += 1
                if value == 'Todos':
                    todos_count += 1
        return {
            'sheet_visible': ws.sheet_state == 'visible',
            'rows': rows,
            'transportador_diversas_count': divers_count,
            'transportador_todos_count': todos_count,
            'transportador_unique_values': sorted({str(v) for v in values if v not in (None, '')}),
        }
    finally:
        wb.close()


def zip_compare(before: Path, after: Path) -> dict:
    with zipfile.ZipFile(before) as zb, zipfile.ZipFile(after) as za:
        before_names = set(zb.namelist())
        after_names = set(za.namelist())
        return {
            'removed_entries': sorted(before_names - after_names),
            'added_entries': sorted(after_names - before_names),
            'backup_zip_test': zb.testzip(),
            'official_zip_test': za.testzip(),
        }


def main() -> int:
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    official_path = resolve_workbook_path()
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    ANALYSIS_DIR.mkdir(parents=True, exist_ok=True)

    backup_path = BACKUP_DIR / f'{official_path.stem}_before_transportador_todos_{timestamp}{official_path.suffix}'
    test_path = ANALYSIS_DIR / f'{official_path.stem}_transportador_todos_test_{timestamp}{official_path.suffix}'
    official_tmp = ANALYSIS_DIR / f'{official_path.stem}_transportador_todos_tmp_{timestamp}{official_path.suffix}'

    shutil.copy2(official_path, backup_path)
    patch_test = patch_workbook(official_path, test_path)
    test_validation = validate(test_path)
    if test_validation['transportador_diversas_count'] != 0:
        raise RuntimeError(f'Cópia de teste ainda tem Diversas: {test_validation}')

    patch_official = patch_workbook(official_path, official_tmp)
    official_validation = validate(official_tmp)
    if official_validation['transportador_diversas_count'] != 0:
        raise RuntimeError(f'Oficial temporária ainda tem Diversas: {official_validation}')

    os.replace(official_tmp, official_path)
    final_validation = validate(official_path)
    compare = zip_compare(backup_path, official_path)

    report = {
        'ok': True,
        'official_path': str(official_path),
        'backup_path': str(backup_path),
        'test_path': str(test_path),
        'patch_test': patch_test,
        'patch_official': patch_official,
        'final_validation': final_validation,
        'zip_compare': compare,
    }
    report_path = ANALYSIS_DIR / f'transportador_todos_regras_especiais_{timestamp}.json'
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding='utf-8')
    report['report_path'] = str(report_path)
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
