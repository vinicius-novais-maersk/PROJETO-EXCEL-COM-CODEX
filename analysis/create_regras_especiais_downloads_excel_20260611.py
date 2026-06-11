from __future__ import annotations

import importlib.util
import json
import shutil
import sys
import time
from datetime import datetime
from pathlib import Path

import pythoncom
import pywintypes
import win32com.client
from openpyxl import load_workbook

ROOT = Path(r'C:\Users\VNO024\OneDrive - Maersk Group\Aplicativos\Github - programas\PROJETO EXCEL COM CODEX')
SOURCE = Path(r'C:\Users\VNO024\Downloads\Base_DSU2026 - TbM - WK24_BACKUP_SEGURO_antes_regras_especiais_20260611.xlsm')
TIMESTAMP = datetime.now().strftime('%Y%m%d_%H%M%S')
OUTPUT = Path(r'C:\Users\VNO024\Downloads') / f'Base_DSU2026 - TbM - WK24_COM_REGRAS_ESPECIAIS_{TIMESTAMP}.xlsm'
BUILDER = ROOT / 'analysis' / 'create_regras_especiais_consolidada_20260611.py'
SHEET_NAME = 'Regras_Especiais'


def load_builder():
    spec = importlib.util.spec_from_file_location('rules_builder', BUILDER)
    if spec is None or spec.loader is None:
        raise RuntimeError(f'Não consegui carregar builder: {BUILDER}')
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def call_with_retry(func, *args, retries=20, delay=0.4, **kwargs):
    last_exc = None
    for _ in range(retries):
        try:
            return func(*args, **kwargs)
        except pywintypes.com_error as exc:
            last_exc = exc
            time.sleep(delay)
            pythoncom.PumpWaitingMessages()
    raise last_exc


def validate_with_openpyxl(path: Path, expected_rows: int) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True, keep_vba=True)
    try:
        if SHEET_NAME not in wb.sheetnames:
            raise RuntimeError(f'Aba {SHEET_NAME} não encontrada')
        ws = wb[SHEET_NAME]
        headers = [ws.cell(4, c).value for c in range(1, 14)]
        rows = 0
        values = []
        for r in range(5, ws.max_row + 1):
            if ws.cell(r, 1).value or ws.cell(r, 5).value:
                rows += 1
                values.append(ws.cell(r, 10).value)
        return {
            'sheet_count': len(wb.sheetnames),
            'sheet_visible': ws.sheet_state == 'visible',
            'rows': rows,
            'expected_rows': expected_rows,
            'headers': headers,
            'transportador_values': sorted({str(v) for v in values if v not in (None, '')}),
            'diversas_count': sum(1 for v in values if v == 'Diversas'),
        }
    finally:
        wb.close()


def main() -> int:
    if not SOURCE.exists():
        raise FileNotFoundError(SOURCE)
    if OUTPUT.exists():
        raise FileExistsError(OUTPUT)

    builder = load_builder()
    rules = builder.build_rules(SOURCE)
    headers = builder.HEADERS
    # Safety: ensure old label is not generated.
    for row in rules:
        if row.get('Transportador_Tipo') == 'Diversas':
            row['Transportador_Tipo'] = 'Todos'
    shutil.copy2(SOURCE, OUTPUT)

    pythoncom.CoInitialize()
    excel = win32com.client.DispatchEx('Excel.Application')
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.EnableEvents = False
    excel.ScreenUpdating = False
    excel.AskToUpdateLinks = False
    excel.AutomationSecurity = 3  # disable macros while editing

    workbook = None
    try:
        workbook = call_with_retry(
            excel.Workbooks.Open,
            str(OUTPUT),
            UpdateLinks=0,
            ReadOnly=False,
            IgnoreReadOnlyRecommended=True,
            AddToMru=False,
            CorruptLoad=0,
        )
        existing = [workbook.Worksheets(i).Name for i in range(1, workbook.Worksheets.Count + 1)]
        if SHEET_NAME in existing:
            raise RuntimeError(f'Aba {SHEET_NAME} já existe no arquivo de saída')

        ws = workbook.Worksheets.Add(After=workbook.Worksheets(workbook.Worksheets.Count))
        ws.Name = SHEET_NAME
        ws.Visible = True

        ws.Range('A1').Value = 'Regras consolidadas de clientes especiais'
        ws.Range('A1:M1').Merge()
        ws.Range('A1').Font.Bold = True
        ws.Range('A1').Font.Size = 14

        ws.Range('A2').Value = 'Fonte de controle. Não substitui a regra oficial automaticamente; serve para validação e manutenção.'
        ws.Range('A2:M2').Merge()
        ws.Range('A2').Font.Italic = True

        matrix = [headers] + [[row.get(header, '') for header in headers] for row in rules]
        start_row = 4
        end_row = start_row + len(matrix) - 1
        end_col = len(headers)
        ws.Range(ws.Cells(start_row, 1), ws.Cells(end_row, end_col)).Value = matrix

        header_range = ws.Range(ws.Cells(start_row, 1), ws.Cells(start_row, end_col))
        header_range.Font.Bold = True

        widths = [10, 18, 12, 34, 42, 42, 42, 16, 16, 24, 38, 34, 70]
        for idx, width in enumerate(widths, start=1):
            ws.Columns(idx).ColumnWidth = width
        ws.Columns(12).WrapText = True
        ws.Columns(13).WrapText = True

        # Validation lists. Do not use table object to reduce workbook structural risk.
        modal_range = ws.Range(ws.Cells(5, 9), ws.Cells(max(end_row, 500), 9))
        modal_range.Validation.Delete()
        modal_range.Validation.Add(Type=3, AlertStyle=1, Operator=1, Formula1='Todos,Longo Curso,Cabotagem')

        tr_range = ws.Range(ws.Cells(5, 10), ws.Cells(max(end_row, 500), 10))
        tr_range.Validation.Delete()
        tr_range.Validation.Add(Type=3, AlertStyle=1, Operator=1, Formula1='Todos,Maersk/Frota Própria,Específico')

        ws.Activate()
        ws.Range('A5').Select()
        excel.ActiveWindow.FreezePanes = True

        call_with_retry(workbook.Save)
        call_with_retry(workbook.Close, SaveChanges=False)
        workbook = None
    finally:
        if workbook is not None:
            try:
                call_with_retry(workbook.Close, SaveChanges=False)
            except Exception:
                pass
        excel.Quit()
        pythoncom.CoUninitialize()

    validation = validate_with_openpyxl(OUTPUT, len(rules))
    report = {
        'ok': True,
        'source': str(SOURCE),
        'output': str(OUTPUT),
        'rules_count': len(rules),
        'validation': validation,
    }
    report_path = ROOT / 'analysis' / f'downloads_regras_especiais_excel_{TIMESTAMP}.json'
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding='utf-8')
    report['report_path'] = str(report_path)
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
