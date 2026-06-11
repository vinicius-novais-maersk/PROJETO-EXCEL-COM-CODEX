from __future__ import annotations

import importlib.util
import json
import os
import re
import shutil
import sys
import zipfile
from datetime import datetime
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

try:
    from scripts.workbook_paths import resolve_workbook_path
except ModuleNotFoundError:
    sys.path.append(str(Path(__file__).resolve().parents[1]))
    from scripts.workbook_paths import resolve_workbook_path

ROOT_DIR = Path(__file__).resolve().parents[1]
ANALYSIS_DIR = ROOT_DIR / 'analysis'
BACKUP_DIR = ROOT_DIR / 'backups'
SHEET_NAME = 'Regras_Especiais'
BUILDER_PATH = ANALYSIS_DIR / 'create_regras_especiais_consolidada_20260611.py'

MAIN_NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
REL_NS = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
PKG_REL_NS = 'http://schemas.openxmlformats.org/package/2006/relationships'
CT_NS = 'http://schemas.openxmlformats.org/package/2006/content-types'
WORKSHEET_REL_TYPE = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet'
WORKSHEET_CONTENT_TYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml'

ET.register_namespace('', MAIN_NS)
ET.register_namespace('r', REL_NS)
ET.register_namespace('', PKG_REL_NS)
ET.register_namespace('', CT_NS)


def col_name(index: int) -> str:
    name = ''
    while index:
        index, rem = divmod(index - 1, 26)
        name = chr(65 + rem) + name
    return name


def load_builder():
    spec = importlib.util.spec_from_file_location('regras_builder', BUILDER_PATH)
    if spec is None or spec.loader is None:
        raise RuntimeError(f'Não consegui carregar o builder: {BUILDER_PATH}')
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def text_cell(ref: str, value: str) -> ET.Element:
    cell = ET.Element(f'{{{MAIN_NS}}}c', {'r': ref, 't': 'inlineStr'})
    is_el = ET.SubElement(cell, f'{{{MAIN_NS}}}is')
    t = ET.SubElement(is_el, f'{{{MAIN_NS}}}t')
    # Preserve leading/trailing spaces if any.
    if value != value.strip():
        t.set('{http://www.w3.org/XML/1998/namespace}space', 'preserve')
    t.text = value
    return cell


def create_row(row_index: int, values: list[str]) -> ET.Element:
    row = ET.Element(f'{{{MAIN_NS}}}row', {'r': str(row_index)})
    for col_idx, value in enumerate(values, start=1):
        if value is None:
            value = ''
        row.append(text_cell(f'{col_name(col_idx)}{row_index}', str(value)))
    return row


def create_sheet_xml(headers: list[str], rules: list[dict[str, str]]) -> bytes:
    worksheet = ET.Element(f'{{{MAIN_NS}}}worksheet')
    end_row = max(5, 4 + len(rules))
    ET.SubElement(worksheet, f'{{{MAIN_NS}}}dimension', {'ref': f'A1:M{end_row}'})

    sheet_views = ET.SubElement(worksheet, f'{{{MAIN_NS}}}sheetViews')
    sheet_view = ET.SubElement(sheet_views, f'{{{MAIN_NS}}}sheetView', {'workbookViewId': '0'})
    ET.SubElement(sheet_view, f'{{{MAIN_NS}}}pane', {
        'ySplit': '4',
        'topLeftCell': 'A5',
        'activePane': 'bottomLeft',
        'state': 'frozen',
    })
    ET.SubElement(sheet_view, f'{{{MAIN_NS}}}selection', {'pane': 'bottomLeft', 'activeCell': 'A5', 'sqref': 'A5'})

    ET.SubElement(worksheet, f'{{{MAIN_NS}}}sheetFormatPr', {'defaultRowHeight': '15'})
    cols = ET.SubElement(worksheet, f'{{{MAIN_NS}}}cols')
    widths = [10, 18, 12, 34, 42, 42, 42, 16, 16, 22, 38, 34, 70]
    for idx, width in enumerate(widths, start=1):
        ET.SubElement(cols, f'{{{MAIN_NS}}}col', {'min': str(idx), 'max': str(idx), 'width': str(width), 'customWidth': '1'})
    ET.SubElement(cols, f'{{{MAIN_NS}}}col', {'min': '15', 'max': '16', 'width': '24', 'customWidth': '1'})

    sheet_data = ET.SubElement(worksheet, f'{{{MAIN_NS}}}sheetData')
    sheet_data.append(create_row(1, ['Regras consolidadas de clientes especiais']))
    sheet_data.append(create_row(2, ['Fonte de controle: consolidado sem apagar as abas antigas. Modal permitido: Todos, Longo Curso, Cabotagem.']))
    sheet_data.append(create_row(4, headers))
    for row_idx, rule in enumerate(rules, start=5):
        sheet_data.append(create_row(row_idx, [rule.get(header, '') for header in headers]))

    merge_cells = ET.SubElement(worksheet, f'{{{MAIN_NS}}}mergeCells', {'count': '2'})
    ET.SubElement(merge_cells, f'{{{MAIN_NS}}}mergeCell', {'ref': 'A1:M1'})
    ET.SubElement(merge_cells, f'{{{MAIN_NS}}}mergeCell', {'ref': 'A2:M2'})

    ET.SubElement(worksheet, f'{{{MAIN_NS}}}autoFilter', {'ref': f'A4:M{end_row}'})

    data_validations = ET.SubElement(worksheet, f'{{{MAIN_NS}}}dataValidations', {'count': '2'})
    dv_modal = ET.SubElement(data_validations, f'{{{MAIN_NS}}}dataValidation', {
        'type': 'list', 'allowBlank': '1', 'showErrorMessage': '1', 'sqref': f'I5:I{max(end_row, 500)}'
    })
    ET.SubElement(dv_modal, f'{{{MAIN_NS}}}formula1').text = '"Todos,Longo Curso,Cabotagem"'
    dv_tr = ET.SubElement(data_validations, f'{{{MAIN_NS}}}dataValidation', {
        'type': 'list', 'allowBlank': '1', 'showErrorMessage': '1', 'sqref': f'J5:J{max(end_row, 500)}'
    })
    ET.SubElement(dv_tr, f'{{{MAIN_NS}}}formula1').text = '"Todos,Maersk/Frota Própria,Específico"'

    ET.SubElement(worksheet, f'{{{MAIN_NS}}}pageMargins', {
        'left': '0.7', 'right': '0.7', 'top': '0.75', 'bottom': '0.75', 'header': '0.3', 'footer': '0.3'
    })
    return ET.tostring(worksheet, encoding='utf-8', xml_declaration=True)


def next_sheet_part(existing_names: list[str]) -> tuple[str, int]:
    nums = []
    for name in existing_names:
        match = re.fullmatch(r'xl/worksheets/sheet(\d+)\.xml', name)
        if match:
            nums.append(int(match.group(1)))
    n = max(nums or [0]) + 1
    return f'xl/worksheets/sheet{n}.xml', n


def parse_xml(data: bytes) -> ET.ElementTree:
    return ET.ElementTree(ET.fromstring(data))


def add_sheet_to_xlsm(input_path: Path, output_path: Path, headers: list[str], rules: list[dict[str, str]]) -> None:
    with zipfile.ZipFile(input_path, 'r') as zin:
        names = zin.namelist()
        workbook_xml = parse_xml(zin.read('xl/workbook.xml'))
        wb_root = workbook_xml.getroot()
        sheets_el = wb_root.find(f'{{{MAIN_NS}}}sheets')
        if sheets_el is None:
            raise RuntimeError('workbook.xml sem elemento sheets')
        existing_sheet_names = [sheet.attrib.get('name') for sheet in sheets_el.findall(f'{{{MAIN_NS}}}sheet')]
        if SHEET_NAME in existing_sheet_names:
            raise RuntimeError(f'A aba {SHEET_NAME} já existe em {input_path}; abortado para não sobrescrever.')

        rels_xml = parse_xml(zin.read('xl/_rels/workbook.xml.rels'))
        rels_root = rels_xml.getroot()
        content_types_xml = parse_xml(zin.read('[Content_Types].xml'))
        ct_root = content_types_xml.getroot()

        sheet_part, _ = next_sheet_part(names)
        target = sheet_part.replace('xl/', '')

        # Next rId and sheetId.
        rid_nums = []
        for rel in rels_root.findall(f'{{{PKG_REL_NS}}}Relationship'):
            rid = rel.attrib.get('Id', '')
            if rid.startswith('rId') and rid[3:].isdigit():
                rid_nums.append(int(rid[3:]))
        new_rid = f'rId{max(rid_nums or [0]) + 1}'

        sheet_ids = []
        for sheet in sheets_el.findall(f'{{{MAIN_NS}}}sheet'):
            sid = sheet.attrib.get('sheetId')
            if sid and sid.isdigit():
                sheet_ids.append(int(sid))
        new_sheet_id = str(max(sheet_ids or [0]) + 1)

        ET.SubElement(sheets_el, f'{{{MAIN_NS}}}sheet', {
            'name': SHEET_NAME,
            'sheetId': new_sheet_id,
            f'{{{REL_NS}}}id': new_rid,
        })
        ET.SubElement(rels_root, f'{{{PKG_REL_NS}}}Relationship', {
            'Id': new_rid,
            'Type': WORKSHEET_REL_TYPE,
            'Target': target,
        })
        ET.SubElement(ct_root, f'{{{CT_NS}}}Override', {
            'PartName': '/' + sheet_part,
            'ContentType': WORKSHEET_CONTENT_TYPE,
        })

        sheet_xml = create_sheet_xml(headers, rules)
        tmp_path = output_path.with_suffix(output_path.suffix + '.tmp')
        with zipfile.ZipFile(tmp_path, 'w', compression=zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == 'xl/workbook.xml':
                    data = ET.tostring(wb_root, encoding='utf-8', xml_declaration=True)
                elif item.filename == 'xl/_rels/workbook.xml.rels':
                    data = ET.tostring(rels_root, encoding='utf-8', xml_declaration=True)
                elif item.filename == '[Content_Types].xml':
                    data = ET.tostring(ct_root, encoding='utf-8', xml_declaration=True)
                zout.writestr(item, data)
            zout.writestr(sheet_part, sheet_xml)
        os.replace(tmp_path, output_path)


def validate(path: Path, expected_rows: int, original_sheets: list[str] | None = None) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True, keep_vba=True)
    try:
        sheetnames = wb.sheetnames
        if SHEET_NAME not in sheetnames:
            raise RuntimeError(f'{SHEET_NAME} não encontrada em {path}')
        ws = wb[SHEET_NAME]
        headers = [ws.cell(4, c).value for c in range(1, 14)]
        rows = 0
        for r in range(5, ws.max_row + 1):
            if ws.cell(r, 1).value or ws.cell(r, 5).value:
                rows += 1
        added = [] if original_sheets is None else [s for s in sheetnames if s not in original_sheets]
        return {
            'path': str(path),
            'headers_ok': headers[0] == 'Ativo' and headers[8] == 'Modal' and headers[9] == 'Transportador_Tipo',
            'rows': rows,
            'expected_rows': expected_rows,
            'added_sheets': added,
            'sheet_count': len(sheetnames),
            'sample': [[ws.cell(r, c).value for c in range(1, 6)] for r in range(5, min(ws.max_row, 9) + 1)],
        }
    finally:
        wb.close()


def main() -> int:
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    official_path = resolve_workbook_path()
    if not official_path.exists():
        raise FileNotFoundError(official_path)

    builder = load_builder()
    rules = builder.build_rules(official_path)
    headers = builder.HEADERS

    original = load_workbook(official_path, read_only=True, data_only=True, keep_vba=True)
    original_sheets = original.sheetnames
    original.close()
    if SHEET_NAME in original_sheets:
        raise RuntimeError(f'A aba {SHEET_NAME} já existe na oficial; não vou sobrescrever.')

    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    ANALYSIS_DIR.mkdir(parents=True, exist_ok=True)
    backup_path = BACKUP_DIR / f'{official_path.stem}_before_regras_especiais_ooxml_{timestamp}{official_path.suffix}'
    test_base = ANALYSIS_DIR / f'{official_path.stem}_regras_especiais_ooxml_test_base_{timestamp}{official_path.suffix}'
    test_path = ANALYSIS_DIR / f'{official_path.stem}_regras_especiais_ooxml_test_{timestamp}{official_path.suffix}'

    shutil.copy2(official_path, backup_path)
    shutil.copy2(official_path, test_base)
    add_sheet_to_xlsm(test_base, test_path, headers, rules)
    test_validation = validate(test_path, len(rules), original_sheets)
    if test_validation['rows'] != len(rules) or test_validation['added_sheets'] != [SHEET_NAME]:
        raise RuntimeError(f'Falha na validação da cópia: {test_validation}')

    official_tmp = ANALYSIS_DIR / f'{official_path.stem}_official_with_regras_tmp_{timestamp}{official_path.suffix}'
    add_sheet_to_xlsm(official_path, official_tmp, headers, rules)
    official_validation = validate(official_tmp, len(rules), original_sheets)
    if official_validation['rows'] != len(rules) or official_validation['added_sheets'] != [SHEET_NAME]:
        raise RuntimeError(f'Falha na validação da oficial temporária: {official_validation}')

    os.replace(official_tmp, official_path)
    final_validation = validate(official_path, len(rules), original_sheets)

    report = {
        'ok': True,
        'official_path': str(official_path),
        'backup_path': str(backup_path),
        'test_path': str(test_path),
        'sheet_name': SHEET_NAME,
        'rules_count': len(rules),
        'test_validation': test_validation,
        'final_validation': final_validation,
    }
    report_path = ANALYSIS_DIR / f'regras_especiais_consolidadas_ooxml_{timestamp}.json'
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding='utf-8')
    report['report_path'] = str(report_path)
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
