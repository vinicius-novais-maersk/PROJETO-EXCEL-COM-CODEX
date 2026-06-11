from __future__ import annotations

import json
import re
import shutil
import sys
import time
import unicodedata
from collections import OrderedDict
from datetime import datetime
from pathlib import Path

import pythoncom
import pywintypes
import win32com.client
from openpyxl import load_workbook

try:
    from scripts.workbook_paths import resolve_workbook_path
except ModuleNotFoundError:
    sys.path.append(str(Path(__file__).resolve().parents[1]))
    from scripts.workbook_paths import resolve_workbook_path

ROOT_DIR = Path(__file__).resolve().parents[1]
BACKUP_DIR = ROOT_DIR / "backups"
ANALYSIS_DIR = ROOT_DIR / "analysis"
SHEET_NAME = "Regras_Especiais"

HEADERS = [
    "Ativo",
    "Status",
    "Execução",
    "Tipo_Regra",
    "Nome",
    "Embarcador",
    "Cliente_Proposta",
    "Porto",
    "Modal",
    "Transportador_Tipo",
    "Transportador_Detalhe",
    "Origem",
    "Observação",
]

MODAL_VALUES = ["Todos", "Longo Curso", "Cabotagem"]
TRANSPORTADOR_VALUES = ["Todos", "Maersk/Frota Própria", "Específico"]


def normalize(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = "".join(ch for ch in unicodedata.normalize("NFKD", text) if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", text).upper()


def is_active(value) -> bool:
    text = normalize(value)
    return text in {"S", "SIM", "YES", "TRUE"}


def clean(value, fallback: str = "") -> str:
    if value is None:
        return fallback
    text = str(value).strip()
    return text if text else fallback


def transport_type(provider_or_brand: str) -> str:
    text = normalize(provider_or_brand)
    if text.startswith("MAERSK") or text == "MAERSK":
        return "Maersk/Frota Própria"
    if text in {"", "TODAS", "TODOS", "DIVERSAS"}:
        return "Todos"
    return "Específico"


def modal_for_name(name: str, default: str = "Todos") -> str:
    text = normalize(name)
    if "CRISTAL MASTER" in text:
        return "Longo Curso"
    if "WESTROCK" in text:
        return "Cabotagem"
    return default


def make_key(row: dict[str, str]) -> tuple[str, ...]:
    return tuple(normalize(row.get(header, "")) for header in HEADERS[:11])


def add_rule(rules: OrderedDict, **kwargs) -> None:
    row = {header: "" for header in HEADERS}
    row.update({key: clean(value) for key, value in kwargs.items() if key in row})
    row.setdefault("Ativo", "Sim")
    if not row["Ativo"]:
        row["Ativo"] = "Sim"
    if not row["Status"]:
        row["Status"] = "Atual"
    if not row["Execução"]:
        row["Execução"] = "OTO"
    if not row["Porto"]:
        row["Porto"] = "Todos"
    if not row["Modal"]:
        row["Modal"] = "Todos"
    if not row["Transportador_Tipo"]:
        row["Transportador_Tipo"] = "Todos"
    key = make_key(row)
    if key in rules:
        existing = rules[key]
        for field in ["Origem", "Observação"]:
            new_value = row.get(field, "")
            if new_value and new_value not in existing[field]:
                existing[field] = f"{existing[field]}; {new_value}" if existing[field] else new_value
    else:
        rules[key] = row


def build_rules(workbook_path: Path) -> list[dict[str, str]]:
    rules: OrderedDict[tuple[str, ...], dict[str, str]] = OrderedDict()

    # 1) Regras hardcoded na fórmula atual de ROE_wk[Especiais], com ajustes confirmados pelo usuário.
    formula_rules = [
        ("Texto em Embarcador/Cliente", "COTRIGUACU COOPERATIVA CENTRAL", "Todos", "Todos", "Todos", "", "Fórmula atual"),
        ("Texto em Embarcador/Cliente", "PLUSVAL AGROAVICOLA LTDA", "Todos", "Todos", "Todos", "", "Fórmula atual"),
        ("Texto em Embarcador/Cliente", "CVALE COOPERATIVA AGROINDUSTRIAL", "Todos", "Todos", "Todos", "", "Fórmula atual"),
        ("Texto em Embarcador/Cliente", "COPACOL", "Todos", "Todos", "Todos", "", "Fórmula atual"),
        ("Texto em Embarcador/Cliente+Transportador", "MULTILIT FIBROCIMENTO LTDA", "Todos", "Todos", "Específico", "VALE DO TIBAGI TRANSPORTES E LOGISTICA L", "Fórmula atual"),
        ("Texto em Embarcador/Cliente+Transportador", "CRISTAL MASTER", "Todos", "Longo Curso", "Maersk/Frota Própria", "MAERSK*", "Fórmula atual + confirmação usuário"),
        ("Texto em Embarcador/Cliente+Transportador", "NIDEC GLOBAL APPLIANCE BRASIL LTDA", "Todos", "Todos", "Maersk/Frota Própria", "MAERSK*", "Fórmula atual"),
        ("Texto em Embarcador/Cliente+Transportador", "SUMITOMO RUBBER DO BRASIL LTDA", "Todos", "Todos", "Maersk/Frota Própria", "MAERSK*", "Fórmula atual"),
        ("Texto em Embarcador/Cliente+Transportador", "BMW DO BRASIL LTDA", "Todos", "Todos", "Maersk/Frota Própria", "MAERSK*", "Fórmula atual"),
        ("Texto em Embarcador/Cliente+Modal", "WESTROCK", "Todos", "Cabotagem", "Todos", "", "Fórmula atual"),
        ("Texto em Embarcador/Cliente+Porto", "MARIO JOSE WERNER & CIA LTDA", "Itajai", "Todos", "Todos", "", "Fórmula atual"),
        ("Cliente+Porto+Transportador", "PROCTER", "Rio", "Todos", "Específico", "IRB*", "Fórmula atual"),
        ("Cliente+Porto+Transportador", "VOLKSWAGEN TRUCK E BUS INDUSTRIA E COMER", "Rio", "Todos", "Específico", "IRB LOGISTICA S.A.", "Fórmula atual"),
        ("Cliente+Porto+Transportador", "AJINOMOTO DO BRASIL INDUSTRIA E COMERCIO", "Santos", "Todos", "Específico", "UNITRADING LOGISTICA IMPORTACAO E EXPORT", "Fórmula atual"),
        ("Cliente Proposta", "SAMSUNG SDS GLOBAL SCL LATIN AMERICA LOG", "Todos", "Todos", "Todos", "", "Fórmula atual"),
        ("Cliente Proposta", "SAMSUNG SDS LATIN AMERICA TECNOLOGIA  E", "Todos", "Todos", "Todos", "", "Fórmula atual"),
        ("Cliente+Embarcador", "NOVELIS DO BRASIL LTDA", "Todos", "Todos", "Todos", "BALL BEVERAGE CAN SOUTH AMERICA*", "Fórmula atual"),
        ("Cliente+Destinatário", "NOVELIS DO BRASIL LTDA", "Todos", "Todos", "Todos", "ADUKARGO TRANSPORTES LOGISTICA ESERVICOS", "Fórmula atual"),
        ("Cliente Proposta", "LG ELECTRONICS DO BRASIL LTDA", "Todos", "Todos", "Todos", "", "Fórmula atual"),
        ("Cliente Proposta", "LG DISTRIBUIDORA DE UTILIDADES LTDA", "Todos", "Todos", "Todos", "", "Fórmula atual"),
        ("Cliente Proposta", "ELGIN SA", "Todos", "Todos", "Todos", "", "Fórmula atual"),
        ("Cliente Proposta", "ELGIN INDUSTRIAL DA AMAZONIA LTDA", "Todos", "Todos", "Todos", "", "Fórmula atual"),
        ("Cliente Proposta", "VIDEOLAR SA", "Todos", "Todos", "Todos", "", "Fórmula atual"),
        ("Cliente Proposta", "VIDEOLAR INNOVA SA", "Todos", "Todos", "Todos", "", "Fórmula atual"),
        ("Cliente Proposta", "PHILCO ELETRONICOS SA", "Todos", "Todos", "Todos", "", "Fórmula atual"),
        ("Cliente Proposta", "VALGROUP AM INDUSTRIA DE MASTERBATCH LTD", "Todos", "Todos", "Todos", "", "Fórmula atual"),
        ("Cliente Proposta", "VALGROUP AM INDUSTRIA DE EMBALAGENS FLEX", "Todos", "Todos", "Todos", "", "Fórmula atual"),
        ("Cliente Proposta", "VALGROUP BRASIL III INDUSTRIA DE EMBALAG", "Todos", "Todos", "Todos", "", "Fórmula atual"),
        ("Cliente Proposta", "ELECTROLUX DO BRASIL SA", "Todos", "Todos", "Todos", "", "Fórmula atual"),
        ("Destinatário", "CONSULADO GERAL AMERICANO NO RIO DE JANE", "Todos", "Todos", "Todos", "", "Fórmula atual"),
        ("Cliente Proposta", "MAERSK A/S", "Todos", "Todos", "Todos", "", "Fórmula atual"),
    ]
    for type_rule, name, port, modal, tr_type, tr_detail, origin in formula_rules:
        obs = "Regra confirmada: somente Longo Curso + Frota" if "CRISTAL MASTER" in normalize(name) else "Vem da fórmula atual de ROE_wk[Especiais]"
        if normalize(name) == "LG ELECTRONICS DO BRASIL LTDA":
            obs = "Regra atual marca sem restringir transportador; lista regional cita Frota. Validar antes de alterar cálculo."
        add_rule(
            rules,
            Ativo="Sim",
            Status="Atual" if "CRISTAL MASTER" not in normalize(name) else "Ajustar fórmula",
            Tipo_Regra=type_rule,
            Nome=name,
            Porto=port,
            Modal=modal,
            Transportador_Tipo=tr_type,
            Transportador_Detalhe=tr_detail,
            Origem=origin,
            Observação=obs,
        )

    # 2) Lista regional informada pelo usuário.
    user_rules = [
        ("BMW DO BRASIL LTDA", "Todos", "Todos", "Maersk/Frota Própria", "MAERSK*"),
        ("COOPAVEL COOPERATIVA AGROINDUSTRIAL", "Todos", "Todos", "Todos", ""),
        ("COPACOL - COOPERATIVA AGROINDUSTRIAL", "Todos", "Todos", "Todos", ""),
        ("COTRIGUACU COOPERATIVA CENTRAL", "Todos", "Todos", "Todos", ""),
        ("CRISTAL MASTER INDUSTRIA E COMERCIO LTDA", "Todos", "Longo Curso", "Maersk/Frota Própria", "MAERSK*"),
        ("CVALE COOPERATIVA AGROINDUSTRIAL", "Todos", "Todos", "Todos", ""),
        ("MARIO JOSE WERNER & CIA LTDA", "Itajai", "Todos", "Todos", ""),
        ("MULTILIT FIBROCIMENTO LTDA", "Todos", "Todos", "Específico", "Vale do Tibagi"),
        ("NIDEC GLOBAL APPLIANCE BRASIL LTDA", "Todos", "Todos", "Maersk/Frota Própria", "MAERSK*"),
        ("PLUSVAL AGROAVICOLA LTDA", "Todos", "Todos", "Todos", ""),
        ("SUMITOMO RUBBER DO BRASIL LTDA", "Todos", "Todos", "Maersk/Frota Própria", "MAERSK*"),
        ("WESTROCK, CELULOSE, PAPEL E EMBALAGENS L", "Todos", "Cabotagem", "Todos", ""),
        ("LG ELECTRONICS DO BRASIL LTDA", "Todos", "Todos", "Maersk/Frota Própria", "MAERSK*"),
    ]
    for name, port, modal, tr_type, tr_detail in user_rules:
        status = "Lista regional"
        obs = "Lista recebida de usuário/região; usar como validação, não como regra global automática."
        if "CRISTAL MASTER" in normalize(name):
            status = "Confirmado"
            obs = "Confirmado pelo usuário: somente Longo Curso + Frota."
        add_rule(
            rules,
            Ativo="Sim",
            Status=status,
            Tipo_Regra="Lista regional usuário",
            Nome=name,
            Porto=port,
            Modal=modal,
            Transportador_Tipo=tr_type,
            Transportador_Detalhe=tr_detail,
            Origem="Lista regional usuário",
            Observação=obs,
        )

    wb = load_workbook(workbook_path, read_only=True, data_only=True, keep_vba=True)

    # 3) Exceptions: tabelas que a fórmula atual consulta ou que são OTO especiais.
    ws = wb["Exceptions"]
    for r in range(2, ws.max_row + 1):
        # Embarcador + Cliente + Provedor, cols S:W
        if is_active(ws.cell(r, 23).value):
            provider = clean(ws.cell(r, 21).value)
            add_rule(
                rules,
                Ativo="Sim",
                Tipo_Regra="Embarcador+Cliente+Transportador",
                Nome=clean(ws.cell(r, 20).value) or clean(ws.cell(r, 19).value),
                Embarcador=clean(ws.cell(r, 19).value),
                Cliente_Proposta=clean(ws.cell(r, 20).value),
                Porto="Todos",
                Modal="Todos",
                Transportador_Tipo=transport_type(provider),
                Transportador_Detalhe=provider,
                Origem="Exceptions!S:W",
                Observação="Regra OTO existente por embarcador + cliente + provedor.",
            )
        # Cliente Proposta Carrossel, cols Y:Z
        if is_active(ws.cell(r, 26).value):
            name = clean(ws.cell(r, 25).value)
            add_rule(
                rules,
                Ativo="Sim",
                Tipo_Regra="Cliente Proposta",
                Nome=name,
                Cliente_Proposta=name,
                Porto="Todos",
                Modal=modal_for_name(name),
                Transportador_Tipo="Todos",
                Origem="Exceptions!Y:Z",
                Observação="Regra OTO existente por cliente.",
            )
        # Cliente + Porto, cols AG:AJ
        if is_active(ws.cell(r, 36).value):
            name = clean(ws.cell(r, 33).value)
            add_rule(
                rules,
                Ativo="Sim",
                Tipo_Regra="Cliente+Porto",
                Nome=name,
                Cliente_Proposta=name,
                Porto=clean(ws.cell(r, 34).value, "Todos"),
                Modal=modal_for_name(name),
                Transportador_Tipo="Todos",
                Origem="Exceptions!AG:AJ",
                Observação="Regra OTO existente por cliente + porto.",
            )
        # Cliente + Marca, cols AL:AO
        if is_active(ws.cell(r, 41).value):
            name = clean(ws.cell(r, 38).value)
            brand = clean(ws.cell(r, 39).value)
            add_rule(
                rules,
                Ativo="Sim",
                Status="Ajustar fórmula" if "CRISTAL MASTER" in normalize(name) else "Atual",
                Tipo_Regra="Cliente+Marca/Transportador",
                Nome=name,
                Cliente_Proposta=name,
                Porto="Todos",
                Modal=modal_for_name(name),
                Transportador_Tipo=transport_type(brand),
                Transportador_Detalhe=brand,
                Origem="Exceptions!AL:AO",
                Observação="Regra OTO existente por cliente + marca/transportador.",
            )
        # Provedor OTO, cols M:N
        if is_active(ws.cell(r, 14).value):
            provider = clean(ws.cell(r, 13).value)
            add_rule(
                rules,
                Ativo="Sim",
                Tipo_Regra="Transportador",
                Nome=provider,
                Porto="Todos",
                Modal="Todos",
                Transportador_Tipo=transport_type(provider),
                Transportador_Detalhe=provider,
                Origem="Exceptions!M:N",
                Observação="Regra OTO existente por transportador.",
            )
        # Cliente Safra, cols P:Q
        if is_active(ws.cell(r, 17).value):
            name = clean(ws.cell(r, 16).value)
            add_rule(
                rules,
                Ativo="Sim",
                Tipo_Regra="Cliente Proposta Safra",
                Nome=name,
                Cliente_Proposta=name,
                Porto="Todos",
                Modal="Todos",
                Transportador_Tipo="Todos",
                Origem="Exceptions!P:Q",
                Observação="Regra OTO Safra existente.",
            )
        # Embarcador + Local, cols AB:AE
        if is_active(ws.cell(r, 31).value):
            shipper = clean(ws.cell(r, 28).value)
            local = clean(ws.cell(r, 29).value)
            add_rule(
                rules,
                Ativo="Sim",
                Tipo_Regra="Embarcador+Local",
                Nome=shipper,
                Embarcador=shipper,
                Porto="Todos",
                Modal="Todos",
                Transportador_Tipo="Todos",
                Transportador_Detalhe=local,
                Origem="Exceptions!AB:AE",
                Observação="Regra OTO existente por embarcador + local de atendimento.",
            )

    # 4) OTO Exceptions oculta: lista histórica/apoio validada pelo usuário.
    ws = wb["OTO Exceptions"]
    for r in range(2, ws.max_row + 1):
        if is_active(ws.cell(r, 5).value):
            name = clean(ws.cell(r, 4).value)
            add_rule(
                rules,
                Ativo="Sim",
                Tipo_Regra="Cliente Proposta",
                Nome=name,
                Cliente_Proposta=name,
                Porto="Todos",
                Modal=modal_for_name(name),
                Transportador_Tipo="Todos",
                Origem="OTO Exceptions!D:E",
                Observação="Registro na aba oculta OTO Exceptions.",
            )
        if is_active(ws.cell(r, 10).value):
            name = clean(ws.cell(r, 7).value)
            brand = clean(ws.cell(r, 8).value)
            add_rule(
                rules,
                Ativo="Sim",
                Status="Ajustar fórmula" if "CRISTAL MASTER" in normalize(name) else "Atual",
                Tipo_Regra="Cliente+Marca/Transportador",
                Nome=name,
                Cliente_Proposta=name,
                Porto="Todos",
                Modal=modal_for_name(name),
                Transportador_Tipo=transport_type(brand),
                Transportador_Detalhe=brand,
                Origem="OTO Exceptions!G:J",
                Observação="Registro na aba oculta OTO Exceptions.",
            )
        if is_active(ws.cell(r, 15).value):
            name = clean(ws.cell(r, 12).value)
            add_rule(
                rules,
                Ativo="Sim",
                Tipo_Regra="Cliente+Porto",
                Nome=name,
                Cliente_Proposta=name,
                Porto=clean(ws.cell(r, 13).value, "Todos"),
                Modal=modal_for_name(name),
                Transportador_Tipo="Todos",
                Origem="OTO Exceptions!L:O",
                Observação="Registro na aba oculta OTO Exceptions.",
            )
        if is_active(ws.cell(r, 21).value):
            provider = clean(ws.cell(r, 19).value)
            add_rule(
                rules,
                Ativo="Sim",
                Tipo_Regra="Embarcador+Cliente+Transportador",
                Nome=clean(ws.cell(r, 18).value) or clean(ws.cell(r, 17).value),
                Embarcador=clean(ws.cell(r, 17).value),
                Cliente_Proposta=clean(ws.cell(r, 18).value),
                Porto="Todos",
                Modal="Todos",
                Transportador_Tipo=transport_type(provider),
                Transportador_Detalhe=provider,
                Origem="OTO Exceptions!Q:U",
                Observação="Registro na aba oculta OTO Exceptions.",
            )
        if is_active(ws.cell(r, 26).value):
            shipper = clean(ws.cell(r, 23).value)
            local = clean(ws.cell(r, 24).value)
            add_rule(
                rules,
                Ativo="Sim",
                Tipo_Regra="Embarcador+Local",
                Nome=shipper,
                Embarcador=shipper,
                Porto="Todos",
                Modal="Todos",
                Transportador_Tipo="Todos",
                Transportador_Detalhe=local,
                Origem="OTO Exceptions!W:Z",
                Observação="Registro na aba oculta OTO Exceptions.",
            )
        if is_active(ws.cell(r, 29).value):
            provider = clean(ws.cell(r, 28).value)
            add_rule(
                rules,
                Ativo="Sim",
                Tipo_Regra="Transportador",
                Nome=provider,
                Porto="Todos",
                Modal="Todos",
                Transportador_Tipo=transport_type(provider),
                Transportador_Detalhe=provider,
                Origem="OTO Exceptions!AB:AC",
                Observação="Registro na aba oculta OTO Exceptions.",
            )
        if is_active(ws.cell(r, 32).value):
            name = clean(ws.cell(r, 31).value)
            add_rule(
                rules,
                Ativo="Sim",
                Tipo_Regra="Cliente Proposta Safra",
                Nome=name,
                Cliente_Proposta=name,
                Porto="Todos",
                Modal="Todos",
                Transportador_Tipo="Todos",
                Origem="OTO Exceptions!AE:AF",
                Observação="Registro na aba oculta OTO Exceptions.",
            )

    wb.close()

    rows = list(rules.values())
    rows.sort(key=lambda row: (normalize(row["Status"]), normalize(row["Tipo_Regra"]), normalize(row["Nome"]), normalize(row["Porto"]), normalize(row["Transportador_Detalhe"])))
    return rows


def call_with_retry(func, *args, retries: int = 30, delay: float = 0.4, **kwargs):
    last_exc = None
    for _ in range(retries):
        try:
            return func(*args, **kwargs)
        except pywintypes.com_error as exc:
            last_exc = exc
            time.sleep(delay)
            pythoncom.PumpWaitingMessages()
    raise last_exc


def write_rules_sheet(workbook_path: Path, rules: list[dict[str, str]]) -> None:
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.EnableEvents = False
    excel.AskToUpdateLinks = False
    workbook = None
    try:
        workbook = call_with_retry(excel.Workbooks.Open, str(workbook_path), UpdateLinks=False, ReadOnly=False)
        existing_names = [workbook.Worksheets(i).Name for i in range(1, workbook.Worksheets.Count + 1)]
        if SHEET_NAME in existing_names:
            raise RuntimeError(f"A aba {SHEET_NAME!r} já existe em {workbook_path}; abortado para não sobrescrever.")

        ws = workbook.Worksheets.Add(After=workbook.Worksheets(workbook.Worksheets.Count))
        ws.Name = SHEET_NAME
        ws.Visible = True

        title = "Regras consolidadas de clientes especiais"
        ws.Range("A1").Value = title
        ws.Range("A1:M1").Merge()
        ws.Range("A1").Font.Bold = True
        ws.Range("A1").Font.Size = 14
        ws.Range("A1").Interior.Color = 0x5D3617  # dark blue in BGR-ish Excel color
        ws.Range("A1").Font.Color = 0xFFFFFF

        note = "Fonte de controle: consolidado sem apagar as abas antigas. Modal permitido: Todos, Longo Curso, Cabotagem."
        ws.Range("A2").Value = note
        ws.Range("A2:M2").Merge()
        ws.Range("A2").Font.Italic = True

        start_row = 4
        matrix = [HEADERS] + [[row.get(header, "") for header in HEADERS] for row in rules]
        end_row = start_row + len(matrix) - 1
        end_col = len(HEADERS)
        target = ws.Range(ws.Cells(start_row, 1), ws.Cells(end_row, end_col))
        target.Value = matrix

        header_range = ws.Range(ws.Cells(start_row, 1), ws.Cells(start_row, end_col))
        header_range.Font.Bold = True
        header_range.Font.Color = 0xFFFFFF
        header_range.Interior.Color = 0x5D3617
        header_range.HorizontalAlignment = -4108

        data_range = ws.Range(ws.Cells(start_row, 1), ws.Cells(end_row, end_col))
        table = ws.ListObjects.Add(1, data_range, None, 1)
        table.Name = "tblRegrasEspeciais"
        table.TableStyle = "TableStyleMedium2"

        widths = [10, 18, 12, 34, 42, 42, 42, 16, 16, 22, 38, 34, 70]
        for idx, width in enumerate(widths, start=1):
            ws.Columns(idx).ColumnWidth = width
        ws.Columns(13).WrapText = True
        ws.Columns(12).WrapText = True
        ws.Columns(11).WrapText = True

        ws.Range("O1").Value = "Listas permitidas"
        ws.Range("O1").Font.Bold = True
        ws.Range("O2").Value = "Modal"
        for i, value in enumerate(MODAL_VALUES, start=3):
            ws.Cells(i, 15).Value = value
        ws.Range("P2").Value = "Transportador_Tipo"
        for i, value in enumerate(TRANSPORTADOR_VALUES, start=3):
            ws.Cells(i, 16).Value = value
        ws.Columns(15).ColumnWidth = 18
        ws.Columns(16).ColumnWidth = 24

        # Data validation for editable rows.
        modal_range = ws.Range(ws.Cells(start_row + 1, 9), ws.Cells(max(end_row, start_row + 200), 9))
        modal_range.Validation.Delete()
        modal_range.Validation.Add(Type=3, AlertStyle=1, Operator=1, Formula1=",".join(MODAL_VALUES))
        tr_range = ws.Range(ws.Cells(start_row + 1, 10), ws.Cells(max(end_row, start_row + 200), 10))
        tr_range.Validation.Delete()
        tr_range.Validation.Add(Type=3, AlertStyle=1, Operator=1, Formula1=",".join(TRANSPORTADOR_VALUES))

        ws.Activate()
        ws.Range("A5").Select()
        excel.ActiveWindow.FreezePanes = True

        call_with_retry(workbook.Save)
        call_with_retry(workbook.Close, SaveChanges=False)
        workbook = None
    finally:
        if workbook is not None:
            try:
                call_with_retry(workbook.Close, SaveChanges=False)
            except Exception:
                pass
        excel.Quit()


def validate_workbook(workbook_path: Path, expected_rows: int, original_sheets: list[str] | None = None) -> dict:
    wb = load_workbook(workbook_path, read_only=True, data_only=True, keep_vba=True)
    try:
        sheetnames = wb.sheetnames
        if SHEET_NAME not in sheetnames:
            raise RuntimeError(f"Aba {SHEET_NAME} não foi encontrada em {workbook_path}")
        ws = wb[SHEET_NAME]
        headers = [ws.cell(4, c).value for c in range(1, len(HEADERS) + 1)]
        if headers != HEADERS:
            raise RuntimeError(f"Cabeçalhos inesperados: {headers}")
        data_rows = 0
        for r in range(5, ws.max_row + 1):
            if ws.cell(r, 1).value in (None, "") and ws.cell(r, 5).value in (None, ""):
                continue
            data_rows += 1
        if data_rows != expected_rows:
            raise RuntimeError(f"Quantidade de linhas divergente: {data_rows} != {expected_rows}")
        added_sheets = []
        if original_sheets is not None:
            added_sheets = [sheet for sheet in sheetnames if sheet not in original_sheets]
        return {
            "path": str(workbook_path),
            "sheet_found": True,
            "row_count": data_rows,
            "sheet_count": len(sheetnames),
            "added_sheets": added_sheets,
            "sample_rows": [
                [ws.cell(r, c).value for c in range(1, 8)]
                for r in range(5, min(ws.max_row, 9) + 1)
            ],
        }
    finally:
        wb.close()


def main() -> int:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    official_path = resolve_workbook_path()
    if not official_path.exists():
        print(json.dumps({"ok": False, "error": f"Workbook not found: {official_path}"}, ensure_ascii=False))
        return 1

    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    ANALYSIS_DIR.mkdir(parents=True, exist_ok=True)

    backup_path = BACKUP_DIR / f"{official_path.stem}_before_regras_especiais_{timestamp}{official_path.suffix}"
    test_path = ANALYSIS_DIR / f"{official_path.stem}_regras_especiais_test_{timestamp}{official_path.suffix}"

    original_wb = load_workbook(official_path, read_only=True, data_only=True, keep_vba=True)
    original_sheets = original_wb.sheetnames
    original_wb.close()
    if SHEET_NAME in original_sheets:
        print(json.dumps({"ok": False, "error": f"A aba {SHEET_NAME} já existe na oficial; não sobrescrevi."}, ensure_ascii=False))
        return 1

    rules = build_rules(official_path)
    shutil.copy2(official_path, backup_path)
    shutil.copy2(official_path, test_path)

    write_rules_sheet(test_path, rules)
    test_validation = validate_workbook(test_path, len(rules), original_sheets)
    if test_validation.get("added_sheets") != [SHEET_NAME]:
        raise RuntimeError(f"Validação da cópia indicou abas inesperadas: {test_validation.get('added_sheets')}")

    write_rules_sheet(official_path, rules)
    official_validation = validate_workbook(official_path, len(rules), original_sheets)

    report = {
        "ok": True,
        "official_path": str(official_path),
        "backup_path": str(backup_path),
        "test_path": str(test_path),
        "sheet_name": SHEET_NAME,
        "rules_count": len(rules),
        "test_validation": test_validation,
        "official_validation": official_validation,
    }
    report_path = ANALYSIS_DIR / f"regras_especiais_consolidadas_{timestamp}.json"
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    report["report_path"] = str(report_path)
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    print(
        "Este script mantém apenas a função build_rules(). "
        "Para aplicar a aba na planilha, use "
        "analysis/create_regras_especiais_consolidada_ooxml_20260611.py."
    )
    raise SystemExit(1)
