from __future__ import annotations

import json
import re
import shutil
import sys
import time
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import pythoncom
import pywintypes
import win32com.client
from openpyxl import load_workbook

SOURCE = Path(r'C:\Users\VNO024\Downloads\Base_DSU2026 - TbM - WK24_COM_REGRAS_ESPECIAIS_20260611_120404.xlsm')
TIMESTAMP = datetime.now().strftime('%Y%m%d_%H%M%S')
OUTPUT = Path(r'C:\Users\VNO024\Downloads') / f'Base_DSU2026 - TbM - WK24_REGRAS_ESPECIAIS_AJUSTADO_{TIMESTAMP}.xlsm'
ROOT = Path(r'C:\Users\VNO024\OneDrive - Maersk Group\Aplicativos\Github - programas\PROJETO EXCEL COM CODEX')
REPORT = ROOT / 'analysis' / f'regras_especiais_simplificada_downloads_{TIMESTAMP}.json'
SHEET_NAME = 'Regras_Especiais'

NEW_HEADERS = [
    'Ativo',
    'Status',
    'Execução',
    'Tipo_Regra',
    'Cliente_Proposta',
    'Região',
    'Porto',
    'Modal',
    'Transportador_Detalhe',
    'Origem',
    'Observação',
]


def norm(value) -> str:
    if value is None:
        return ''
    text = str(value).strip()
    text = ''.join(ch for ch in unicodedata.normalize('NFKD', text) if not unicodedata.combining(ch))
    return re.sub(r'\s+', ' ', text).upper()


def clean(value, fallback='') -> str:
    if value is None:
        return fallback
    text = str(value).strip()
    return text if text else fallback


def split_candidate_names(*values):
    names = []
    for value in values:
        text = clean(value)
        if not text:
            continue
        names.append(text)
    # preserve order, unique normalized
    seen = set()
    out = []
    for name in names:
        key = norm(name)
        if key and key not in seen:
            seen.add(key)
            out.append(name)
    return out


def build_region_maps(wb):
    ws = wb['ROE_wk']
    rows = ws.iter_rows(values_only=True)
    headers = next(rows)
    idx = {norm(h).lower(): i for i, h in enumerate(headers)}
    needed = ['cliente proposta', 'embarcador', 'region', 'porto', 'provedor', 'especiais']
    missing = [name for name in needed if name not in idx]
    if missing:
        raise RuntimeError(f'Colunas ausentes em ROE_wk: {missing}')

    by_name = defaultdict(set)
    by_name_port = defaultdict(set)
    by_name_provider = defaultdict(set)

    for row in rows:
        if row[idx['especiais']] != 'Especial':
            continue
        region = clean(row[idx['region']], 'Sem Region')
        port = clean(row[idx['porto']], 'Sem Porto')
        provider = clean(row[idx['provedor']], 'Sem Provedor')
        for name in split_candidate_names(row[idx['cliente proposta']], row[idx['embarcador']]):
            key = norm(name)
            by_name[key].add(region)
            by_name_port[(key, norm(port))].add(region)
            by_name_provider[key].add(provider)
    return by_name, by_name_port, by_name_provider


def infer_region(candidates, port, by_name, by_name_port):
    regions = set()
    port_norm = norm(port)
    for name in candidates:
        key = norm(name)
        if port_norm and port_norm not in {'TODOS', 'TODAS'}:
            regions.update(by_name_port.get((key, port_norm), set()))
        if not regions:
            regions.update(by_name.get(key, set()))
    if not regions:
        return 'Todos'
    return '; '.join(sorted(regions))


def infer_transportador(detail, tipo):
    detail = clean(detail)
    tipo = clean(tipo)
    if detail:
        return detail
    if tipo:
        if norm(tipo) in {'TODOS', 'DIVERSAS'}:
            return 'Todos'
        if 'MAERSK' in norm(tipo) or 'FROTA' in norm(tipo):
            return 'MAERSK*'
        return tipo
    return 'Todos'


def read_and_transform(source: Path):
    wb = load_workbook(source, read_only=True, data_only=True, keep_vba=True)
    try:
        if SHEET_NAME not in wb.sheetnames:
            raise RuntimeError(f'Aba {SHEET_NAME} não encontrada em {source}')
        by_name, by_name_port, _ = build_region_maps(wb)
        ws = wb[SHEET_NAME]
        header_row = None
        old_headers = []
        for candidate_row in range(1, min(ws.max_row, 10) + 1):
            candidate_headers = [ws.cell(candidate_row, c).value for c in range(1, ws.max_column + 1)]
            normalized = {str(h) for h in candidate_headers if h not in (None, '')}
            if {'Ativo', 'Status', 'Cliente_Proposta'}.issubset(normalized):
                header_row = candidate_row
                old_headers = candidate_headers
                break
        if header_row is None:
            raise RuntimeError(f'Não encontrei a linha de cabeçalhos em {SHEET_NAME}')
        old_idx = {str(h): i + 1 for i, h in enumerate(old_headers) if h not in (None, '')}
        required = ['Ativo', 'Status', 'Execução', 'Tipo_Regra', 'Cliente_Proposta', 'Porto', 'Modal', 'Transportador_Detalhe', 'Origem', 'Observação']
        missing = [h for h in required if h not in old_idx]
        if missing:
            raise RuntimeError(f'Colunas ausentes em {SHEET_NAME}: {missing}')

        out = []
        seen = set()
        for r in range(header_row + 1, ws.max_row + 1):
            if not ws.cell(r, old_idx['Ativo']).value and not ws.cell(r, old_idx.get('Cliente_Proposta', 1)).value:
                continue
            cliente = clean(ws.cell(r, old_idx['Cliente_Proposta']).value)
            nome = clean(ws.cell(r, old_idx['Nome']).value) if 'Nome' in old_idx else ''
            embarcador = clean(ws.cell(r, old_idx['Embarcador']).value) if 'Embarcador' in old_idx else ''
            candidates = split_candidate_names(cliente, nome, embarcador)
            cliente_final = cliente or nome or embarcador or 'Sem Cliente'
            porto = clean(ws.cell(r, old_idx['Porto']).value, 'Todos')
            regiao = infer_region(candidates or [cliente_final], porto, by_name, by_name_port)
            transportador = infer_transportador(
                ws.cell(r, old_idx['Transportador_Detalhe']).value,
                ws.cell(r, old_idx['Transportador_Tipo']).value if 'Transportador_Tipo' in old_idx else '',
            )
            row = {
                'Ativo': clean(ws.cell(r, old_idx['Ativo']).value, 'Sim'),
                'Status': clean(ws.cell(r, old_idx['Status']).value),
                'Execução': clean(ws.cell(r, old_idx['Execução']).value, 'OTO'),
                'Tipo_Regra': clean(ws.cell(r, old_idx['Tipo_Regra']).value),
                'Cliente_Proposta': cliente_final,
                'Região': regiao,
                'Porto': porto,
                'Modal': clean(ws.cell(r, old_idx['Modal']).value, 'Todos'),
                'Transportador_Detalhe': transportador,
                'Origem': clean(ws.cell(r, old_idx['Origem']).value),
                'Observação': clean(ws.cell(r, old_idx['Observação']).value),
            }
            key = tuple(norm(row[h]) for h in NEW_HEADERS)
            if key not in seen:
                seen.add(key)
                out.append(row)
        return out, old_headers
    finally:
        wb.close()


def call_with_retry(func, *args, retries=20, delay=0.4, **kwargs):
    last_exc = None
    for _ in range(retries):
        try:
            return func(*args, **kwargs)
        except pywintypes.com_error as exc:
            last_exc = exc
            time.sleep(delay)
            pythoncom.PumpWaitingMessages()
    raise last_exc


def write_with_excel(source: Path, output: Path, rows):
    shutil.copy2(source, output)
    pythoncom.CoInitialize()
    excel = win32com.client.DispatchEx('Excel.Application')
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.EnableEvents = False
    excel.ScreenUpdating = False
    excel.AskToUpdateLinks = False
    excel.AutomationSecurity = 3
    workbook = None
    try:
        workbook = call_with_retry(
            excel.Workbooks.Open,
            str(output),
            UpdateLinks=0,
            ReadOnly=False,
            IgnoreReadOnlyRecommended=True,
            AddToMru=False,
            CorruptLoad=0,
        )
        # Remove old consolidated sheet from this output copy only.
        for i in range(workbook.Worksheets.Count, 0, -1):
            if workbook.Worksheets(i).Name == SHEET_NAME:
                workbook.Worksheets(i).Delete()
                break

        ws = workbook.Worksheets.Add(After=workbook.Worksheets(workbook.Worksheets.Count))
        ws.Name = SHEET_NAME
        ws.Visible = True
        matrix = [NEW_HEADERS] + [[row[h] for h in NEW_HEADERS] for row in rows]
        start_row = 1
        end_row = start_row + len(matrix) - 1
        end_col = len(NEW_HEADERS)
        ws.Range(ws.Cells(start_row, 1), ws.Cells(end_row, end_col)).Value = matrix
        ws.Range(ws.Cells(start_row, 1), ws.Cells(start_row, end_col)).Font.Bold = True

        widths = [10, 18, 12, 34, 46, 22, 18, 16, 38, 34, 70]
        for idx, width in enumerate(widths, start=1):
            ws.Columns(idx).ColumnWidth = width
        ws.Columns(10).WrapText = True
        ws.Columns(11).WrapText = True

        # Keep only useful validation. Transportador_Detalhe is free text/name of provider.
        modal_range = ws.Range(ws.Cells(2, 8), ws.Cells(max(end_row, 500), 8))
        modal_range.Validation.Delete()
        modal_range.Validation.Add(Type=3, AlertStyle=1, Operator=1, Formula1='Todos,Longo Curso,Cabotagem')

        ws.Activate()
        ws.Range('A2').Select()
        excel.ActiveWindow.FreezePanes = True
        call_with_retry(workbook.Save)
        call_with_retry(workbook.Close, SaveChanges=False)
        workbook = None
    finally:
        if workbook is not None:
            try:
                call_with_retry(workbook.Close, SaveChanges=False)
            except Exception:
                pass
        excel.Quit()
        pythoncom.CoUninitialize()


def validate(path: Path, expected_rows: int):
    wb = load_workbook(path, read_only=True, data_only=True, keep_vba=True)
    try:
        ws = wb[SHEET_NAME]
        headers = [ws.cell(1, c).value for c in range(1, len(NEW_HEADERS) + 1)]
        rows = 0
        has_removed_headers = False
        for h in ['Nome', 'Embarcador', 'Transportador_Tipo']:
            if h in headers:
                has_removed_headers = True
        regioes = set()
        transportadores = set()
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 1).value or ws.cell(r, 5).value:
                rows += 1
                regioes.add(str(ws.cell(r, 6).value or ''))
                transportadores.add(str(ws.cell(r, 9).value or ''))
        return {
            'sheet_count': len(wb.sheetnames),
            'sheet_visible': ws.sheet_state == 'visible',
            'headers': headers,
            'headers_ok': headers == NEW_HEADERS,
            'removed_headers_present': has_removed_headers,
            'rows': rows,
            'expected_rows': expected_rows,
            'sample_regioes': sorted(regioes)[:10],
            'transportador_blank_count': sum(1 for r in range(2, ws.max_row + 1) if (ws.cell(r, 1).value or ws.cell(r, 5).value) and not ws.cell(r, 9).value),
            'sample_transportadores': sorted(t for t in transportadores if t)[:10],
        }
    finally:
        wb.close()


def main():
    if not SOURCE.exists():
        raise FileNotFoundError(SOURCE)
    rows, old_headers = read_and_transform(SOURCE)
    write_with_excel(SOURCE, OUTPUT, rows)
    validation = validate(OUTPUT, len(rows))
    report = {
        'ok': True,
        'source': str(SOURCE),
        'output': str(OUTPUT),
        'old_headers': old_headers,
        'new_headers': NEW_HEADERS,
        'rows': len(rows),
        'validation': validation,
    }
    REPORT.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding='utf-8')
    report['report_path'] = str(REPORT)
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == '__main__':
    main()
