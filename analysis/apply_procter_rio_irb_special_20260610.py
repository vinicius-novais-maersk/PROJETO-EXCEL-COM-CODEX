from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import tempfile
import xml.etree.ElementTree as ET
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import load_workbook


NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

INSERT_BEFORE = (
    '        AND(_xlpm.cliente="VOLKSWAGEN TRUCK E BUS INDUSTRIA E COMER",'
    '_xlpm.provedor="IRB LOGISTICA S.A.",_xlpm.porto="Rio"),'
)

RULE = (
    '        AND(NOT(ISERROR(FIND("PROCTER",_xlpm.cliente))),'
    'LEFT(_xlpm.provedor,3)="IRB",_xlpm.porto="Rio"),\r\n'
)

TARGET_COLUMNS = {
    "cliente": 13,
    "provedor": 4,
    "porto": 51,
    "booking": 8,
    "otd_ajustado_BB": 54,
    "atrasado_BD": 56,
    "oto_out_BL": 64,
    "especiais_BO": 67,
    "atraso_rev_BP": 68,
    "is_oto_exception_CA": 79,
    "is_move_exception_CB": 80,
}


def _worksheet_path(workbook_path: Path, sheet_name: str) -> str:
    with ZipFile(workbook_path) as archive:
        workbook_xml = ET.fromstring(archive.read("xl/workbook.xml"))
        rels_xml = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        relmap = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels_xml}

        sheets = workbook_xml.find(f"{{{NS_MAIN}}}sheets")
        if sheets is None:
            raise RuntimeError("Workbook has no sheets collection")

        for sheet in sheets:
            if sheet.attrib["name"] == sheet_name:
                rel_id = sheet.attrib[f"{{{NS_REL}}}id"]
                target = relmap[rel_id]
                return "xl/" + target.lstrip("/") if not target.startswith("xl/") else target

    raise RuntimeError(f"Sheet not found: {sheet_name}")


def _find_target_rows(workbook_path: Path) -> list[dict]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True, keep_links=False)
    try:
        sheet = workbook["ROE_wk"]
        rows: list[dict] = []
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            cliente = str(row[TARGET_COLUMNS["cliente"] - 1] or "")
            provedor = str(row[TARGET_COLUMNS["provedor"] - 1] or "")
            porto = str(row[TARGET_COLUMNS["porto"] - 1] or "")
            if "PROCTER" in cliente.upper() and provedor.upper().startswith("IRB") and porto == "Rio":
                rows.append(
                    {
                        "row": row_num,
                        "booking": row[TARGET_COLUMNS["booking"] - 1],
                        "cliente": cliente,
                        "provedor": provedor,
                        "porto": porto,
                        "otd_ajustado_BB": row[TARGET_COLUMNS["otd_ajustado_BB"] - 1],
                        "atrasado_BD": row[TARGET_COLUMNS["atrasado_BD"] - 1],
                        "oto_out_BL": row[TARGET_COLUMNS["oto_out_BL"] - 1],
                        "especiais_BO": row[TARGET_COLUMNS["especiais_BO"] - 1],
                        "atraso_rev_BP": row[TARGET_COLUMNS["atraso_rev_BP"] - 1],
                        "is_oto_exception_CA": row[TARGET_COLUMNS["is_oto_exception_CA"] - 1],
                        "is_move_exception_CB": row[TARGET_COLUMNS["is_move_exception_CB"] - 1],
                    }
                )
        return rows
    finally:
        workbook.close()


def _replace_cell_cached_value(sheet_xml: str, cell_ref: str, value: str) -> str:
    pattern = re.compile(rf'(<c\b[^>]*\br="{re.escape(cell_ref)}"[^>]*>)(.*?)(</c>)', re.DOTALL)
    match = pattern.search(sheet_xml)
    if not match:
        raise RuntimeError(f"Cell not found in XML: {cell_ref}")

    start, body, end = match.groups()
    if re.search(r"<v>.*?</v>|<v\s*/>", body, flags=re.DOTALL):
        body = re.sub(r"<v>.*?</v>|<v\s*/>", f"<v>{value}</v>", body, count=1, flags=re.DOTALL)
    else:
        body = body + f"<v>{value}</v>"

    return sheet_xml[: match.start()] + start + body + end + sheet_xml[match.end() :]


def _set_full_recalc_on_open(workbook_xml: str) -> str:
    calc_match = re.search(r"<calcPr\b[^>]*/>", workbook_xml)
    if not calc_match:
        return workbook_xml.replace("</workbook>", '<calcPr calcMode="auto" fullCalcOnLoad="1" forceFullCalc="1"/></workbook>')

    calc_tag = calc_match.group(0)
    for attr, value in {
        "calcMode": "auto",
        "fullCalcOnLoad": "1",
        "forceFullCalc": "1",
    }.items():
        if re.search(rf'\b{attr}="[^"]*"', calc_tag):
            calc_tag = re.sub(rf'\b{attr}="[^"]*"', f'{attr}="{value}"', calc_tag)
        else:
            calc_tag = calc_tag.replace("/>", f' {attr}="{value}"/>')

    return workbook_xml[: calc_match.start()] + calc_tag + workbook_xml[calc_match.end() :]


def apply_patch(workbook_path: Path, output_path: Path | None = None) -> dict:
    workbook_path = workbook_path.resolve()
    output_path = output_path.resolve() if output_path else workbook_path
    sheet_path = _worksheet_path(workbook_path, "ROE_wk")
    target_rows_before = _find_target_rows(workbook_path)
    if not target_rows_before:
        raise RuntimeError("No Procter + IRB + Rio rows found before patch")

    tmp_dir: Path | None = None
    tmp_file: Path | None = None
    try:
        with ZipFile(workbook_path, "r") as source:
            sheet_xml = source.read(sheet_path).decode("utf-8")
            workbook_xml = source.read("xl/workbook.xml").decode("utf-8")
            already_present = 'FIND("PROCTER",_xlpm.cliente)' in sheet_xml and 'LEFT(_xlpm.provedor,3)="IRB"' in sheet_xml
            insert_count = 0
            if not already_present:
                insert_count = sheet_xml.count(INSERT_BEFORE)
                if insert_count == 0:
                    raise RuntimeError("Could not find insertion point in ROE_wk[Especiais] XML")
                sheet_xml = sheet_xml.replace(INSERT_BEFORE, RULE + INSERT_BEFORE)

            for row in target_rows_before:
                row_num = row["row"]
                sheet_xml = _replace_cell_cached_value(sheet_xml, f"BO{row_num}", "Especial")
                sheet_xml = _replace_cell_cached_value(sheet_xml, f"BD{row_num}", "0")
                sheet_xml = _replace_cell_cached_value(sheet_xml, f"BP{row_num}", "0")

            workbook_xml = _set_full_recalc_on_open(workbook_xml)

            tmp_dir = Path(tempfile.mkdtemp(prefix="dsu_procter_patch_"))
            tmp_file = tmp_dir / output_path.name
            with ZipFile(tmp_file, "w", ZIP_DEFLATED) as target:
                for info in source.infolist():
                    data = source.read(info.filename)
                    if info.filename == sheet_path:
                        data = sheet_xml.encode("utf-8")
                    elif info.filename == "xl/workbook.xml":
                        data = workbook_xml.encode("utf-8")
                    target.writestr(info, data)

        if tmp_file is None:
            raise RuntimeError("Patch file was not created")
        if output_path == workbook_path:
            os.replace(tmp_file, workbook_path)
        else:
            shutil.copyfile(tmp_file, output_path)
    finally:
        if tmp_dir is not None:
            shutil.rmtree(tmp_dir, ignore_errors=True)

    target_rows_after = _find_target_rows(output_path)
    summary = {
        "target_count": len(target_rows_after),
        "special_count_BO": sum(1 for row in target_rows_after if row["especiais_BO"] == "Especial"),
        "late_flag_count_BD": sum(1 for row in target_rows_after if row["atrasado_BD"] == 1),
        "atraso_rev_count_BP": sum(1 for row in target_rows_after if row["atraso_rev_BP"] == 1),
        "otd_still_late_count_BB": sum(1 for row in target_rows_after if row["otd_ajustado_BB"] == "Atrasado"),
    }

    return {
        "source": str(workbook_path),
        "output": str(output_path),
        "sheet_xml": sheet_path,
        "changed_formula_xml": not already_present,
        "formula_insertions": insert_count,
        "summary": summary,
        "rows_before": target_rows_before,
        "rows_after": target_rows_after,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--output-workbook", type=Path)
    parser.add_argument("--output-json", type=Path, required=True)
    args = parser.parse_args()

    result = apply_patch(args.workbook, args.output_workbook)
    args.output_json.parent.mkdir(parents=True, exist_ok=True)
    args.output_json.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(result["summary"], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
