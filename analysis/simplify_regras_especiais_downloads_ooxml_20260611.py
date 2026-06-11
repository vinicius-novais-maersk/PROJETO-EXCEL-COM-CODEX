from __future__ import annotations

import json
import re
import shutil
import sys
import unicodedata
import zipfile
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

SOURCE = Path(r'C:\Users\VNO024\Downloads\Base_DSU2026 - TbM - WK24_COM_REGRAS_ESPECIAIS_20260611_120404.xlsm')
TIMESTAMP = datetime.now().strftime('%Y%m%d_%H%M%S')
OUTPUT = Path(r'C:\Users\VNO024\Downloads') / f'Base_DSU2026 - TbM - WK24_REGRAS_ESPECIAIS_AJUSTADO_{TIMESTAMP}.xlsm'
ROOT = Path(r'C:\Users\VNO024\OneDrive - Maersk Group\Aplicativos\Github - programas\PROJETO EXCEL COM CODEX')
REPORT = ROOT / 'analysis' / f'regras_especiais_simplificada_ooxml_downloads_{TIMESTAMP}.json'
SHEET_NAME = 'Regras_Especiais'

MAIN_NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
REL_NS = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
PKG_REL_NS = 'http://schemas.openxmlformats.org/package/2006/relationships'
ET.register_namespace('', MAIN_NS)
ET.register_namespace('r', REL_NS)
ET.register_namespace('', PKG_REL_NS)

NEW_HEADERS = [
    'Ativo', 'Status', 'Execução', 'Tipo_Regra', 'Cliente_Proposta',
    'Região', 'Porto', 'Modal', 'Transportador_Detalhe', 'Origem', 'Observação'
]


def norm(value) -> str:
    if value is None:
        return ''
    text = str(value).strip()
    text = ''.join(ch for ch in unicodedata.normalize('NFKD', text) if not unicodedata.combining(ch))
    return re.sub(r'\s+', ' ', text).upper()


def clean(value, fallback='') -> str:
    if value is None:
        return fallback
    text = str(value).strip()
    return text if text else fallback


def split_candidate_names(*values):
    seen = set()
    names = []
    for value in values:
        text = clean(value)
        key = norm(text)
        if key and key not in seen:
            seen.add(key)
            names.append(text)
    return names


def build_region_maps(wb):
    ws = wb['ROE_wk']
    rows = ws.iter_rows(values_only=True)
    headers = next(rows)
    idx = {norm(h).lower(): i for i, h in enumerate(headers)}
    by_name = defaultdict(set)
    by_name_port = defaultdict(set)
    for row in rows:
        if row[idx['especiais']] != 'Especial':
            continue
        region = clean(row[idx['region']], 'Sem Region')
        port = clean(row[idx['porto']], 'Sem Porto')
        for name in split_candidate_names(row[idx['cliente proposta']], row[idx['embarcador']]):
            key = norm(name)
            by_name[key].add(region)
            by_name_port[(key, norm(port))].add(region)
    return by_name, by_name_port


def infer_region(candidates, port, by_name, by_name_port):
    regions = set()
    port_norm = norm(port)
    for name in candidates:
        key = norm(name)
        if port_norm and port_norm not in {'TODOS', 'TODAS'}:
            regions.update(by_name_port.get((key, port_norm), set()))
        if not regions:
            regions.update(by_name.get(key, set()))
    return '; '.join(sorted(regions)) if regions else 'Todos'


def infer_transportador(detail, tipo):
    detail = clean(detail)
    tipo = clean(tipo)
    if detail:
        return detail
    tipo_norm = norm(tipo)
    if tipo_norm in {'', 'TODOS', 'DIVERSAS'}:
        return 'Todos'
    if 'MAERSK' in tipo_norm or 'FROTA' in tipo_norm:
        return 'MAERSK*'
    return tipo


def read_transform(source: Path):
    wb = load_workbook(source, read_only=True, data_only=True, keep_vba=True)
    try:
        by_name, by_name_port = build_region_maps(wb)
        ws = wb[SHEET_NAME]
        header_row = None
        old_headers = None
        for r in range(1, min(ws.max_row, 10) + 1):
            row_headers = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            names = {str(v) for v in row_headers if v not in (None, '')}
            if {'Ativo', 'Status', 'Cliente_Proposta'}.issubset(names):
                header_row = r
                old_headers = row_headers
                break
        if header_row is None or old_headers is None:
            raise RuntimeError('Não encontrei os cabeçalhos da aba Regras_Especiais')
        old_idx = {str(h): i + 1 for i, h in enumerate(old_headers) if h not in (None, '')}
        rows = []
        seen = set()
        for r in range(header_row + 1, ws.max_row + 1):
            if not ws.cell(r, old_idx['Ativo']).value and not ws.cell(r, old_idx['Cliente_Proposta']).value:
                continue
            cliente = clean(ws.cell(r, old_idx['Cliente_Proposta']).value)
            nome = clean(ws.cell(r, old_idx['Nome']).value) if 'Nome' in old_idx else ''
            embarcador = clean(ws.cell(r, old_idx['Embarcador']).value) if 'Embarcador' in old_idx else ''
            cliente_final = cliente or nome or embarcador or 'Sem Cliente'
            candidates = split_candidate_names(cliente, nome, embarcador, cliente_final)
            porto = clean(ws.cell(r, old_idx['Porto']).value, 'Todos')
            row = {
                'Ativo': clean(ws.cell(r, old_idx['Ativo']).value, 'Sim'),
                'Status': clean(ws.cell(r, old_idx['Status']).value),
                'Execução': clean(ws.cell(r, old_idx['Execução']).value, 'OTO'),
                'Tipo_Regra': clean(ws.cell(r, old_idx['Tipo_Regra']).value),
                'Cliente_Proposta': cliente_final,
                'Região': infer_region(candidates, porto, by_name, by_name_port),
                'Porto': porto,
                'Modal': clean(ws.cell(r, old_idx['Modal']).value, 'Todos'),
                'Transportador_Detalhe': infer_transportador(
                    ws.cell(r, old_idx['Transportador_Detalhe']).value,
                    ws.cell(r, old_idx['Transportador_Tipo']).value if 'Transportador_Tipo' in old_idx else '',
                ),
                'Origem': clean(ws.cell(r, old_idx['Origem']).value),
                'Observação': clean(ws.cell(r, old_idx['Observação']).value),
            }
            key = tuple(norm(row[h]) for h in NEW_HEADERS)
            if key not in seen:
                seen.add(key)
                rows.append(row)
        return rows, old_headers
    finally:
        wb.close()


def col_name(index: int) -> str:
    result = ''
    while index:
        index, rem = divmod(index - 1, 26)
        result = chr(65 + rem) + result
    return result


def text_cell(ref: str, value: str) -> ET.Element:
    c = ET.Element(f'{{{MAIN_NS}}}c', {'r': ref, 't': 'inlineStr'})
    is_el = ET.SubElement(c, f'{{{MAIN_NS}}}is')
    t = ET.SubElement(is_el, f'{{{MAIN_NS}}}t')
    text = '' if value is None else str(value)
    if text != text.strip():
        t.set('{http://www.w3.org/XML/1998/namespace}space', 'preserve')
    t.text = text
    return c


def make_row(r: int, values: list[str]) -> ET.Element:
    row = ET.Element(f'{{{MAIN_NS}}}row', {'r': str(r)})
    for c_idx, value in enumerate(values, start=1):
        row.append(text_cell(f'{col_name(c_idx)}{r}', value))
    return row


def make_sheet_xml(rows: list[dict[str, str]]) -> bytes:
    end_row = 1 + len(rows)
    end_col = len(NEW_HEADERS)
    root = ET.Element(f'{{{MAIN_NS}}}worksheet')
    ET.SubElement(root, f'{{{MAIN_NS}}}dimension', {'ref': f'A1:{col_name(end_col)}{end_row}'})
    views = ET.SubElement(root, f'{{{MAIN_NS}}}sheetViews')
    view = ET.SubElement(views, f'{{{MAIN_NS}}}sheetView', {'workbookViewId': '0'})
    ET.SubElement(view, f'{{{MAIN_NS}}}pane', {'ySplit': '1', 'topLeftCell': 'A2', 'activePane': 'bottomLeft', 'state': 'frozen'})
    ET.SubElement(root, f'{{{MAIN_NS}}}sheetFormatPr', {'defaultRowHeight': '15'})
    cols = ET.SubElement(root, f'{{{MAIN_NS}}}cols')
    widths = [10, 18, 12, 34, 46, 22, 18, 16, 38, 34, 70]
    for i, width in enumerate(widths, start=1):
        ET.SubElement(cols, f'{{{MAIN_NS}}}col', {'min': str(i), 'max': str(i), 'width': str(width), 'customWidth': '1'})
    data = ET.SubElement(root, f'{{{MAIN_NS}}}sheetData')
    data.append(make_row(1, NEW_HEADERS))
    for i, row in enumerate(rows, start=2):
        data.append(make_row(i, [row[h] for h in NEW_HEADERS]))
    validations = ET.SubElement(root, f'{{{MAIN_NS}}}dataValidations', {'count': '1'})
    dv = ET.SubElement(validations, f'{{{MAIN_NS}}}dataValidation', {
        'type': 'list', 'allowBlank': '1', 'showErrorMessage': '1', 'sqref': f'H2:H{max(end_row, 500)}'
    })
    ET.SubElement(dv, f'{{{MAIN_NS}}}formula1').text = '"Todos,Longo Curso,Cabotagem"'
    ET.SubElement(root, f'{{{MAIN_NS}}}pageMargins', {'left': '0.7', 'right': '0.7', 'top': '0.75', 'bottom': '0.75', 'header': '0.3', 'footer': '0.3'})
    return ET.tostring(root, encoding='utf-8', xml_declaration=True)


def find_sheet_part(zf: zipfile.ZipFile):
    workbook = ET.fromstring(zf.read('xl/workbook.xml'))
    rels = ET.fromstring(zf.read('xl/_rels/workbook.xml.rels'))
    rel_map = {rel.attrib['Id']: rel.attrib['Target'] for rel in rels.findall(f'{{{PKG_REL_NS}}}Relationship')}
    for sheet in workbook.find(f'{{{MAIN_NS}}}sheets').findall(f'{{{MAIN_NS}}}sheet'):
        if sheet.attrib.get('name') == SHEET_NAME:
            rid = sheet.attrib.get(f'{{{REL_NS}}}id')
            target = rel_map[rid]
            return target.lstrip('/') if target.startswith('/') else 'xl/' + target.lstrip('/')
    raise RuntimeError(f'Aba {SHEET_NAME} não encontrada no pacote')


def replace_sheet(source: Path, output: Path, rows: list[dict[str, str]]):
    sheet_xml = make_sheet_xml(rows)
    shutil.copy2(source, output)
    tmp = output.with_suffix(output.suffix + '.tmp')
    with zipfile.ZipFile(output, 'r') as zin:
        sheet_part = find_sheet_part(zin)
        with zipfile.ZipFile(tmp, 'w', compression=zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == sheet_part:
                    data = sheet_xml
                zout.writestr(item, data)
    tmp.replace(output)
    return sheet_part


def validate(path: Path, expected_rows: int):
    wb = load_workbook(path, read_only=True, data_only=True, keep_vba=True)
    try:
        ws = wb[SHEET_NAME]
        headers = [ws.cell(1, c).value for c in range(1, len(NEW_HEADERS) + 1)]
        rows = sum(1 for r in range(2, ws.max_row + 1) if ws.cell(r, 1).value or ws.cell(r, 5).value)
        return {
            'sheet_count': len(wb.sheetnames),
            'sheet_visible': ws.sheet_state == 'visible',
            'headers': headers,
            'headers_ok': headers == NEW_HEADERS,
            'rows': rows,
            'expected_rows': expected_rows,
            'removed_columns_present': any(h in headers for h in ['Nome', 'Embarcador', 'Transportador_Tipo']),
            'transportador_blank_count': sum(1 for r in range(2, ws.max_row + 1) if (ws.cell(r, 1).value or ws.cell(r, 5).value) and not ws.cell(r, 9).value),
            'sample': [[ws.cell(r, c).value for c in range(1, len(NEW_HEADERS)+1)] for r in range(1, min(ws.max_row, 6)+1)],
        }
    finally:
        wb.close()


def main():
    if not SOURCE.exists():
        raise FileNotFoundError(SOURCE)
    rows, old_headers = read_transform(SOURCE)
    sheet_part = replace_sheet(SOURCE, OUTPUT, rows)
    validation = validate(OUTPUT, len(rows))
    report = {
        'ok': True,
        'source': str(SOURCE),
        'output': str(OUTPUT),
        'sheet_part': sheet_part,
        'old_headers': old_headers,
        'new_headers': NEW_HEADERS,
        'rows': len(rows),
        'validation': validation,
    }
    REPORT.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding='utf-8')
    report['report_path'] = str(REPORT)
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == '__main__':
    main()
